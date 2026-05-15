import sys
from pathlib import Path
import openpyxl
sys.stdout.reconfigure(encoding="utf-8")

R = Path(r"C:\Users\DidWac\Prosero Dropbox\Didrik Wachtmeister\Phoenix Foundation\April alla filer\Get testfiles\extracted")
for period, fname in [("202603", "153_Balans 03-2026.xlsx"), ("202604", "153_Balance 04-2026.xlsx")]:
    p = R / period / "Finland" / "Referens" / fname
    if not p.exists():
        print(f"MISSING: {p}"); continue
    print(f"\n=== {period}: {fname} ===")
    wb = openpyxl.load_workbook(str(p), data_only=True)
    ws = wb.active
    for r in range(1, min(ws.max_row + 1, 12)):
        row = [ws.cell(row=r, column=c).value for c in range(1, min(ws.max_column + 1, 10))]
        print(f"  R{r}: {row}")
    wb.close()
