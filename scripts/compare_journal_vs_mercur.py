"""Konto-nivå-jämförelse: fact_journal_* (raw vouchers) vs backup_from_mercur.

Bakgrund: Mercur lagrar månadsrörelse per (bolag, period, konto), ofta uppdelad
i dim-rader (kostnadsställe, projekt). DB-källan beror på bolag:
  SE        → fact_journal_sie  (raw #VER aggregerat per period+konto, sign-flip)
  NO/DK     → fact_journal_saft (GeneralLedgerEntries aggregerat, sign-flip)
  FI/DK/DE  → fact_balances source_kind='IMP' (redan monthly)
För SE/NO/DK skiljer sig DB-tabellen från fact_balances SIE/SAFT — de senare är
YTD-snapshots och föråldras när FY-laddningen kör med --override.

Loop per bolag: undviker statement_timeout (Axlås har >1M journal-rader).

Jämför per (bolag, period, konto). YTD-summor + per-månads-diff. Bucketize.
Skriver xlsx-rapport till _uploads/ för djupare granskning.

Tecken: SIE/SAFT lagras med intäkt -, Mercur intäkt +. SIE/SAFT-sidan flippas
× -1 vid aggregering. IMP behöver ingen flip (Mercur-konvention redan).
"""
from __future__ import annotations

import os
import sys
from collections import defaultdict
from datetime import datetime

import psycopg
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_FILE = os.path.join(REPO, "_uploads", "Konto-nivå journal vs Mercur.xlsx")
PERIODS = ("202601", "202602", "202603", "202604")
TOLERANS = 1.0  # < 1 kr = ok


def get_db_monthly_se(con, cid: int) -> dict[tuple[str, str], float]:
    """SE: SUM från fact_journal_sie per (period, account), sign-flippat."""
    cur = con.execute(
        """SELECT period, account_code, SUM(-amount) AS amt
           FROM fact_journal_sie
           WHERE company_id = %s AND period BETWEEN %s AND %s
             AND LEFT(account_code, 1) IN ('3','4','5','6','7','8','9')
           GROUP BY period, account_code""",
        [cid, PERIODS[0], PERIODS[-1]],
    )
    return {(r[0], r[1]): float(r[2] or 0) for r in cur.fetchall()}


def get_db_monthly_saft(con, cid: int) -> dict[tuple[str, str], float]:
    """NO/DK: SUM från fact_journal_saft per (period, account), sign-flippat."""
    cur = con.execute(
        """SELECT period, account_code, SUM(-amount) AS amt
           FROM fact_journal_saft
           WHERE company_id = %s AND period BETWEEN %s AND %s
             AND LEFT(account_code, 1) IN ('3','4','5','6','7','8','9')
           GROUP BY period, account_code""",
        [cid, PERIODS[0], PERIODS[-1]],
    )
    return {(r[0], r[1]): float(r[2] or 0) for r in cur.fetchall()}


def get_db_monthly_imp(con, cid: int) -> dict[tuple[str, str], float]:
    """FI/DK/DE: fact_balances source_kind='IMP' per (period, account).
    IMP är redan monthly i Mercur-konvention — ingen flip."""
    cur = con.execute(
        """SELECT period, account_code, SUM(amount) AS amt
           FROM fact_balances
           WHERE company_id = %s AND source_kind = 'IMP' AND scenario = 'A'
             AND period BETWEEN %s AND %s
             AND LEFT(account_code, 1) IN ('3','4','5','6','7','8','9')
           GROUP BY period, account_code""",
        [cid, PERIODS[0], PERIODS[-1]],
    )
    return {(r[0], r[1]): float(r[2] or 0) for r in cur.fetchall()}


def get_mercur_monthly(con, cid: int) -> dict[tuple[str, str], float]:
    """SUM över dim-rader i backup_from_mercur per (period, account)."""
    cur = con.execute(
        """SELECT period, account_code, SUM(amount) AS amt
           FROM backup_from_mercur
           WHERE company_id = %s AND scenario = 'A'
             AND period BETWEEN %s AND %s
             AND LEFT(account_code, 1) IN ('3','4','5','6','7','8','9')
           GROUP BY period, account_code""",
        [cid, PERIODS[0], PERIODS[-1]],
    )
    return {(r[0], r[1]): float(r[2] or 0) for r in cur.fetchall()}


def db_loader_for_country(country: str):
    """Välj rätt journal-laddare per land. None = bolaget skippas helt."""
    if country == "Sweden":
        return get_db_monthly_se
    if country in ("Norway", "Denmark"):
        # Vissa DK-bolag är IMP (FI/DK/DE INL), andra SAFT (054 Prosero, 081 Actas).
        # Vi probar SAFT först, faller tillbaka på IMP om SAFT-källan är tom.
        return get_db_monthly_saft
    if country in ("Finland", "Germany", "CENTR", "CA"):
        return get_db_monthly_imp
    return None


def bucket(diff: float) -> str:
    a = abs(diff)
    if a < TOLERANS:
        return "ok"
    if a < 100:
        return "1-100"
    if a < 1000:
        return "100-1k"
    if a < 10000:
        return "1k-10k"
    return ">=10k"


def main() -> None:
    url = os.environ.get("DATABASE_URL_ETL") or os.environ.get("DATABASE_URL")
    if not url:
        sys.exit("DATABASE_URL_ETL saknas i env")

    print(f"[1/3] Hämtar bolagsregister + Mercur-omfattning ...")
    with psycopg.connect(url) as con:
        cur = con.execute(
            """SELECT DISTINCT m.company_id, dc.name, dc.country
               FROM backup_from_mercur m
               JOIN dim_company dc ON dc.company_id = m.company_id
               WHERE m.scenario = 'A' AND m.period BETWEEN %s AND %s
               ORDER BY m.company_id""",
            [PERIODS[0], PERIODS[-1]],
        )
        companies = cur.fetchall()
    print(f"   {len(companies)} bolag i scope")

    print(f"[2/3] Loopar per bolag (kan ta ~1-2 min för stora SE-bolag) ...")

    # Per-bolag-resultat
    rows: list[tuple] = []  # (cid, name, country, period, account, mercur, db, diff, db_source)
    company_stats: list[dict] = []

    for cid, name, country in companies:
        loader = db_loader_for_country(country)
        if loader is None:
            print(f"  SKIP {cid:>4} {name[:30]:<30}  ({country})")
            continue

        try:
            with psycopg.connect(url) as con:
                db_data = loader(con, cid)
                m_data = get_mercur_monthly(con, cid)
        except psycopg.errors.QueryCanceled:
            print(f"  TIMEOUT {cid:>4} {name[:30]:<30}  ({country}) — hoppar över")
            company_stats.append(dict(
                cid=cid, name=name, country=country, status="timeout",
                cells=0, ok=0, ge10k=0, abs_diff=0.0, top_diff=None,
            ))
            continue

        # Probe: om SAFT-loader returnerade tomt och bolaget är DK, prova IMP
        if not db_data and country == "Denmark":
            try:
                with psycopg.connect(url) as con:
                    db_data = get_db_monthly_imp(con, cid)
                db_source = "IMP"
            except psycopg.errors.QueryCanceled:
                pass
        else:
            db_source = {"Sweden": "JOURNAL_SIE", "Norway": "JOURNAL_SAFT",
                         "Denmark": "JOURNAL_SAFT", "Finland": "IMP",
                         "Germany": "IMP", "CENTR": "IMP", "CA": "IMP"}.get(country, "?")

        keys = set(db_data) | set(m_data)
        n_ok = n_lt100 = n_lt1k = n_lt10k = n_ge10k = 0
        abs_diff_sum = 0.0
        top_diff_key, top_diff_val = None, 0.0

        for k in keys:
            m_v = m_data.get(k, 0.0)
            d_v = db_data.get(k, 0.0)
            d = m_v - d_v
            b = bucket(d)
            if b == "ok": n_ok += 1
            elif b == "1-100": n_lt100 += 1
            elif b == "100-1k": n_lt1k += 1
            elif b == "1k-10k": n_lt10k += 1
            else: n_ge10k += 1
            abs_diff_sum += abs(d)
            if abs(d) > abs(top_diff_val):
                top_diff_key, top_diff_val = k, d
            if abs(d) >= TOLERANS:  # spara bara diff-rader i details
                rows.append((cid, name, country, k[0], k[1], m_v, d_v, d, db_source))

        total = len(keys)
        ok_pct = 100 * n_ok / total if total else 0
        company_stats.append(dict(
            cid=cid, name=name, country=country, status="ok",
            cells=total, ok=n_ok, ge10k=n_ge10k, abs_diff=abs_diff_sum,
            ok_pct=ok_pct, top_diff=top_diff_key, top_diff_val=top_diff_val,
            db_source=db_source,
        ))
        marker = "✓" if n_ge10k == 0 else ("." if n_ge10k < 5 else "!")
        print(f"  {marker} {cid:>4} {name[:30]:<30} {country:<10} "
              f"cells={total:>4} ok={n_ok:>4} ({ok_pct:>4.1f}%) ≥10k={n_ge10k:>3}  src={db_source}")

    # Per land
    print()
    print("[3/3] Sammanfattning per land ...")
    by_country: dict[str, dict[str, float]] = defaultdict(lambda: dict(cells=0, ok=0, ge10k=0, abs=0.0))
    for s in company_stats:
        if s["status"] != "ok":
            continue
        c = by_country[s["country"]]
        c["cells"] += s["cells"]
        c["ok"] += s["ok"]
        c["ge10k"] += s["ge10k"]
        c["abs"] += s["abs_diff"]
    print(f'{"Land":<12} {"cells":>8} {"ok":>8} {"ok%":>7} {">=10k":>7} {"abs(diff) sum":>16}')
    grand = dict(cells=0, ok=0, ge10k=0, abs=0.0)
    for country, s in sorted(by_country.items(), key=lambda x: -x[1]["cells"]):
        pct = 100 * s["ok"] / s["cells"] if s["cells"] else 0
        print(f'{country:<12} {int(s["cells"]):>8,} {int(s["ok"]):>8,} {pct:>6.1f}% '
              f'{int(s["ge10k"]):>7,} {s["abs"]:>16,.0f}')
        grand["cells"] += s["cells"]; grand["ok"] += s["ok"]
        grand["ge10k"] += s["ge10k"]; grand["abs"] += s["abs"]
    pct = 100 * grand["ok"] / grand["cells"] if grand["cells"] else 0
    print(f'{"TOTAL":<12} {int(grand["cells"]):>8,} {int(grand["ok"]):>8,} {pct:>6.1f}% '
          f'{int(grand["ge10k"]):>7,} {grand["abs"]:>16,.0f}')

    # XLSX-rapport
    print()
    print(f"Skriver {os.path.basename(OUT_FILE)} ...")
    wb = Workbook()
    bold = Font(bold=True)
    hl = PatternFill("solid", fgColor="FFE082")

    # Per-bolag-sammanfattning
    ws = wb.active
    ws.title = "Per bolag"
    headers = ["CID", "Namn", "Land", "DB-källa", "Status", "Celler", "OK",
               "OK%", ">=10k", "abs(diff) sum", "Topp diff konto", "Topp diff (period)", "Topp diff belopp"]
    ws.append(headers)
    for c in ws[1]:
        c.font = bold
    for s in sorted(company_stats, key=lambda x: -(x.get("abs_diff") or 0)):
        td = s.get("top_diff")
        ws.append([
            s["cid"], s["name"], s["country"], s.get("db_source", "?"), s["status"],
            s.get("cells", 0), s.get("ok", 0),
            round(s.get("ok_pct", 0), 1) if s["status"] == "ok" else None,
            s.get("ge10k", 0), round(s.get("abs_diff", 0), 2),
            td[1] if td else None, td[0] if td else None,
            round(s.get("top_diff_val", 0), 2) if td else None,
        ])

    # Per-land-sammanfattning
    ws2 = wb.create_sheet("Per land")
    ws2.append(["Land", "Celler", "OK", "OK%", ">=10k", "abs(diff) sum"])
    for c in ws2[1]:
        c.font = bold
    for country, s in sorted(by_country.items(), key=lambda x: -x[1]["cells"]):
        pct = 100 * s["ok"] / s["cells"] if s["cells"] else 0
        ws2.append([country, int(s["cells"]), int(s["ok"]), round(pct, 1),
                    int(s["ge10k"]), round(s["abs"], 0)])
    ws2.append(["TOTAL", int(grand["cells"]), int(grand["ok"]),
                round(100*grand["ok"]/grand["cells"], 1) if grand["cells"] else 0,
                int(grand["ge10k"]), round(grand["abs"], 0)])
    for c in ws2[ws2.max_row]:
        c.font = bold
        c.fill = hl

    # Detaljer: alla diff-rader
    ws3 = wb.create_sheet("Diff-rader")
    ws3.append(["CID", "Namn", "Land", "Period", "Konto", "Mercur", "DB", "Diff", "DB-källa"])
    for c in ws3[1]:
        c.font = bold
    for r in sorted(rows, key=lambda x: -abs(x[7]))[:5000]:  # max 5000 rader
        ws3.append(list(r))

    wb.save(OUT_FILE)
    print(f"Klar: {OUT_FILE}")


if __name__ == "__main__":
    main()
