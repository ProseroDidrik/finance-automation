import sys
from pathlib import Path
import openpyxl
sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(r"C:\Users\DidWac\Prosero Dropbox\Didrik Wachtmeister\Phoenix Foundation\April alla filer\Get testfiles\extracted\202604")
paths = [
    ROOT / "Finland" / "output" / "145_Prosero Security Oy_202604_INL.xlsx",
    ROOT / "Denmark" / "output" / "216_Sikom_202604_INL.xlsx",
]
for path in paths:
    print(f"\n=== {path.name} ===")
    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb.active
    accs = []
    for r in range(2, ws.max_row + 1):
        acc = ws.cell(row=r, column=1).value
        v = ws.cell(row=r, column=3).value
        if not isinstance(v, (int, float)):
            continue
        accs.append((acc, v))
    print(f"  Antal rader: {len(accs)}")
    print(f"  Min konto: {min(accs, key=lambda x: int(str(x[0])))[0]}")
    print(f"  Max konto: {max(accs, key=lambda x: int(str(x[0])))[0]}")
    print(f"  Första 5: {accs[:5]}")
    print(f"  Sista 5: {accs[-5:]}")
    wb.close()
