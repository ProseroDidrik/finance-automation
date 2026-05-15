import sys
from pathlib import Path
import openpyxl
sys.stdout.reconfigure(encoding="utf-8")

p = Path(r"C:\Users\DidWac\Prosero Dropbox\Didrik Wachtmeister\Phoenix Foundation\April alla filer\Get testfiles\extracted\202604\Denmark\Referens\216_Balance pr. 300426 SIKOM Danmark.xlsx")
wb = openpyxl.load_workbook(str(p), data_only=True)
ws = wb.active

bold_rows = []
nonbold_rows = []
for r in range(7, ws.max_row + 1):
    acc = ws.cell(row=r, column=1).value
    name = ws.cell(row=r, column=2).value
    v = ws.cell(row=r, column=3).value
    if not isinstance(v, (int, float)) or v == 0:
        continue
    # check if account cell or name cell is bold/underline
    is_bold = False
    is_underline = False
    for c in range(1, 4):
        cell = ws.cell(row=r, column=c)
        if cell.font:
            if cell.font.b:
                is_bold = True
            if cell.font.u:
                is_underline = True
    info = (r, acc, str(name)[:50], v, is_bold, is_underline)
    if is_bold:
        bold_rows.append(info)
    else:
        nonbold_rows.append(info)

print(f"Bold rows: {len(bold_rows)}  Non-bold: {len(nonbold_rows)}")
print(f"Sum BOLD:     {sum(x[3] for x in bold_rows):>15,.2f}")
print(f"Sum NON-BOLD: {sum(x[3] for x in nonbold_rows):>15,.2f}")
print(f"\nExempel BOLD rader:")
for x in bold_rows[:20]:
    print(f"  R{x[0]:3d} acc={x[1]!r:<10} '{x[2]:<48}' {x[3]:>15,.2f}  bold={x[4]} und={x[5]}")
