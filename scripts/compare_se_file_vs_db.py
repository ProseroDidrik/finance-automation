# -*- coding: utf-8 -*-
"""Jamfor Mercur-manadsrapporten mot Postgres-warehouse (fact_balances).

Kalla : _uploads/Svenska bolag for jamforelse.xlsx  (Mercur, manadsdata)
Mot   : fact_balances  (SE, resultatrakning, 2026 jan-apr)
Ut    : _uploads/SE-jamforelse fil vs warehouse v4.xlsx

Verkligt utfall for ett svenskt bolag = SIE-bas + MAN-A + IMP_ADJ-A. MAN och
IMP_ADJ ar additiva justeringslager ovanpa SIE-basen (inte alternativa kallor)
och alltid scenario A. DB-sidan byggs darfor som:
    SIE-bas (best source)  +  MAN-A  +  IMP_ADJ-A

Periodsemantik (empiriskt verifierad 2026-05-21):
  SIE_PSALDO    -> manadsrorelse  (#PSALDO ar per-manad, jamfors rakt av)
  SIE_VER/SIE   -> YTD            (manad = m - (m-1), jan = m sjalv)
  MAN / IMP_ADJ -> manadsrorelse  (period_type='monthly', jamfors rakt av)
Best source SE (warehouse-prioritet): SIE_PSALDO > SIE_VER > SIE.
Tecken: SIE-rader foljer SIE-konvention (intakt -), filen Mercur (intakt +)
        -> SIE-sidan flippas x-1 (inga P_-konton finns i SIE-data).
        MAN/IMP_ADJ lagras i Mercur-konvention -> ingen flip (x+1).
Kontoklass: dim_account_map, narmaste GROUP-forfader uppat via parent_id.
        MAN/IMP_ADJ:s P-koder (P_30 ...) ar account_id-noder utan company_id
        -> egen pcode-karta; vanliga konton via (cid, account_code).

Tva ETL-fynd som jamforelsen synliggor (se djupdyk-bladet):
  1. load_sie.py:RE_PSALDO matchar {[^}]*} -> laddar bade {}-totalen OCH
     dimensionssplit-rader -> SIE_PSALDO blir 2-3x for dim-exporterande bolag.
  2. SIE_VER (syntetiserad ur verifikat) ar ofullstandig for bolag vars
     SIE-fil saknar kompletta #VER -> SIE (#RES) ar da narmare facit.
"""
import os
import re
import subprocess
import sys
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC_FILE = None
for _f in os.listdir(os.path.join(REPO, "_uploads")):
    if _f.lower().startswith("svenska bolag") and _f.lower().endswith(".xlsx") \
            and not _f.startswith("~$"):
        SRC_FILE = os.path.join(REPO, "_uploads", _f)
OUT_FILE = os.path.join(REPO, "_uploads", "SE-jamforelse fil vs warehouse v4.xlsx")
PERIODS = ["202601", "202602", "202603", "202604"]
SRC_PRIORITY = ["SIE_PSALDO", "SIE_VER", "SIE"]


def get_db_url():
    if os.environ.get("DATABASE_URL"):
        return
    out = subprocess.run(
        "az keyvault secret show --vault-name kv-finauto-6427 "
        "--name database-url --query value -o tsv",
        capture_output=True, text=True, shell=True,
    )
    url = out.stdout.strip()
    if not url.startswith("postgres"):
        sys.exit("Kunde inte hamta DATABASE_URL: " + (out.stderr or out.stdout)[:300])
    os.environ["DATABASE_URL"] = url


# --------------------------------------------------------------------------
# 1. Mercur-filen
# --------------------------------------------------------------------------
def parse_file():
    wb = openpyxl.load_workbook(SRC_FILE, data_only=True)
    ws = wb[wb.sheetnames[0]]
    periodcol = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=2, column=c).value
        if v is not None and re.fullmatch(r"\d{6}", str(v)):
            periodcol[str(v)] = c
    file_data, cid_names, cid_order = {}, {}, []
    cur = None
    for r in range(3, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        if a is None:
            continue
        a = str(a).strip()
        if a.startswith("CID:"):
            m = re.match(r"CID:\s*(\d+)\s+(.*)", a)
            if m:
                cur = int(m.group(1))
                cid_names[cur] = m.group(2).strip()
                if cur not in cid_order:
                    cid_order.append(cur)
        elif a.startswith("TOP:"):
            cur = None
        elif a.startswith("GROUP:") and cur is not None:
            grp = a.split(":", 1)[1].strip()
            file_data[(cur, grp)] = {
                p: float(ws.cell(row=r, column=periodcol[p]).value)
                for p in PERIODS
                if isinstance(ws.cell(row=r, column=periodcol[p]).value, (int, float))
            }
    return file_data, cid_names, cid_order


# --------------------------------------------------------------------------
# 2. Kontoklass-kartor: leaf  (company_id, account_code) -> GROUP
#                       pcode (account_id, t.ex. P_30)   -> GROUP
# --------------------------------------------------------------------------
def build_account_group_map(con, groups):
    groupset = set(groups)
    rows = con.execute(
        "SELECT account_id, parent_id, is_aggregated, company_id, account_code "
        "FROM dim_account_map"
    ).fetchall()
    parent, info = {}, {}
    for account_id, parent_id, is_agg, company_id, account_code in rows:
        parent[account_id] = parent_id
        info[account_id] = (is_agg, company_id, account_code)

    def nearest_group(account_id):
        seen, node = set(), account_id
        while node is not None and node not in seen:
            seen.add(node)
            if node in groupset:
                return node
            node = parent.get(node)
        return None

    leaf_map, pcode_map = {}, {}
    for account_id, (is_agg, company_id, account_code) in info.items():
        if not is_agg and company_id is not None and account_code is not None:
            g = nearest_group(account_id)
            if g is not None:
                leaf_map[(int(company_id), str(account_code))] = g
        elif isinstance(account_id, str) and account_id.startswith("P_"):
            # P-koder ar account_id-noder (company_id/account_code = NULL).
            # MAN/IMP_ADJ bokfors pa dem -> behover en egen grupp-karta.
            g = nearest_group(account_id)
            if g is not None:
                pcode_map[account_id] = g
    return leaf_map, pcode_map


# --------------------------------------------------------------------------
# 3. DB-sidan
# --------------------------------------------------------------------------
def fetch_rows(con, se_ids):
    return con.execute(
        """
        SELECT company_id, account_code, period, source_kind, amount
        FROM fact_balances
        WHERE scenario='A'
          AND source_kind IN ('SIE_PSALDO','SIE_VER','SIE','MAN','IMP_ADJ')
          AND period IN ('202601','202602','202603','202604')
          AND company_id = ANY(%s)
        """,
        (list(se_ids),),
    ).fetchall()


def pick_best(rows):
    have = defaultdict(set)
    for cid, _acc, _per, sk, _amt in rows:
        have[cid].add(sk)
    best = {}
    for cid, kinds in have.items():
        for sk in SRC_PRIORITY:
            if sk in kinds:
                best[cid] = sk
                break
    return best


def derive_monthly(pv, sk):
    """pv = {period: amount}. Returnera {period: manadsvarde}."""
    if sk == "SIE_PSALDO":                       # redan manadsdata
        return {p: pv.get(p, 0.0) for p in PERIODS}
    monthly, prev = {}, 0.0                       # SIE_VER/SIE -> YTD, derivera
    for p in PERIODS:
        cur = pv.get(p)
        if cur is None:
            cur = prev
        monthly[p] = cur - prev
        prev = cur
    return monthly


def compute_db(rows, sk_for_cid, leaf_map):
    """sk_for_cid: cid -> source_kind att anvanda. Returnerar (cid,grp)->{per:val}."""
    raw = defaultdict(lambda: defaultdict(float))
    for cid, acc, per, sk, amt in rows:
        if sk == sk_for_cid.get(cid):
            raw[(cid, str(acc))][per] += float(amt)
    db_group = defaultdict(lambda: defaultdict(float))
    for (cid, acc), pv in raw.items():
        monthly = derive_monthly(pv, sk_for_cid[cid])
        flip = 1.0 if acc.startswith("P_") else -1.0
        grp = leaf_map.get((cid, acc))
        if grp is None:
            continue
        tgt = db_group[(cid, grp)]
        for p in PERIODS:
            tgt[p] += monthly[p] * flip
    return db_group


def compute_adjustments(rows, leaf_map, pcode_map):
    """MAN-A + IMP_ADJ-A som additivt lager ovanpa SIE-basen.

    Bada ar redan manadsdata (period_type='monthly') och lagras i
    Mercur-konvention -> ingen YTD-derivering, ingen SIE-flip (x+1, samma
    teckenkonvention som filen). P-koder mappas via pcode_map (account_id),
    vanliga konton via leaf_map (cid, account_code). BS-ben (1410/1470/2990
    ...) saknar P&L-grupp och faller bort - korrekt, RR-jamforelse.
    Returnerar ((cid,grp)->{per:val}, stats).
    """
    adj = defaultdict(lambda: defaultdict(float))
    stats = {"mapped": 0, "dropped": 0}
    for cid, acc, per, sk, amt in rows:
        if sk not in ("MAN", "IMP_ADJ"):
            continue
        acc = str(acc)
        grp = pcode_map.get(acc) if acc.startswith("P_") \
            else leaf_map.get((int(cid), acc))
        if grp is None:                       # BS-ben / oklassat -> ej RR
            stats["dropped"] += 1
            continue
        stats["mapped"] += 1
        adj[(int(cid), grp)][per] += float(amt)
    return adj, stats


# --------------------------------------------------------------------------
# 4. Klassning + arbetsbok
# --------------------------------------------------------------------------
HEAD = Font(bold=True, color="FFFFFF")
HEADFILL = PatternFill("solid", fgColor="305496")
TITLE = Font(bold=True, size=12)
BOLD = Font(bold=True)
FILL = {
    "ok": PatternFill("solid", fgColor="E2EFDA"),
    "periodiseringsbrus": PatternFill("solid", fgColor="FFF2CC"),
    "avvik": PatternFill("solid", fgColor="F8CBAD"),
    "saknas_i_db": PatternFill("solid", fgColor="D9D9D9"),
}
NUMFMT = "#,##0.00"


def style_header(ws, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=1, column=c)
        cell.font, cell.fill = HEAD, HEADFILL
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def widths(ws, spec):
    for col, w in spec:
        ws.column_dimensions[col].width = w


def main():
    get_db_url()
    sys.path.insert(0, REPO)
    from db import connect

    print("[1/6] Laser Mercur-filen ...")
    file_data, cid_names, cid_order = parse_file()
    groups = sorted({g for (_c, g) in file_data})

    con = connect(read_only=True)

    print("[2/6] SE-bolag + kontoklass-karta ...")
    se_rows = con.execute(
        "SELECT company_id, name, COALESCE(kind,'') FROM dim_company "
        "WHERE country='Sweden'"
    ).fetchall()
    se_meta = {int(r[0]): {"name": r[1], "kind": r[2].lower()} for r in se_rows}
    scope = [c for c in cid_order if c in se_meta]
    leaf_map, pcode_map = build_account_group_map(con, groups)

    # Konsolideringskarta: parent -> barnbolag (dim_company.parent_id = kol G
    # i Dotterbolagslistan). Konsoliderade bolag har ingen egen SIE-data utan
    # harleds som summan av barnbolagen.
    children = defaultdict(list)
    for r in con.execute(
        "SELECT company_id, parent_id FROM dim_company WHERE parent_id IS NOT NULL"
    ).fetchall():
        children[int(r[1])].append(int(r[0]))
    consolidated = [c for c in scope if se_meta[c]["kind"] == "consolidated"]

    print("[3/6] Hamtar fact_balances + bygger DB-vyer ...")
    need = set(scope)
    for p in consolidated:
        need |= set(children.get(p, []))
    rows = fetch_rows(con, need)
    best = pick_best(rows)
    db_best = compute_db(rows, best, leaf_map)
    db_sie = compute_db(rows, {c: "SIE" for c in need}, leaf_map)

    # MAN-A + IMP_ADJ-A: additiva justeringslager ovanpa SIE-basen. Laggs in
    # i bada vyerna fore konsolidering sa parent-summorna far med dem.
    db_adj, adj_stats = compute_adjustments(rows, leaf_map, pcode_map)
    for view in (db_best, db_sie):
        for (cid, grp), pv in db_adj.items():
            tgt = view[(cid, grp)]
            for p in PERIODS:
                tgt[p] += pv.get(p, 0.0)
    print(f"        MAN/IMP_ADJ-lager: {adj_stats['mapped']} rader till "
          f"P&L-grupp, {adj_stats['dropped']} BS-ben/oklassade (ej i RR)")

    # Konsolidera: parent (cid,grp) = summa over barnbolagens varden.
    cons_missing = {}                       # parent -> [barn utan SIE-data]
    for p in consolidated:
        kids = children.get(p, [])
        best[p] = f"KONSOL:{len(kids)}"
        cons_missing[p] = [k for k in kids
                           if not any((k, g) in db_best for g in groups)]
        for view in (db_best, db_sie):
            for grp in groups:
                acc = None
                for k in kids:
                    cv = view.get((k, grp))
                    if cv is None:
                        continue
                    if acc is None:
                        acc = defaultdict(float)
                    for per in PERIODS:
                        acc[per] += cv[per]
                if acc is not None:
                    view[(p, grp)] = acc

    print("[4/6] Jamfor ...")
    monthly_rows, ytd_rows = [], []
    health = {}                       # cid -> (verdikt, abs_file, abs_best, abs_sie)
    for cid in scope:
        name = se_meta[cid]["name"]
        src = best.get(cid)
        grp_set = {g for (c, g) in file_data if c == cid}
        grp_set |= {g for (c, g) in db_best if c == cid}
        abs_file = abs_best = abs_sie = 0.0
        rows_for_cid = []
        for grp in sorted(grp_set):
            fvals = file_data.get((cid, grp), {})
            dvals = db_best.get((cid, grp))
            svals = db_sie.get((cid, grp))
            file_ytd = sum(fvals.get(p, 0.0) for p in PERIODS)
            db_ytd = sum(dvals.get(p, 0.0) for p in PERIODS) if dvals is not None else None
            sie_ytd = sum(svals.get(p, 0.0) for p in PERIODS) if svals is not None else None
            abs_file += abs(file_ytd)
            if db_ytd is not None:
                abs_best += abs(db_ytd)
            if sie_ytd is not None:
                abs_sie += abs(sie_ytd)
            if db_ytd is None:
                ydiff = None
                if abs(file_ytd) < 1.0:          # tom i bada kallorna
                    ytd_ok, ytd_status = True, "ok"
                else:
                    ytd_ok, ytd_status = False, "saknas_i_db"
            else:
                ydiff = file_ytd - db_ytd
                ytd_ok = abs(ydiff) <= max(1.0, 0.005 * abs(file_ytd))
                ytd_status = "ok" if ytd_ok else "avvik"
            rows_for_cid.append((grp, fvals, dvals, file_ytd, db_ytd, sie_ytd,
                                 ydiff, ytd_status, ytd_ok))

        # ---- per-bolag DB-halsa --------------------------------------------
        # err = summa avvikelse mot facit (filen) per grupp; jamfor best_source
        # mot SIE (#RES). None -> hela facit-beloppet raknas som fel.
        err_best = sum(abs(fy - (dy if dy is not None else 0.0))
                       for _g, _f, _d, fy, dy, _s, *_ in rows_for_cid)
        err_sie = sum(abs(fy - (sy if sy is not None else 0.0))
                      for _g, _f, _d, fy, _dy, sy, *_ in rows_for_cid)
        tol = max(200.0, 0.01 * abs_file)
        if src is None:
            verdict = "ingen SIE-data"
        elif err_best <= tol:
            verdict = "ren"
        elif err_sie <= tol and err_sie < 0.3 * err_best:
            if src and src.startswith("KONSOL"):
                mech = "konsoliderad - barnbolags kallfel"
            elif src == "SIE_PSALDO":
                mech = "SIE_PSALDO dim-dubblering"
            else:
                mech = "SIE_VER ofullstandig"
            verdict = f"best_source avviker ({mech}) - SIE #RES stammer"
        else:
            verdict = "avvik kvarstar - granska"
        health[cid] = (verdict, abs_file, abs_best, abs_sie)

        # ---- rader ----------------------------------------------------------
        for grp, fvals, dvals, file_ytd, db_ytd, sie_ytd, ydiff, ytd_status, ytd_ok \
                in rows_for_cid:
            ytd_rows.append([
                cid, name, grp, src or "-", round(file_ytd, 2),
                round(db_ytd, 2) if db_ytd is not None else None,
                round(ydiff, 2) if ydiff is not None else None,
                round(ydiff / file_ytd * 100, 2)
                if (ydiff is not None and abs(file_ytd) > 0.005) else None,
                round(sie_ytd, 2) if sie_ytd is not None else None,
                ytd_status,
            ])
            for p in PERIODS:
                fv = fvals.get(p)
                dv = dvals.get(p) if dvals is not None else None
                if dv is None:
                    if abs(fv or 0.0) < 0.005:
                        status, diag = "ok", ""        # tom i bada kallorna
                    else:
                        status = "saknas_i_db"
                        diag = "filen har varde, DB saknar konto/mappning"
                elif abs(fv or 0.0) < 0.005 and abs(dv or 0.0) < 0.005:
                    status, diag = "ok", ""
                else:
                    diff0 = (fv or 0.0) - (dv or 0.0)
                    if abs(diff0) <= max(1.0, 0.005 * abs(fv or 0.0)):
                        status, diag = "ok", ""
                    elif not ytd_ok:
                        status = "avvik"
                        diag = health[cid][0] if health[cid][0] not in ("ren", "") else ""
                    elif src == "SIE_PSALDO":
                        status, diag = "avvik", "PSALDO ar manadsdata - manadsavvik trots YTD-ok"
                    else:
                        status = "periodiseringsbrus"
                        diag = "manadssplit skiljer (SIE_VER verifikatdatum), YTD stammer"
                diff = (fv or 0.0) - (dv or 0.0) if dv is not None else None
                monthly_rows.append([
                    cid, name, grp, p, src or "-",
                    round(fv, 2) if fv is not None else None,
                    round(dv, 2) if dv is not None else None,
                    round(diff, 2) if diff is not None else None,
                    round(diff / fv * 100, 2)
                    if (diff is not None and fv not in (None, 0)) else None,
                    status, diag,
                ])

    print("[5/6] Sammanstaller ...")
    cid_summary = []
    for cid in scope:
        mine = [r for r in monthly_rows if r[0] == cid]
        yt = [r for r in ytd_rows if r[0] == cid]
        cnt = defaultdict(int)
        for r in mine:
            cnt[r[9]] += 1
        verdict, abs_file, abs_best, abs_sie = health[cid]
        cid_summary.append([
            cid, se_meta[cid]["name"], best.get(cid, "-"), verdict,
            cnt["ok"], cnt["periodiseringsbrus"], cnt["avvik"], cnt["saknas_i_db"],
            sum(1 for r in yt if r[9] == "avvik"),
            round(max((abs(r[6]) for r in yt if r[6] is not None), default=0.0), 2),
            round(abs_file, 0), round(abs_best, 0), round(abs_sie, 0),
        ])

    grp_summary = []
    for grp in groups:
        mine = [r for r in monthly_rows if r[2] == grp]
        yt = [r for r in ytd_rows if r[2] == grp]
        cnt = defaultdict(int)
        for r in mine:
            cnt[r[9]] += 1
        grp_summary.append([
            grp, cnt["ok"], cnt["periodiseringsbrus"], cnt["avvik"],
            cnt["saknas_i_db"], sum(1 for r in yt if r[9] == "avvik"),
        ])

    # ---- arbetsbok --------------------------------------------------------
    print("[6/6] Skriver " + os.path.basename(OUT_FILE) + " ...")
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Manadsmatris 2026"
    ws.append(["CID", "Namn", "GROUP", "Period", "Kalla", "Filen", "Databas",
               "Diff (fil-db)", "Diff %", "Status", "Diagnos"])
    for r in monthly_rows:
        ws.append(r)
    style_header(ws, 11)
    for i, r in enumerate(monthly_rows, start=2):
        if r[9] in FILL:
            ws.cell(row=i, column=10).fill = FILL[r[9]]
        for c in (6, 7, 8):
            ws.cell(row=i, column=c).number_format = NUMFMT
    widths(ws, zip("ABCDEFGHIJK", (6, 32, 23, 9, 12, 14, 14, 14, 8, 20, 48)))

    ws = wb.create_sheet("YTD-kontroll jan-apr")
    ws.append(["CID", "Namn", "GROUP", "Kalla", "Filen YTD", "Databas YTD",
               "Diff (fil-db)", "Diff %", "Alt: SIE #RES YTD", "Status"])
    for r in ytd_rows:
        ws.append(r)
    style_header(ws, 10)
    for i, r in enumerate(ytd_rows, start=2):
        if r[9] in FILL:
            ws.cell(row=i, column=10).fill = FILL[r[9]]
        for c in (5, 6, 7, 9):
            ws.cell(row=i, column=c).number_format = NUMFMT
    widths(ws, zip("ABCDEFGHIJ", (6, 32, 23, 12, 15, 15, 15, 8, 17, 13)))

    ws = wb.create_sheet("Sammanfattning CID")
    ws.append(["CID", "Namn", "Kalla", "DB-halsa", "Manad OK", "Periodbrus",
               "Manad avvik", "Saknas", "YTD-avvik grp", "Max abs YTD-diff",
               "abs Filen", "abs DB", "abs DB-SIE(#RES)"])
    for r in cid_summary:
        ws.append(r)
    style_header(ws, 13)
    for i in range(2, len(cid_summary) + 2):
        for c in (10, 11, 12, 13):
            ws.cell(row=i, column=c).number_format = NUMFMT
    widths(ws, zip("ABCDEFGHIJKLM",
                   (6, 30, 12, 38, 9, 11, 12, 8, 14, 17, 15, 15, 17)))

    ws = wb.create_sheet("Sammanfattning GROUP")
    ws.append(["GROUP", "Manad OK", "Periodbrus", "Manad avvik", "Saknas",
               "YTD-avvik (CID)"])
    for r in grp_summary:
        ws.append(r)
    style_header(ws, 6)
    widths(ws, zip("ABCDEF", (24, 10, 12, 12, 9, 16)))

    write_deepdive(wb, scope, se_meta, best, health, monthly_rows, ytd_rows,
                   consolidated, children, cons_missing)
    wb.save(OUT_FILE)

    # ---- konsol ----------------------------------------------------------
    tot = len(monthly_rows)
    cnt = defaultdict(int)
    for r in monthly_rows:
        cnt[r[9]] += 1
    print("\n=== RESULTAT ===")
    print(f"SE-bolag i scope: {len(scope)}   Manadsrader: {tot}")
    for k in ("ok", "periodiseringsbrus", "avvik", "saknas_i_db"):
        print(f"  {k:20s}: {cnt[k]:5d}  ({cnt[k] / tot * 100:5.1f}%)")
    print("\nDB-halsa per bolag:")
    vc = defaultdict(list)
    for cid in scope:
        vc[health[cid][0]].append(cid)
    for v, ids in sorted(vc.items(), key=lambda t: -len(t[1])):
        print(f"  {v:42s}: {len(ids):2d}  {ids}")
    print("\nKlar: " + OUT_FILE)


# --------------------------------------------------------------------------
# 5. Djupdyk-blad
# --------------------------------------------------------------------------
def write_deepdive(wb, scope, se_meta, best, health, monthly_rows, ytd_rows,
                   consolidated, children, cons_missing):
    ws = wb.create_sheet("Avvikelse-djupdyk")
    ws.column_dimensions["A"].width = 4
    for col in "BCDEFG":
        ws.column_dimensions[col].width = 22
    ws.column_dimensions["B"].width = 32

    r = [1]

    def line(text="", *, title=False, bold=False, cells=None, fill=None):
        if cells:
            for j, v in enumerate(cells):
                c = ws.cell(row=r[0], column=2 + j, value=v)
                if bold:
                    c.font = BOLD
                if fill:
                    c.fill = fill
                if isinstance(v, (int, float)):
                    c.number_format = NUMFMT
        else:
            c = ws.cell(row=r[0], column=2, value=text)
            if title:
                c.font = TITLE
            elif bold:
                c.font = BOLD
        r[0] += 1

    line("Avvikelse-djupdyk - Mercur-fil vs warehouse (SE, RR, 2026 jan-apr)",
         title=True)
    line()
    line("DB-sidan = SIE-bas + MAN-A + IMP_ADJ-A (additiva justeringslager, "
         "Mercur-konv, scenario A).", bold=True)
    line("Slutsats: kontoklass-mappningen (dim_account_map) ar korrekt - DB "
         "reproducerar filens", bold=True)
    line("YTD-summor pa oret for bolag med ren kalldata. ETL-fynden nedan "
         "forklarar i princip alla")
    line("vasentliga avvikelser. Manadsavvik for SIE_VER-bolag ar i ovrigt "
         "periodiseringsbrus")
    line("(verifikatdatum vs Mercurs periodisering) - vantat, inget fel.")
    line()

    # --- Fynd 1 -----------------------------------------------------------
    line("FYND 1 - SIE_PSALDO-lanen i load_sie.py har tva fel", bold=True)
    line()
    line("FYND 1a - fel period_type-tagg (allvarligast - drabbar webappen)",
         bold=True)
    line("#PSALDO-raderna ar MANADSRORELSER, men load_sie.py taggar "
         "SIE_PSALDO med")
    line("period_type='ytd' (verifierat: samtliga 53 868 rader). "
         "report_pnl.sql rad 122-123")
    line("grenar pa period_type och raknar darfor fel for de 17 "
         "SIE_PSALDO-bolagen:")
    line("    amount_ytd   = cur.amount         -> ger EN manad, inte YTD")
    line("    amount_month = cur.amount - prev  -> differens av tva "
         "manadsrorelser = brus")
    line("Bevis - bolag 11, konto 6590: SIE_PSALDO jan-apr ar "
         "839 / 838 / 894 / 857 (fluktuerar,")
    line("dvs manadsdata - en YTD-serie vore monotont stigande). "
         "report_pnl YTD-apr ger 857,")
    line("korrekt YTD ar 3 427 (summan av de fyra).")
    line("Atgard: tagga SIE_PSALDO period_type='monthly'. OBS: den har "
         "jamforelsen behandlar")
    line("SIE_PSALDO korrekt som manad - workbookens DB-siffror ar rätt "
         "trots taggen.")
    line()
    line("FYND 1b - #PSALDO med dimensioner dubbelladdas", bold=True)
    line("load_sie.py RE_PSALDO matchar  \\{[^}]*\\}  -> traffar BADE "
         "{}-totalen OCH varje")
    line("dimensionssplit-rad. Bevis - bolag 186, konto 3010, 202604:")
    line("    #PSALDO 0 202604 3010 {} -1127159.49           <- kontototal")
    line("    #PSALDO 0 202604 3010 {1 \"10\"} -516326.70      <- dim 1 (summa = totalen)")
    line("    #PSALDO 0 202604 3010 {6 \"101014\"} -5471.00    <- dim 6 (summa = totalen igen)")
    line("-> SIE_PSALDO blir 2-3x for dim-exporterande bolag (23, 75, 186 "
         "nedan).")
    line("Fix: matcha endast  \\{\\s*\\}  och summera per konto. Forutsatter "
         "att filen alltid")
    line("emitterar {}-totalen - racka #PSALDO-rader utan {} per bolag innan "
         "deploy. Bolag 32 har")
    line("enbart {}-rader (flera per konto, summering korrekt dar).")
    line()
    line(cells=["CID", "Namn", "Kalla", "DB-halsa", "abs Filen", "abs DB",
                "abs DB-SIE(#RES)"], bold=True, fill=HEADFILL)
    for j in range(7):
        ws.cell(row=r[0] - 1, column=2 + j).font = HEAD
    aff1 = [c for c in scope if "SIE_PSALDO dim" in health[c][0]]
    for cid in sorted(aff1, key=lambda c: -health[c][2]):
        v, af, ab, asie = health[cid]
        line(cells=[cid, se_meta[cid]["name"], best.get(cid, "-"), v,
                    round(af, 0), round(ab, 0), round(asie, 0)])
    if not aff1:
        line("(inga)")
    line()

    # --- Fynd 2 -----------------------------------------------------------
    line("FYND 2 - SIE_VER ofullstandig for bolag med glesa verifikat", bold=True)
    line("SIE_VER syntetiseras ur #VER/#TRANS. Nar en SIE-fil saknar "
         "kompletta verifikat blir")
    line("SIE_VER ofullstandig. Da ar SIE (#RES, YTD) narmare facit. "
         "Ex bolag 14, Personnel YTD:")
    line("    Filen 6 336 944 | SIE_VER 692 842 | SIE (#RES) 6 336 944  "
         "<- SIE traffar exakt")
    line("Warehouse-prioritet SIE_VER > SIE valjer har fel kalla. "
         "Atgard: ladda om SIE med")
    line("--include-journal, eller lat best_source falla till SIE nar "
         "SIE_VER ar markant glesare.")
    line()
    line(cells=["CID", "Namn", "Kalla", "DB-halsa", "abs Filen", "abs DB",
                "abs DB-SIE(#RES)"], bold=True, fill=HEADFILL)
    for j in range(7):
        ws.cell(row=r[0] - 1, column=2 + j).font = HEAD
    aff2 = [c for c in scope if "SIE_VER ofull" in health[c][0]]
    for cid in sorted(aff2, key=lambda c: -health[c][1]):
        v, af, ab, asie = health[cid]
        line(cells=[cid, se_meta[cid]["name"], best.get(cid, "-"), v,
                    round(af, 0), round(ab, 0), round(asie, 0)])
    if not aff2:
        line("(inga)")
    line()

    # --- Kvarstaende avvik ------------------------------------------------
    line("FYND 3 - kvarstaende YTD-avvik (DB inkl. MAN-A + IMP_ADJ-A)",
         bold=True)
    line("DB-sidan ar nu SIE-bas + MAN-A + IMP_ADJ-A. Monster i tabellen "
         "nedan:")
    line(" (a) Personnel kraftigt under i DB (bolag 18, 152) - samma "
         "SIE_VER-glesa-verifikat som fynd 2,")
    line("     men bolaget har aven annat avvik sa det fastnar inte i "
         "auto-klassningen.")
    line(" (b) Avvik som kvarstar EFTER att MAN-A lagts till -> fact_balances "
         "MAN-A matchar inte")
    line("     filens justeringar (annan storlek / annan P&L-rad). Bor utredas "
         "separat - ex bolag 13:")
    line("     fact MAN-A = P_30 +1.0M (-> Net Sales), men filens avvik ar "
         "0.5M / 1.6M / 0.48M.")
    line(" (c) Bolag 7: konto 3616/4616 (intern 'VF Passera'-debitering, "
         "+/-1.94M) klassas som")
    line("     Net Sales/Materialkost via dim_account_map men som internpost "
         "i Mercur. Netto = 0.")
    line()
    line(cells=["CID", "GROUP", "Kalla", "Filen YTD", "Databas YTD",
                "Diff (fil-db)"], bold=True, fill=HEADFILL)
    for j in range(6):
        ws.cell(row=r[0] - 1, column=2 + j).font = HEAD
    explained = {c for c in scope
                 if "SIE_PSALDO dim" in health[c][0]
                 or "SIE_VER ofull" in health[c][0]
                 or health[c][0] == "ingen SIE-data"}
    rest = sorted(
        [yr for yr in ytd_rows
         if yr[9] == "avvik" and yr[6] is not None and yr[0] not in explained],
        key=lambda yr: -abs(yr[6]))
    for yr in rest[:30]:
        line(cells=[yr[0], yr[2], yr[3], yr[4], yr[5], yr[6]])
    if not rest:
        line("(inga)")
    line()
    # --- Konsoliderade bolag ----------------------------------------------
    line("KONSOLIDERADE BOLAG - DB harledd som summa av barnbolag", bold=True)
    line("Filen listar 12 konsoliderade bolag som saknar egen SIE-data. "
         "DB-sidan summeras")
    line("fran barnbolagen (dim_company.parent_id). Barnen arver sina "
         "kallfel - ex bolag 24")
    line("(Dala Las) = 72+73+74, dar 72 har #PSALDO dim-dubblering (fynd 1b). "
         "OBS: summering")
    line("eliminerar inte koncerninterna poster (INTER/INTEX/KONC) - "
         "avvik dar kan vara vantat.")
    line()
    line(cells=["Konsoliderat (CID)", "Namn", "Barnbolag", "Barn utan data"],
         bold=True, fill=HEADFILL)
    for j in range(4):
        ws.cell(row=r[0] - 1, column=2 + j).font = HEAD
    for p in consolidated:
        kids = children.get(p, [])
        miss = cons_missing.get(p, [])
        line(cells=[p, se_meta[p]["name"],
                    ", ".join(str(k) for k in kids),
                    ", ".join(str(k) for k in miss) if miss else "-"])
    line()
    line("Bolag med DB-halsa 'ren' och inga rader ovan: warehouse-data "
         "stammer mot Mercur-filen", bold=True)
    line("pa YTD-niva. Aterstaende manadsdiff for dem ar periodiseringsbrus.")


if __name__ == "__main__":
    main()
