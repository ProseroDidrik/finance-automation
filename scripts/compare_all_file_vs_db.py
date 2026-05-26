# -*- coding: utf-8 -*-
"""Jamfor Mercur-manadsrapporten (ALLA lander) mot Postgres-warehouse.

Kalla : _uploads/alla bolag*.xlsx   (Mercur, manadsdata)
Mot   : fact_balances               (alla lander, resultatrakning, 2026 jan-apr)
Ut    : _uploads/Alla bolag - jamforelse fil vs warehouse.xlsx

Syskon till compare_se_file_vs_db.py - samma idé, men alla lander och bara
kallorna IMP/SIE/SAF-T (filen ar exporterad "ENBART IMP SIE SAFT").

Per-land-semantik (warehouse describe_schema):
  SE  best_source SIE_PSALDO > SIE_VER > SIE   (PSALDO = manad, VER/SIE = YTD)
  NO  SAFT (YTD)
  FI / DK / DE / CENTR  IMP (manad)
Tecken: SIE-konv-kallor (SIE*/SAFT) lagras intakt-negativt -> flippas x-1 sa de
        far Mercur-konvention som filen. IMP ar redan Mercur-konv (ingen flip).
        P_-konton ar alltid Mercur-konv (+1).
Valuta: filens 'Value Local FX'-block (kol 55-106) = lokal valuta = exakt det
        fact_balances.amount lagrar. 'Value SEK'-blocket ignoreras - Mercurs
        SEK-omrakning misslyckades ("Valutaomrakning misslyckades" i headern).
Kontoklass: dim_account_map, narmaste GROUP-forfader uppat via parent_id.
Konsoliderade bolag (kind='consolidated') skippas - se filter i main().

Forvantat brus i NO-SAFT:
  Bolag 158 och 189 (Tripletex) avviker konsekvent ~3 % pa intaktskonton
  for att ClosingBalance > sum(GL-entries) i samma fil. Inte ETL-bug. Se
  docs/warehouse_semantics.md (sektion "Tripletex ClosingBalance vs GL").
"""
import os
import re
import subprocess
import sys
import zipfile
import xml.etree.ElementTree as ET
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC_FILE = None
for _f in sorted(os.listdir(os.path.join(REPO, "_uploads"))):
    if _f.lower().startswith("alla bolag") and _f.lower().endswith(".xlsx") \
            and not _f.startswith("~$"):
        SRC_FILE = os.path.join(REPO, "_uploads", _f)
OUT_FILE = os.path.join(REPO, "_uploads",
                        "Alla bolag - jamforelse fil vs warehouse.xlsx")
PERIODS = ["202601", "202602", "202603", "202604"]

# Kallor som filen innehaller. Prioritet per land (hogst forst).
ALL_KINDS = ("SIE_PSALDO", "SIE_VER", "SIE", "SAFT", "IMP")
PRIORITY = {
    "Sweden":  ["SIE_PSALDO", "SIE_VER", "SIE", "IMP"],
    "Norway":  ["SAFT", "IMP"],
    "Finland": ["IMP"],
    # Bolag 81 (Actas DK) levererar SAF-T istället för INL-Excel (undantag,
    # se CLAUDE.md) — DK måste därför söka SAFT före IMP, annars hittar
    # compare inget för 81 och konsol 132 visas tomt.
    "Denmark": ["SAFT", "IMP"],
    "Germany": ["IMP"],
    # CENTR-bolag levererar blandade format: SIE (50, 51, 53), SAFT (52 NO-orgnr,
    # 54 DK-CVR) eller IMP. Sök i prioritetsordning så att bolaget hittas
    # oavsett källtyp.
    "CENTR":   ["SIE_PSALDO", "SIE_VER", "SIE", "SAFT", "IMP"],
    "CA":      ["SIE_PSALDO", "SIE_VER", "SIE", "IMP"],
}
SIE_CONV = {"SIE", "SIE_PSALDO", "SIE_VER", "SAFT"}   # intakt-negativt -> flip
MONTHLY_KINDS = {"SIE_PSALDO", "IMP"}                  # ovriga ar YTD
COUNTRY_ORDER = ["Sweden", "Norway", "Finland", "Denmark", "Germany",
                 "CENTR", "CA"]


def get_db_url():
    if os.environ.get("DATABASE_URL"):
        return
    out = subprocess.run(
        "az keyvault secret show --vault-name kv-finauto-6427 "
        "--name database-url --query value -o tsv",
        capture_output=True, text=True, shell=True,
    )
    url = out.stdout.strip()
    if not url.startswith("postgres"):
        sys.exit("Kunde inte hamta DATABASE_URL: " + (out.stderr or out.stdout)[:300])
    os.environ["DATABASE_URL"] = url


# --------------------------------------------------------------------------
# 1. Mercur-filen  (ra XML - openpyxl klarar inte filens style-XML)
# --------------------------------------------------------------------------
def _ln(tag):
    return tag.rsplit("}", 1)[-1]


def _col_idx(ref):
    n = 0
    for ch in re.match(r"([A-Z]+)", ref).group(1):
        n = n * 26 + (ord(ch) - 64)
    return n


def read_sheet(path):
    """Las forsta bladet ur en xlsx via ra XML -> {rad: {kol: varde}}."""
    with zipfile.ZipFile(path) as z:
        shared = []
        if "xl/sharedStrings.xml" in z.namelist():
            for si in ET.fromstring(z.read("xl/sharedStrings.xml")):
                shared.append("".join(t.text or ""
                                      for t in si.iter() if _ln(t.tag) == "t"))
        wbroot = ET.fromstring(z.read("xl/workbook.xml"))
        rid = next([v for k, v in el.attrib.items() if _ln(k) == "id"][0]
                   for el in wbroot.iter() if _ln(el.tag) == "sheet")
        rels = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
        target = {r.attrib["Id"]: r.attrib["Target"] for r in rels}[rid]
        if not target.startswith("xl/"):
            target = "xl/" + target.lstrip("/")
        sroot = ET.fromstring(z.read(target))
    rows = {}
    for row in sroot.iter():
        if _ln(row.tag) != "row":
            continue
        cells = {}
        for c in row:
            if _ln(c.tag) != "c" or not c.attrib.get("r"):
                continue
            ci = _col_idx(c.attrib["r"])
            t = c.attrib.get("t")
            v = None
            for ch in c:
                if _ln(ch.tag) == "v":
                    v = ch.text
                elif _ln(ch.tag) == "is":
                    v = "".join(x.text or ""
                                for x in ch.iter() if _ln(x.tag) == "t")
            if v is None:
                continue
            if t == "s":
                v = shared[int(v)]
            elif t in (None, "n"):
                try:
                    v = float(v)
                except (ValueError, TypeError):
                    pass
            cells[ci] = v
        rows[int(row.attrib["r"])] = cells
    return rows


def parse_file():
    rows = read_sheet(SRC_FILE)
    # Periodkolumner ur 'Value Local FX'-blocket (rad 7 = blocketikett,
    # rad 8 = period). SEK-blocket ignoreras.
    r7, r8 = rows.get(7, {}), rows.get(8, {})
    periodcol = {}
    for c, label in r7.items():
        if label == "Value Local FX":
            p = r8.get(c)
            if p is not None:
                ps = str(p).split(".")[0]
                if re.fullmatch(r"\d{6}", ps) and ps in PERIODS:
                    periodcol[ps] = c
    missing = [p for p in PERIODS if p not in periodcol]
    if missing:
        sys.exit("Saknar 'Value Local FX'-kolumner for: " + ", ".join(missing))

    file_data, cid_names, cid_order = {}, {}, []
    cur = None
    for r in sorted(rows):
        a = rows[r].get(1)
        if a is None:
            continue
        a = str(a).strip()
        if a.startswith("CID:"):
            m = re.match(r"CID:\s*(\d+)\s+(.*)", a)
            if m:
                cur = int(m.group(1))
                cid_names[cur] = m.group(2).strip()
                if cur not in cid_order:
                    cid_order.append(cur)
        elif a.startswith("TOP:"):
            cur = None
        elif a.startswith("GROUP:") and cur is not None:
            grp = a.split(":", 1)[1].strip()
            file_data[(cur, grp)] = {
                p: float(rows[r][periodcol[p]])
                for p in PERIODS
                if isinstance(rows[r].get(periodcol[p]), (int, float))
            }
    return file_data, cid_names, cid_order


# --------------------------------------------------------------------------
# 2. Kontoklass-karta: leaf (company_id, account_code) -> GROUP
# --------------------------------------------------------------------------
def build_account_group_map(con, groups):
    groupset = set(groups)
    rows = con.execute(
        "SELECT account_id, parent_id, is_aggregated, company_id, account_code "
        "FROM dim_account_map"
    ).fetchall()
    parent, info = {}, {}
    for account_id, parent_id, is_agg, company_id, account_code in rows:
        parent[account_id] = parent_id
        info[account_id] = (is_agg, company_id, account_code)

    def nearest_group(account_id):
        seen, node = set(), account_id
        while node is not None and node not in seen:
            seen.add(node)
            if node in groupset:
                return node
            node = parent.get(node)
        return None

    leaf_map = {}
    for account_id, (is_agg, company_id, account_code) in info.items():
        if is_agg or company_id is None or account_code is None:
            continue
        g = nearest_group(account_id)
        if g is not None:
            leaf_map[(int(company_id), str(account_code))] = g
    return leaf_map


# --------------------------------------------------------------------------
# 3. DB-sidan
# --------------------------------------------------------------------------
def fetch_rows(con, ids):
    return con.execute(
        """
        SELECT company_id, account_code, period, source_kind, amount
        FROM fact_balances
        WHERE scenario = 'A'
          AND source_kind = ANY(%s)
          AND period = ANY(%s)
          AND company_id = ANY(%s)
        """,
        (list(ALL_KINDS), PERIODS, list(ids)),
    ).fetchall()


def pick_best(rows, country_of):
    have = defaultdict(set)
    for cid, _acc, _per, sk, _amt in rows:
        have[cid].add(sk)
    best = {}
    for cid, kinds in have.items():
        for sk in PRIORITY.get(country_of.get(cid, ""), ["IMP"]):
            if sk in kinds:
                best[cid] = sk
                break
    return best


def derive_monthly(pv, sk):
    """pv = {period: amount}. Returnera {period: manadsvarde}.

    PERIODS ligger inom ett rakenskapsar (2026) -> YTD-derivering startar pa 0.
    """
    if sk in MONTHLY_KINDS:                       # redan manadsdata
        return {p: pv.get(p, 0.0) for p in PERIODS}
    monthly, prev = {}, 0.0                        # YTD -> derivera
    for p in PERIODS:
        cur = pv.get(p)
        if cur is None:
            cur = prev
        monthly[p] = cur - prev
        prev = cur
    return monthly


def compute_db(rows, sk_for_cid, leaf_map):
    """Returnerar (cid, grp) -> {period: manadsvarde i Mercur-konvention}."""
    raw = defaultdict(lambda: defaultdict(float))
    for cid, acc, per, sk, amt in rows:
        if sk == sk_for_cid.get(cid):
            raw[(cid, str(acc))][per] += float(amt)
    db_group = defaultdict(lambda: defaultdict(float))
    for (cid, acc), pv in raw.items():
        sk = sk_for_cid[cid]
        monthly = derive_monthly(pv, sk)
        # Flip till Mercur-konvention: SIE-konv-kallor x-1, P_-konton/IMP x+1.
        flip = 1.0 if acc.startswith("P_") else (-1.0 if sk in SIE_CONV else 1.0)
        grp = leaf_map.get((cid, acc))
        if grp is None:
            continue
        tgt = db_group[(cid, grp)]
        for p in PERIODS:
            tgt[p] += monthly[p] * flip
    return db_group


# --------------------------------------------------------------------------
# 4. Stil + arbetsbok
# --------------------------------------------------------------------------
HEAD = Font(bold=True, color="FFFFFF")
HEADFILL = PatternFill("solid", fgColor="305496")
BOLD = Font(bold=True)
FILL = {
    "ok": PatternFill("solid", fgColor="E2EFDA"),
    "periodiseringsbrus": PatternFill("solid", fgColor="FFF2CC"),
    "avvik": PatternFill("solid", fgColor="F8CBAD"),
    "saknas_i_db": PatternFill("solid", fgColor="D9D9D9"),
}
NUMFMT = "#,##0.00"


def style_header(ws, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=1, column=c)
        cell.font, cell.fill = HEAD, HEADFILL
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def widths(ws, spec):
    for col, w in spec:
        ws.column_dimensions[col].width = w


def main():
    if SRC_FILE is None:
        sys.exit("Hittar ingen '_uploads/alla bolag*.xlsx'.")
    get_db_url()
    sys.path.insert(0, REPO)
    from db import connect

    print("[1/6] Laser Mercur-filen: " + os.path.basename(SRC_FILE))
    file_data, cid_names, cid_order = parse_file()
    groups = sorted({g for (_c, g) in file_data})

    con = connect(read_only=True)

    print("[2/6] Bolagsregister + kontoklass-karta ...")
    meta = {}
    for r in con.execute(
        "SELECT company_id, name, COALESCE(country,''), COALESCE(kind,''), "
        "parent_id FROM dim_company"
    ).fetchall():
        meta[int(r[0])] = {"name": r[1], "country": r[2],
                           "kind": (r[3] or "").lower(),
                           "parent_id": int(r[4]) if r[4] is not None else None}
    country_of = {c: m["country"] for c, m in meta.items()}
    # Konsoliderade bolag (kind='consolidated') skippas - jamforelsen ar bara
    # meningsfull for bolag som har egen kalldata. Mercur-filen listar dem som
    # egna CID-rader men de speglar bara summan av barnbolagen.
    scope = [c for c in cid_order
             if c in meta and meta[c]["kind"] != "consolidated"]
    skipped = [c for c in cid_order if c not in meta]
    skipped_consolidated = [c for c in cid_order
                            if c in meta and meta[c]["kind"] == "consolidated"]
    leaf_map = build_account_group_map(con, groups)

    children = defaultdict(list)
    for c, m in meta.items():
        if m["parent_id"] is not None:
            children[m["parent_id"]].append(c)
    consolidated = []  # ingen derivering nodvandig - se filter ovan

    print("[3/6] Hamtar fact_balances + bygger DB-vyer ...")
    need = set(scope)
    for p in consolidated:
        need |= set(children.get(p, []))
    rows = fetch_rows(con, need)
    best = pick_best(rows, country_of)
    db = compute_db(rows, best, leaf_map)

    # Konsoliderade bolag: (parent, grp) = summa over barnbolagens varden.
    for p in consolidated:
        kids = children.get(p, [])
        best[p] = "KONSOL:%d" % len(kids)
        for grp in groups:
            acc = None
            for k in kids:
                cv = db.get((k, grp))
                if cv is None:
                    continue
                if acc is None:
                    acc = defaultdict(float)
                for per in PERIODS:
                    acc[per] += cv[per]
            if acc is not None:
                db[(p, grp)] = acc

    print("[4/6] Jamfor ...")
    monthly_rows, ytd_rows = [], []
    health = {}                       # cid -> (verdikt, abs_file, abs_db)
    for cid in scope:
        name = meta[cid]["name"]
        land = meta[cid]["country"]
        src = best.get(cid)
        grp_set = {g for (c, g) in file_data if c == cid}
        grp_set |= {g for (c, g) in db if c == cid}
        abs_file = abs_db = 0.0
        rows_for_cid = []
        for grp in sorted(grp_set):
            fvals = file_data.get((cid, grp), {})
            dvals = db.get((cid, grp))
            file_ytd = sum(fvals.get(p, 0.0) for p in PERIODS)
            db_ytd = (sum(dvals.get(p, 0.0) for p in PERIODS)
                      if dvals is not None else None)
            abs_file += abs(file_ytd)
            if db_ytd is not None:
                abs_db += abs(db_ytd)
            if db_ytd is None:
                ydiff = None
                if abs(file_ytd) < 1.0:
                    ytd_ok, ytd_status = True, "ok"
                else:
                    ytd_ok, ytd_status = False, "saknas_i_db"
            else:
                ydiff = file_ytd - db_ytd
                ytd_ok = abs(ydiff) <= max(1.0, 0.005 * abs(file_ytd))
                ytd_status = "ok" if ytd_ok else "avvik"
            rows_for_cid.append((grp, fvals, dvals, file_ytd, db_ytd,
                                 ydiff, ytd_status, ytd_ok))

        err = sum(abs(fy - (dy if dy is not None else 0.0))
                  for _g, _f, _d, fy, dy, *_ in rows_for_cid)
        tol = max(200.0, 0.01 * abs_file)
        if src is None:
            verdict = "ingen data"
        elif err <= tol:
            verdict = "ren"
        else:
            verdict = "avvik"
        health[cid] = (verdict, abs_file, abs_db)

        for grp, fvals, dvals, file_ytd, db_ytd, ydiff, ytd_status, ytd_ok \
                in rows_for_cid:
            ytd_rows.append([
                cid, name, land, grp, src or "-", round(file_ytd, 2),
                round(db_ytd, 2) if db_ytd is not None else None,
                round(ydiff, 2) if ydiff is not None else None,
                round(ydiff / file_ytd * 100, 2)
                if (ydiff is not None and abs(file_ytd) > 0.005) else None,
                ytd_status,
            ])
            for p in PERIODS:
                fv = fvals.get(p)
                dv = dvals.get(p) if dvals is not None else None
                if dv is None:
                    if abs(fv or 0.0) < 0.005:
                        status = "ok"
                    else:
                        status = "saknas_i_db"
                elif abs(fv or 0.0) < 0.005 and abs(dv or 0.0) < 0.005:
                    status = "ok"
                else:
                    diff0 = (fv or 0.0) - (dv or 0.0)
                    if abs(diff0) <= max(1.0, 0.005 * abs(fv or 0.0)):
                        status = "ok"
                    elif not ytd_ok:
                        status = "avvik"
                    else:
                        status = "periodiseringsbrus"
                diff = (fv or 0.0) - (dv or 0.0) if dv is not None else None
                monthly_rows.append([
                    cid, name, land, grp, p, src or "-",
                    round(fv, 2) if fv is not None else None,
                    round(dv, 2) if dv is not None else None,
                    round(diff, 2) if diff is not None else None,
                    round(diff / fv * 100, 2)
                    if (diff is not None and fv not in (None, 0)) else None,
                    status,
                ])

    print("[5/6] Sammanstaller ...")
    cid_summary = []
    for cid in scope:
        mine = [r for r in monthly_rows if r[0] == cid]
        yt = [r for r in ytd_rows if r[0] == cid]
        cnt = defaultdict(int)
        for r in mine:
            cnt[r[10]] += 1
        verdict, abs_file, abs_db = health[cid]
        cid_summary.append([
            cid, meta[cid]["name"], meta[cid]["country"], best.get(cid, "-"),
            verdict, cnt["ok"], cnt["periodiseringsbrus"], cnt["avvik"],
            cnt["saknas_i_db"], sum(1 for r in yt if r[9] == "avvik"),
            round(max((abs(r[7]) for r in yt if r[7] is not None), default=0.0), 2),
            round(abs_file, 0), round(abs_db, 0),
        ])

    land_summary = []
    for land in COUNTRY_ORDER:
        mine = [r for r in monthly_rows if r[2] == land]
        if not mine and land not in (meta[c]["country"] for c in scope):
            continue
        cnt = defaultdict(int)
        for r in mine:
            cnt[r[10]] += 1
        nbolag = sum(1 for c in scope if meta[c]["country"] == land)
        land_summary.append([
            land, nbolag, cnt["ok"], cnt["periodiseringsbrus"],
            cnt["avvik"], cnt["saknas_i_db"],
        ])

    grp_summary = []
    for grp in groups:
        mine = [r for r in monthly_rows if r[3] == grp]
        cnt = defaultdict(int)
        for r in mine:
            cnt[r[10]] += 1
        grp_summary.append([
            grp, cnt["ok"], cnt["periodiseringsbrus"], cnt["avvik"],
            cnt["saknas_i_db"],
        ])

    # ---- arbetsbok --------------------------------------------------------
    print("[6/6] Skriver " + os.path.basename(OUT_FILE) + " ...")
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Manadsmatris 2026"
    ws.append(["CID", "Namn", "Land", "GROUP", "Period", "Kalla", "Filen",
               "Databas", "Diff (fil-db)", "Diff %", "Status"])
    for r in monthly_rows:
        ws.append(r)
    style_header(ws, 11)
    for i, r in enumerate(monthly_rows, start=2):
        if r[10] in FILL:
            ws.cell(row=i, column=11).fill = FILL[r[10]]
        for c in (7, 8, 9):
            ws.cell(row=i, column=c).number_format = NUMFMT
    widths(ws, zip("ABCDEFGHIJK", (6, 32, 9, 23, 9, 12, 14, 14, 14, 8, 18)))

    ws = wb.create_sheet("YTD-kontroll jan-apr")
    ws.append(["CID", "Namn", "Land", "GROUP", "Kalla", "Filen YTD",
               "Databas YTD", "Diff (fil-db)", "Diff %", "Status"])
    for r in ytd_rows:
        ws.append(r)
    style_header(ws, 10)
    for i, r in enumerate(ytd_rows, start=2):
        if r[9] in FILL:
            ws.cell(row=i, column=10).fill = FILL[r[9]]
        for c in (6, 7, 8):
            ws.cell(row=i, column=c).number_format = NUMFMT
    widths(ws, zip("ABCDEFGHIJ", (6, 32, 9, 23, 12, 15, 15, 15, 8, 13)))

    ws = wb.create_sheet("Sammanfattning CID")
    ws.append(["CID", "Namn", "Land", "Kalla", "DB-halsa", "Manad OK",
               "Periodbrus", "Manad avvik", "Saknas", "YTD-avvik grp",
               "Max abs YTD-diff", "abs Filen", "abs DB"])
    for r in cid_summary:
        ws.append(r)
    style_header(ws, 13)
    for i in range(2, len(cid_summary) + 2):
        for c in (11, 12, 13):
            ws.cell(row=i, column=c).number_format = NUMFMT
    widths(ws, zip("ABCDEFGHIJKLM",
                   (6, 30, 9, 12, 14, 9, 11, 12, 8, 14, 17, 15, 15)))

    ws = wb.create_sheet("Sammanfattning Land")
    ws.append(["Land", "Bolag", "Manad OK", "Periodbrus", "Manad avvik",
               "Saknas"])
    for r in land_summary:
        ws.append(r)
    style_header(ws, 6)
    widths(ws, zip("ABCDEF", (12, 8, 10, 12, 12, 9)))

    ws = wb.create_sheet("Sammanfattning GROUP")
    ws.append(["GROUP", "Manad OK", "Periodbrus", "Manad avvik", "Saknas"])
    for r in grp_summary:
        ws.append(r)
    style_header(ws, 5)
    widths(ws, zip("ABCDE", (24, 10, 12, 12, 9)))

    # ---- Avvikelser: storsta YTD-glappen ---------------------------------
    ws = wb.create_sheet("Avvikelser")
    ws.append(["CID", "Namn", "Land", "GROUP", "Kalla", "Filen YTD",
               "Databas YTD", "Glapp (fil-db)", "Status"])
    avv = []
    for r in ytd_rows:
        if r[9] in ("avvik", "saknas_i_db"):
            gap = r[7] if r[7] is not None else r[5]   # ydiff annars file_ytd
            avv.append((abs(gap or 0.0), r, gap))
    avv.sort(key=lambda t: -t[0])
    for _a, r, gap in avv:
        ws.append([r[0], r[1], r[2], r[3], r[4], r[5], r[6],
                   round(gap, 2) if gap is not None else None, r[9]])
    style_header(ws, 9)
    for i in range(2, len(avv) + 2):
        for c in (6, 7, 8):
            ws.cell(row=i, column=c).number_format = NUMFMT
        st = ws.cell(row=i, column=9).value
        if st in FILL:
            ws.cell(row=i, column=9).fill = FILL[st]
    widths(ws, zip("ABCDEFGHI", (6, 32, 9, 23, 12, 15, 15, 15, 13)))

    wb.save(OUT_FILE)

    # ---- konsol ----------------------------------------------------------
    tot = len(monthly_rows)
    cnt = defaultdict(int)
    for r in monthly_rows:
        cnt[r[10]] += 1
    print("\n=== RESULTAT ===")
    print("Bolag i scope: %d   Manadsrader: %d" % (len(scope), tot))
    if skipped:
        print("Bolag i filen som saknas i dim_company (hoppade): %s" % skipped)
    if skipped_consolidated:
        print("Konsoliderade bolag (skippade - ingen egen kalldata): %s"
              % skipped_consolidated)
    for k in ("ok", "periodiseringsbrus", "avvik", "saknas_i_db"):
        pct = (cnt[k] / tot * 100) if tot else 0.0
        print("  %-20s: %5d  (%5.1f%%)" % (k, cnt[k], pct))
    print("\nDB-halsa per bolag:")
    vc = defaultdict(list)
    for cid in scope:
        vc[health[cid][0]].append(cid)
    for v, ids in sorted(vc.items(), key=lambda t: -len(t[1])):
        print("  %-12s: %3d  %s" % (v, len(ids), ids))
    print("\nKlar: " + OUT_FILE)


if __name__ == "__main__":
    main()
