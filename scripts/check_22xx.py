"""Kolla 22XX-konton i existerande 202603 INL-filer (alla FI-bolag)."""
import sys
from pathlib import Path
import openpyxl
sys.stdout.reconfigure(encoding="utf-8")

R = Path(r"C:\Users\DidWac\Prosero Dropbox\Didrik Wachtmeister\Phoenix Foundation\April alla filer\Get testfiles\extracted\202604\Finland\output")

for p in sorted(R.glob("*_INL.xlsx")):
    wb = openpyxl.load_workbook(str(p), data_only=True)
    ws = wb.active
    hits = []
    for r in range(2, ws.max_row + 1):
        acc = ws.cell(row=r, column=1).value
        name = ws.cell(row=r, column=2).value
        v = ws.cell(row=r, column=3).value
        if not isinstance(v, (int, float)):
            continue
        try:
            acc_int = int(str(acc).strip())
        except (ValueError, TypeError):
            continue
        s = str(acc_int)
        prefix4 = int(s[:4]) if len(s) > 4 else acc_int
        if 2200 <= prefix4 <= 2299 and v != 0:
            hits.append((acc, str(name)[:35], v))
    if hits:
        print(f"\n{p.name}:")
        for h in hits:
            print(f"  {h[0]:>6}  {h[1]:<37}  {h[2]:>14,.2f}")
    wb.close()
