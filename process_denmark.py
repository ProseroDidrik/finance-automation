#!/usr/bin/env python3
"""
process_denmark.py  –  Danish Saldobalance XLSX → INL.xlsx

Kör från C:\\Users\\DidWac\\dev\\finance-automation\\ :
    py process_denmark.py               # kör alla bolag
    py process_denmark.py 229 242       # kör specifika bolag
    py process_denmark.py --dry-run     # visa utan att skriva

Vad scriptet gör:
  1. Läser Saldobalance XLSX per bolag, delar konton i IS/BS per bolagskonfiguration
  2. Stoppar vid "Nulkontrol"-sektionen (IS-konton visas annars dubbelt)
  3. Hoppar över bold+underline summerings-rader (bolag 178)
  4. Skriver {kod}_{Namn}_{YYYYMM}_INL.xlsx till Denmark/output/
  5. Flyttar källfiler till Denmark/Referens/

Bolagsspecifika IS/BS-gränser (4-siffrig kontoprefixnivå):
  178: IS = 0–4999,  BS = 5000+   (hoppa över bold+underline summary-rader)
  190: IS = 0–4999,  BS = 5000+
  216: IS = 0–9999   (enbart resultaträkning, inga BS-konton)
  229: IS = 0–4999,  BS = 5000+   (BS från YTD-fil: FEB-saldo − MAR-saldo)
  242: IS = 0–799,   BS = 800+
"""

import argparse
import re
import sys

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
from datetime import date
from pathlib import Path

from shared import (
    load_dotterbolag, move_to_referens_safe, save_inl_xlsx,
    load_config, log, DUPE_RE, glob_one, begin_run,
)

try:
    import openpyxl
except ImportError:
    sys.exit("Saknar openpyxl — kör:  py -m pip install openpyxl")

# ── Paths ──────────────────────────────────────────────────────────────────────
_BASE = Path(__file__).resolve().parent

GET_TESTFILES = Path(load_config()["base_path"])
DENMARK_DIR  = GET_TESTFILES / "extracted" / "Denmark"
OUTPUT_DIR   = DENMARK_DIR / "output"
REFERENS_DIR = DENMARK_DIR / "Referens"
DOTTERBOLAG  = _BASE / "_params" / "Dotterbolagslista.xlsx"

# ── Company definitions ────────────────────────────────────────────────────────
# is_max      : max 4-digit account prefix (inclusive) counted as IS/P&L
# bs_min      : min 4-digit account prefix (inclusive) counted as BS; None = no BS
# file_glob   : glob pattern to find the monthly Saldobalance file
# ytd_indicator: substring in the YTD file's name that distinguishes it from monthly
# use_ytd_bs  : if True, read BS from the YTD file instead of the monthly file
COMPANY_DEFS: dict[str, dict] = {
    "178": dict(
        is_max=4999, bs_min=5000, skip_formatting=True,
        file_glob="178_*.xlsx", exclude=[1999, 6140],
    ),
    "190": dict(
        is_max=4999, bs_min=5000, skip_formatting=False,
        file_glob="190_*.xlsx",
    ),
    # 216 SIKOM: enbart resultaträkning (RESULTATOPGØRELSE), inga BS-konton.
    # Kontona lagras som 5-siffriga strängar i källan ("01020"–"07XXX") och
    # bevaras byte-för-byte ut till INL.xlsx (se acc_raw i read_saldobalance).
    # Filen innehåller bold subtotaler ("I ALT", "RESULTAT FØR ...") som
    # skapar dubbelräkning — skip_bold=True filtrerar bort dem.
    "216": dict(
        is_max=9999, bs_min=None, skip_formatting=False, skip_bold=True,
        file_glob="216_*.xlsx",
    ),
    "229": dict(
        is_max=4999, bs_min=5000, skip_formatting=False,
        file_glob="229_Saldobalance*.xlsx", ytd_indicator="01-01", use_ytd_bs=True,
    ),
    "242": dict(
        is_max=799, bs_min=800, skip_formatting=False,
        file_glob="242_Saldobalance*.xlsx", ytd_indicator="01-01",
    ),
}

# ── Period helper ──────────────────────────────────────────────────────────────
def prev_month_period() -> str:
    today = date.today()
    if today.month == 1:
        return f"{today.year - 1}12"
    return f"{today.year}{today.month - 1:02d}"


# ── File discovery ─────────────────────────────────────────────────────────────
def _find_files(
    directory: Path, glob_pattern: str, ytd_indicator: str | None = None
) -> tuple[Path, Path | None]:
    """Return (main_file, ytd_file_or_None) for a company.

    When ytd_indicator is given, separates YTD from monthly files by the
    indicator substring. Falls back gracefully for January (no non-YTD file).
    Prefers non-duplicate files (unique_path copies have ` (2)+` suffix).
    Raises FileNotFoundError if no files match.
    """
    matches = sorted(directory.glob(glob_pattern))
    if not matches:
        raise FileNotFoundError(directory / glob_pattern)

    if ytd_indicator:
        ytd_matches  = [f for f in matches if ytd_indicator in f.name]
        main_matches = [f for f in matches if ytd_indicator not in f.name]
        # Take the latest alphabetically within each group — later months sort higher
        ytd  = ytd_matches[-1]  if ytd_matches  else None
        main = main_matches[-1] if main_matches else None
        if main is None and ytd is not None:
            main, ytd = ytd, None  # January edge case: only YTD file exists
        if main is None:
            raise FileNotFoundError(directory / glob_pattern)
        return main, ytd
    else:
        return matches[-1], None


# ── Amount parsing (Danish: . thousands, , decimal) ───────────────────────────
def parse_amount(val) -> float:
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return 0.0 if val != val else float(val)  # NaN → 0
    s = str(val).strip().replace("\xa0", "").replace(" ", "")
    if not s or s in ("-", "+"):
        return 0.0
    # "1.234,56" → remove . (thousands), , → .
    if re.match(r"^-?\d{1,3}(\.\d{3})+(,\d*)?$", s):
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


# ── Account normalisation ──────────────────────────────────────────────────────
def normalize4(account: int) -> int:
    """4-digit prefix of any account number (4, 5, 6 digits)."""
    s = str(account)
    return int(s[:4]) if len(s) > 4 else account




# ── Saldobalance column detection ─────────────────────────────────────────────
def find_columns(ws) -> tuple[int | None, dict]:
    """
    Scan the first 20 rows for the header row.
    Returns (header_row_index, cols_dict).
    cols_dict keys: acc, name, debet, kredit, saldo (any may be absent).
    """
    KEYWORDS = ("konto", "debet", "kredit", "saldo", "navn", "tekst", "beskrivelse", "periode")

    for row_idx, row in enumerate(ws.iter_rows(max_row=20)):
        texts = [
            str(c.value).strip().lower() if c.value is not None else ""
            for c in row
        ]
        if sum(1 for t in texts if any(kw in t for kw in KEYWORDS)) < 1:
            continue

        cols: dict = {}
        for j, t in enumerate(texts):
            if "acc"    not in cols and "konto" in t:
                cols["acc"] = j
            if "name"   not in cols and any(kw in t for kw in ("navn", "tekst", "beskrivelse")):
                cols["name"] = j
            if "debet"  not in cols and "debet" in t:
                cols["debet"] = j
            if "kredit" not in cols and "kredit" in t:
                cols["kredit"] = j
            if "saldo"  not in cols and ("saldo" in t or "periode" in t):
                cols["saldo"] = j

        # Require saldo > 0 when it's the only useful column found, to avoid
        # matching title rows like "Periode: 01-03-2026" at col 0 (which would
        # clash with the default acc=0).
        has_accs  = "acc" in cols or ("debet" in cols and "kredit" in cols)
        has_saldo = "saldo" in cols and cols["saldo"] > 0
        if has_accs or has_saldo:
            return row_idx, cols

    return None, {}


# ── Saldobalance reader ────────────────────────────────────────────────────────
def read_saldobalance(
    filepath: Path,
    is_max_4d: int,
    bs_min_4d: int | None,
    skip_formatting: bool = False,
    skip_bold: bool = False,
    exclude: list[int] | None = None,
) -> tuple[list, list]:
    """
    Returns (is_rows, bs_rows), each a list of (account_code_str, name_str, amount_float).
    account_code_str bevarar källcellens strängform (inkl. inledande nollor som i
    216:s "0XXXX"). Internt används en int-tolkning (acc_int) för normalize4,
    seen_accs-dedup och exclude-check.

    Amounts use kredit-debet sign convention (income positive, costs negative).
    Stops at a Nulkontrol ACCOUNT ROW (account present + "nulkontrol" in any cell).
    Section-header rows (empty account) named "Nulkontrol" are skipped, not treated
    as stop signals — this handles månedsopdelt files (e.g. 229) where the file has
    a second IS+BS section under a "Nulkontrol" heading.
    Duplicate account numbers are skipped to avoid double-counting from that second
    section.
    If skip_formatting=True, rows where any of the first 3 cells is bold+underline
    are skipped (handles 178's summary/total rows).
    """
    wb = openpyxl.load_workbook(str(filepath), data_only=True)
    ws = wb.active

    header_row_idx, cols = find_columns(ws)
    if header_row_idx is None:
        wb.close()
        raise ValueError(f"Kunde inte hitta header-rad i {filepath.name}")

    col_acc    = cols.get("acc",   0)
    col_name   = cols.get("name",  1 if cols.get("acc", 0) == 0 else 0)
    col_debet  = cols.get("debet")
    col_kredit = cols.get("kredit")
    col_saldo  = cols.get("saldo")

    is_rows: list = []
    bs_rows: list = []
    seen_accs: set[int] = set()
    in_nulkontrol = False  # True once we pass the Nulkontrol section header

    for row in ws.iter_rows(min_row=header_row_idx + 2):
        if col_acc >= len(row):
            continue
        acc_cell = row[col_acc]

        # Detect the Nulkontrol section-header row (empty account containing
        # "Nulkontrol"): e.g. row 132 in månedsopdelt files.  After this point
        # IS accounts have YTD-cumulative values, not monthly movements.
        if acc_cell.value is None or str(acc_cell.value).strip() == "":
            if any("nulkontrol" in str(c.value).lower() for c in row if c.value is not None):
                in_nulkontrol = True
            continue

        # Bevara källcellens strängform (incl. inledande nollor) när cellen är
        # en ren digit-sträng — t.ex. 216:s "01020". För numeriska celler eller
        # exotiska representationer faller vi tillbaka till int-formen.
        # acc_int används internt för normalize4/seen_accs/exclude; acc_raw
        # skrivs till INL.xlsx (→ fact_balances.account_code).
        try:
            acc_int = int(float(str(acc_cell.value).strip()))
        except (ValueError, TypeError):
            continue
        acc_raw = str(acc_int)
        if isinstance(acc_cell.value, str):
            s = acc_cell.value.strip()
            if s.isdigit() and int(s) == acc_int:
                acc_raw = s

        # Stop on a numeric "Nulkontrol" account row (e.g. acc_int=9990 in 178).
        if any("nulkontrol" in str(c.value).lower() for c in row if c.value is not None):
            break

        # Skip accounts already processed (månedsopdelt double-section files)
        if acc_int in seen_accs:
            continue
        seen_accs.add(acc_int)

        if exclude and acc_int in exclude:
            continue

        # In the Nulkontrol section, IS accounts carry YTD values — skip them.
        # BS accounts (after the YTD-IS section) are still valid to collect.
        acc4 = normalize4(acc_int)
        if in_nulkontrol and acc4 <= is_max_4d:
            continue

        # Skip bold+underline summary rows (178)
        if skip_formatting:
            skip = False
            for cell in row[:3]:
                f = cell.font
                if f and f.bold and f.underline:
                    skip = True
                    break
            if skip:
                continue

        # Skip bold-only summary rows (216: "I ALT"/"RESULTAT FØR..." subtotaler)
        if skip_bold:
            skip = False
            for cell in row[:3]:
                f = cell.font
                if f and f.bold:
                    skip = True
                    break
            if skip:
                continue

        # Name
        name = ""
        if col_name < len(row) and row[col_name].value is not None:
            name = str(row[col_name].value).strip()

        # Amount: sign convention = kredit − debet (income positive, costs negative).
        # Saldo/Periode columns already store debet−kredit, so negate them.
        if col_saldo is not None and col_saldo < len(row):
            amt = -parse_amount(row[col_saldo].value)
        elif col_debet is not None and col_kredit is not None:
            d = parse_amount(row[col_debet].value  if col_debet  < len(row) else None)
            k = parse_amount(row[col_kredit].value if col_kredit < len(row) else None)
            amt = k - d
        elif col_debet is not None and col_debet < len(row):
            amt = -parse_amount(row[col_debet].value)
        else:
            continue

        amt = round(amt, 2)
        if amt == 0.0:
            continue

        if acc4 <= is_max_4d:
            is_rows.append((acc_raw, name, amt))
        elif bs_min_4d is not None and acc4 >= bs_min_4d:
            bs_rows.append((acc_raw, name, amt))

    wb.close()
    return is_rows, bs_rows


# ── YTD BS reader (månedsopdelt) ───────────────────────────────────────────────
def read_bs_from_ytd(filepath: Path, bs_min_4d: int) -> list:
    """
    Reads BS accounts from a YTD månedsopdelt file.
    Finds the last two month columns (e.g. FEB 2026, MAR 2026) and returns
    BS rows as (acc, name, prev_val − curr_val) in kredit-debet convention.
    """
    wb = openpyxl.load_workbook(str(filepath), data_only=True)
    ws = wb.active

    prev_col: int | None = None
    curr_col: int | None = None
    header_row_idx: int | None = None

    for row_idx, row in enumerate(ws.iter_rows(max_row=20)):
        month_cols = []
        for j, cell in enumerate(row):
            if cell.value is None:
                continue
            t = str(cell.value).strip().lower()
            # Match "jan 2026", "feb2026", "mar 2026" etc.
            if re.match(r"^[a-zæøå]{3}\s*\d{4}$", t):
                month_cols.append(j)
        if len(month_cols) >= 2:
            prev_col = month_cols[-2]
            curr_col = month_cols[-1]
            header_row_idx = row_idx
            break

    if header_row_idx is None or prev_col is None or curr_col is None:
        wb.close()
        raise ValueError(f"Fann inga månadskolumner i {filepath.name}")

    # Detect acc/name columns from same or neighbouring rows
    col_acc, col_name = 0, 2
    for row_idx2, row in enumerate(ws.iter_rows(max_row=header_row_idx + 1)):
        texts = [str(c.value).strip().lower() if c.value is not None else "" for c in row]
        for j, t in enumerate(texts):
            if "konto" in t:
                col_acc = j
            if any(kw in t for kw in ("navn", "tekst", "beskrivelse")):
                col_name = j

    bs_rows: list = []
    seen_accs: set[int] = set()

    for row in ws.iter_rows(min_row=header_row_idx + 2):
        if col_acc >= len(row):
            continue
        acc_cell = row[col_acc]
        if acc_cell.value is None:
            continue
        # Bevara källcellens strängform (incl. inledande nollor) när cellen är
        # en ren digit-sträng. För numeriska celler faller vi tillbaka till
        # int-formen. acc_int används för normalize4/seen_accs; acc_raw skrivs ut.
        try:
            acc_int = int(float(str(acc_cell.value).strip()))
        except (ValueError, TypeError):
            continue
        acc_raw = str(acc_int)
        if isinstance(acc_cell.value, str):
            s = acc_cell.value.strip()
            if s.isdigit() and int(s) == acc_int:
                acc_raw = s

        if any("nulkontrol" in str(c.value).lower() for c in row if c.value is not None):
            break

        if acc_int in seen_accs:
            continue
        seen_accs.add(acc_int)

        acc4 = normalize4(acc_int)
        if acc4 < bs_min_4d:
            continue

        name = ""
        if col_name < len(row) and row[col_name].value is not None:
            name = str(row[col_name].value).strip()

        prev_val = parse_amount(row[prev_col].value if prev_col < len(row) else None)
        curr_val = parse_amount(row[curr_col].value if curr_col < len(row) else None)
        amt = round(prev_val - curr_val, 2)
        if amt == 0.0:
            continue

        bs_rows.append((acc_raw, name, amt))

    wb.close()
    return bs_rows


# ── Referens move ──────────────────────────────────────────────────────────────
def move_to_referens(filename: str, dry_run: bool) -> None:
    src = DENMARK_DIR / filename
    if not src.exists():
        return
    move_to_referens_safe(src, REFERENS_DIR, dry_run)


def archive_by_prefix(prefix: str, dry_run: bool) -> None:
    """Flytta alla {prefix}_* filer till Referens/ utan att skapa INL.xlsx."""
    files = sorted(f for f in DENMARK_DIR.glob(f"{prefix}_*") if f.is_file())
    if not files:
        return
    log("INFO", prefix, f"Arkiverar {len(files)} fil(er) → Referens/")
    for f in files:
        move_to_referens(f.name, dry_run)


# ── Process one saldobalance company ──────────────────────────────────────────
def process_company(
    code: str,
    friendly: str,
    period: str,
    filepath: Path,
    is_max_4d: int,
    bs_min_4d: int | None,
    skip_formatting: bool,
    dry_run: bool,
    exclude: list[int] | None = None,
    ytd_filepath: Path | None = None,
    skip_bold: bool = False,
) -> str:
    log("INFO", code, f"{friendly}  Fil: {filepath.name}")

    if not filepath.exists():
        log("SKIP", code, "Filen saknas (redan i Referens?)")
        return "skip"

    try:
        is_rows, bs_rows = read_saldobalance(
            filepath, is_max_4d, bs_min_4d,
            skip_formatting=skip_formatting, skip_bold=skip_bold, exclude=exclude,
        )
        if ytd_filepath is not None and bs_min_4d is not None:
            if ytd_filepath.exists():
                bs_rows = read_bs_from_ytd(ytd_filepath, bs_min_4d)
            else:
                log("WARN", code, f"YTD-fil saknas: {ytd_filepath.name}")
    except Exception as e:
        log("ERROR", code, f"Läsfel: {e}")
        return "error"

    total = sum(r[2] for r in is_rows + bs_rows)
    is_warn = abs(total) >= 1.0
    check = "OK" if not is_warn else f"KONTROLLERA ({total:.2f})"
    log("INFO", code, f"IS={len(is_rows)}, BS={len(bs_rows)}  Summa={total:.4f}  {check}")

    out_name = f"{code}_{friendly}_{period}_INL.xlsx"
    out_path = OUTPUT_DIR / out_name

    if not dry_run:
        save_inl_xlsx(is_rows, bs_rows, out_path)
    dry_prefix = "[DRY] " if dry_run else ""
    log("WARN" if is_warn else "OK", code, f"{dry_prefix}Sparad: {out_name}")

    return "warn" if is_warn else "ok"


# ── Main ───────────────────────────────────────────────────────────────────────
def main() -> None:
    global DENMARK_DIR, OUTPUT_DIR, REFERENS_DIR

    parser = argparse.ArgumentParser(
        description="Bearbeta danska Saldobalance-filer → INL.xlsx"
    )
    parser.add_argument(
        "codes", nargs="*",
        help="Bolagskoder att köra (standard: alla). Ex: py process_denmark.py 229 242",
    )
    parser.add_argument(
        "--dry-run", "-n", action="store_true",
        help="Visa vad som skulle hända utan att skriva några filer",
    )
    parser.add_argument(
        "--period", "-p", metavar="YYYYMM", default=None,
        help="Period att köra (t.ex. 202604). Standard: föregående månad.",
    )
    args = parser.parse_args()

    if args.period:
        DENMARK_DIR  = GET_TESTFILES / "extracted" / args.period / "Denmark"
        OUTPUT_DIR   = DENMARK_DIR / "output"
        REFERENS_DIR = DENMARK_DIR / "Referens"

    period = args.period or prev_month_period()
    begin_run("process_denmark", period)
    dry_label = "  [DRY RUN]" if args.dry_run else ""
    log("START", "process_denmark.py", f"period {period}{dry_label}")

    if not DENMARK_DIR.exists():
        sys.exit(f"[ERROR]  Denmark-mappen saknas: {DENMARK_DIR}")
    if not DOTTERBOLAG.exists():
        sys.exit(f"[ERROR]  Dotterbolagslistan saknas: {DOTTERBOLAG}")

    friendlies = load_dotterbolag(DOTTERBOLAG)

    if not args.dry_run:
        OUTPUT_DIR.mkdir(exist_ok=True)
        REFERENS_DIR.mkdir(exist_ok=True)

    all_codes = sorted(COMPANY_DEFS.keys())
    codes_to_run = args.codes if args.codes else all_codes
    stats: dict[str, int] = {"ok": 0, "warn": 0, "skip": 0, "error": 0}

    for code in codes_to_run:
        if code not in COMPANY_DEFS:
            log("ERROR", code, "Okänd bolagskod")
            stats["error"] += 1
            continue

        cfg      = COMPANY_DEFS[code]
        friendly = friendlies.get(int(code), f"Bolag{code}")

        try:
            filepath, ytd_filepath = _find_files(
                DENMARK_DIR, cfg["file_glob"], cfg.get("ytd_indicator")
            )
        except FileNotFoundError:
            log("SKIP", code, "Källfil saknas (redan i Referens?)")
            stats["skip"] += 1
            continue

        status = process_company(
            code=code,
            friendly=friendly,
            period=period,
            filepath=filepath,
            is_max_4d=cfg["is_max"],
            bs_min_4d=cfg["bs_min"],
            skip_formatting=cfg["skip_formatting"],
            dry_run=args.dry_run,
            exclude=cfg.get("exclude"),
            ytd_filepath=ytd_filepath if cfg.get("use_ytd_bs") else None,
            skip_bold=cfg.get("skip_bold", False),
        )
        stats[status] = stats.get(status, 0) + 1

        # Move all remaining {code}_* source files to Referens
        if not args.dry_run:
            REFERENS_DIR.mkdir(exist_ok=True)
        for f in sorted(DENMARK_DIR.glob(f"{code}_*")):
            if f.is_file():
                move_to_referens(f.name, args.dry_run)

    if not args.codes or "54" in args.codes:
        archive_by_prefix("054", args.dry_run)

    log("DONE", "process_denmark.py",
        f"{stats['ok']} OK  {stats['warn']} WARN  {stats['skip']} SKIP  {stats['error']} ERROR")


if __name__ == "__main__":
    main()
