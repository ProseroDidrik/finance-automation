"""render_xlsx.py — bygg Excel-rapporten ur dash + validation.

Anpassad från Cowork v13:s build_xlsx.py. Ändringar:
  * render(dash, validation, fyo, out_path) — paths som argument (ingen /sessions/-path).
  * BUGGFIX: helår-2025-värden läses ur c['periods']['202512'] (build_ru_aggregat
    sätter ALDRIG c['full_year_2025'] — Cowork-koden läste ett fält som inte fanns).
  * BUGGFIX: proxy-test via flaggan FULL_YEAR_PROXY_2025 (ej c['full_year_only_2025']).
  * Aaro-fliken uppskjuten till separat PR (dashboards/ytd_nyckeltal/aaro/).
  * Tål validation=None (bygger då bara per-bolag + Metod, utan facit-dots).
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

TITLE = Font(name='Arial', size=14, bold=True, color='1F2937')
H2 = Font(name='Arial', size=12, bold=True, color='2563EB')
HDR = Font(name='Arial', size=10, bold=True, color='FFFFFF')
HDR_FILL = PatternFill('solid', start_color='4B5563')
NORMAL = Font(name='Arial', size=10)
MUTED = Font(name='Arial', size=9, color='6B7280', italic=True)
THIN = Side(style='thin', color='E5E7EB')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
PROXY_FILL = PatternFill('solid', start_color='EDE9FE')   # lila = helår-proxy
SAFTVER_FILL = PatternFill('solid', start_color='D1FAE5')  # grön = SAFT_VER-syntes

DISPLAY_TGS = ['Total Sales', 'Total Direct Cost', 'Bruttovinst', 'Personnel',
               'Consultants', 'Other External Costs', 'Premises',
               'Transportation', 'Depreciation', 'Justerad EBITDA']
NUMFMT = '#,##0.0;(#,##0.0);-'


def emoji_status(diff_pct, is_proxy=False):
    if is_proxy:
        return '🔸'
    if diff_pct is None:
        return '⚪'
    a = abs(diff_pct)
    return '🟢' if a < 0.01 else ('🟡' if a < 0.05 else '🔴')


def _is_proxy(c):
    return 'FULL_YEAR_PROXY_2025' in (c.get('flags') or [])


def _sheet_sammanfattning(wb, dash, val):
    ws = wb.create_sheet('Sammanfattning')
    fyo = set(val.get('full_year_only_cids', []))
    ws['A1'] = 'Nyckeltal Prosero — YTD apr 2026 vs 2025'
    ws['A1'].font = TITLE
    ws['A2'] = (f'{len(fyo)} bolag har endast helårs-SAFT för 2025 (🔸). Övriga NO-bolag '
                'har syntetiserad YTD 2025 ur SAF-T-journalen (SAFT_VER).')
    ws['A2'].font = MUTED

    headers = ['Top Group', 'Facit (MSEK)', 'Warehouse (MSEK)', 'Diff (MSEK)', 'Diff %', 'Status']

    def konc_block(title, facit_map, wh_map, r):
        ws.cell(row=r, column=1, value=title).font = H2
        r += 1
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=r, column=ci, value=h)
            c.font = HDR; c.fill = HDR_FILL; c.border = BORDER
        r += 1
        for tg in DISPLAY_TGS:
            fac = (facit_map.get(tg, 0) or 0) / 1e6
            wh = (wh_map.get(tg, 0) or 0) / 1e6
            ws.cell(row=r, column=1, value=tg).font = NORMAL
            ws.cell(row=r, column=2, value=round(fac, 3))
            ws.cell(row=r, column=3, value=round(wh, 3))
            ws.cell(row=r, column=4, value=f'=B{r}-C{r}')
            ws.cell(row=r, column=5, value=f'=IF(B{r}=0,"",D{r}/B{r})')
            ws.cell(row=r, column=5).number_format = '0.0%'
            ws.cell(row=r, column=6,
                    value=f'=IF(B{r}=0,"—",IF(ABS(E{r})<0.05,"OK",IF(ABS(E{r})<0.2,"Avvik.","Stor diff")))')
            for cc in range(2, 5):
                ws.cell(row=r, column=cc).number_format = NUMFMT
            r += 1
        return r + 1

    r = 4
    r = konc_block('KONCERNTOTAL — YTD 2026', val['utfall_facit'], val['utfall_wh'], r)

    # 2025 koncern-WH: summa per RU 202504, exkl. helår-proxy (saknar månadsdata).
    wh25 = {}
    for tg in DISPLAY_TGS:
        if tg == 'Bruttovinst':
            wh25[tg] = sum(
                ((c.get('kpis', {}).get('202504', {}) or {}).get('gross_profit', 0) or 0)
                for c in dash['companies'] if not _is_proxy(c))
        elif tg == 'Justerad EBITDA':
            wh25[tg] = 0
        else:
            from validate import TG_KPI
            kpi = TG_KPI.get(tg)
            wh25[tg] = sum(
                abs((c.get('kpis', {}).get('202504', {}) or {}).get(kpi, 0) or 0)
                for c in dash['companies'] if not _is_proxy(c))
    r = konc_block('KONCERNTOTAL — YTD 2025', val['utfall_facit_25'], wh25, r)

    ws.cell(row=r, column=1,
            value=f'Notera: 2025-koncernen exkluderar {len(fyo)} helår-proxy-bolag (visas i Validering 2025).').font = MUTED
    for col, w in zip('ABCDEF', [32, 15, 16, 15, 12, 14]):
        ws.column_dimensions[col].width = w


def _sheet_per_bolag(wb, dash):
    ws = wb.create_sheet('Nyckeltal per bolag')
    ws['A1'] = 'Nyckeltal per bolag — färgkod 2026 + 2025'
    ws['A1'].font = TITLE
    ws['A2'] = ('Status 2025: 🔸 = helår-proxy (SAFT 202512). Lila rad = helår-proxy, '
                'grön rad = NO-bolag med syntetiserad YTD (SAFT_VER).')
    ws['A2'].font = MUTED
    hdrs = ['Status 26', 'Status 25', 'Bolag', 'Land', 'Typ', 'Oms 2026 YTD (KSEK)', 'Oms 2025 YTD',
            'Δ Oms %', 'BV 2026 (KSEK)', 'BV 2025', 'Δ BV %', 'Pers 2026', 'Pers 2025', 'Δ Pers %',
            'Konsult 2026', 'Konsult 2025', 'Δ Konsult %', 'FTE apr-26', 'Δ FTE', 'Oms/FTE (KSEK)']
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=4, column=ci, value=h)
        c.font = HDR; c.fill = HDR_FILL; c.border = BORDER

    companies = sorted(dash['companies'],
                       key=lambda c: -((c.get('kpis', {}).get('202604', {}) or {}).get('sales') or 0))
    r = 5
    for c in companies:
        f = c.get('facit') or {}
        is_proxy = _is_proxy(c)
        has_saftver = (not is_proxy) and any(
            (c.get('periods', {}).get(p, {}) or {}) for p in ('202504',)) and c.get('country') == 'Norway'
        s26 = emoji_status(f.get('ts_diff_pct'))
        s25 = emoji_status(f.get('ts_2025_diff_pct'), is_proxy=is_proxy)
        k25 = c.get('kpis', {}).get('202504', {}) or {}
        k26 = c.get('kpis', {}).get('202604', {}) or {}

        def kval(d, k):
            v = d.get(k)
            return None if v is None else round(v / 1000, 1)

        ws.cell(row=r, column=1, value=s26)
        ws.cell(row=r, column=2, value=s25)
        ws.cell(row=r, column=3, value=c['name'])
        ws.cell(row=r, column=4, value=c.get('country', ''))
        ws.cell(row=r, column=5, value=c.get('kind', ''))
        ws.cell(row=r, column=6, value=kval(k26, 'sales'))
        ws.cell(row=r, column=7, value=kval(k25, 'sales'))
        ws.cell(row=r, column=8, value=f'=IF(G{r}=0,"",(F{r}-G{r})/G{r})')
        ws.cell(row=r, column=9, value=kval(k26, 'gross_profit'))
        ws.cell(row=r, column=10, value=kval(k25, 'gross_profit'))
        ws.cell(row=r, column=11, value=f'=IF(J{r}=0,"",(I{r}-J{r})/J{r})')
        ws.cell(row=r, column=12, value=kval(k26, 'personnel'))
        ws.cell(row=r, column=13, value=kval(k25, 'personnel'))
        ws.cell(row=r, column=14, value=f'=IF(M{r}=0,"",(L{r}-M{r})/M{r})')
        ws.cell(row=r, column=15, value=kval(k26, 'consultants'))
        ws.cell(row=r, column=16, value=kval(k25, 'consultants'))
        ws.cell(row=r, column=17, value=f'=IF(P{r}=0,"",(O{r}-P{r})/P{r})')
        fte26, fte25 = c.get('fte_apr_2026'), c.get('fte_apr_2025')
        fte_delta = (fte26 - fte25) if (fte26 is not None and fte25 is not None) else None
        ws.cell(row=r, column=18, value=fte26)
        ws.cell(row=r, column=19, value=fte_delta)
        ws.cell(row=r, column=20, value=f'=IF(R{r}>0,F{r}/R{r},"")')
        ws.cell(row=r, column=20).number_format = '#,##0.0'
        for col in (8, 11, 14, 17):
            ws.cell(row=r, column=col).number_format = '0.0%'
        for col in (6, 7, 9, 10, 12, 13, 15, 16):
            ws.cell(row=r, column=col).number_format = NUMFMT
        if is_proxy:
            for col in range(1, 21):
                ws.cell(row=r, column=col).fill = PROXY_FILL
        elif has_saftver:
            for col in (2,):  # markera bara 2025-status-cellen grön
                ws.cell(row=r, column=col).fill = SAFTVER_FILL
        r += 1

    widths = [9, 9, 24, 11, 13, 16, 13, 10, 14, 12, 10, 12, 12, 10, 13, 13, 12, 11, 8, 14]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'F5'


def _sheet_validering_2026(wb, val):
    ws = wb.create_sheet('Validering 2026')
    ws['A1'] = 'Validering YTD apr 2026'
    ws['A1'].font = TITLE
    hdrs = ['Bolag', 'Land', 'Typ', 'Sales facit', 'Sales WH', 'Sales Δ%',
            'Pers facit', 'Pers WH', 'Pers Δ%', 'BV facit', 'BV WH', 'BV Δ%',
            'Konsult facit', 'Konsult WH', 'Konsult Δ%']
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.font = HDR; c.fill = HDR_FILL; c.border = BORDER
    r = 4
    for row in val['rows']:
        ws.cell(row=r, column=1, value=row['mercur_name'])
        ws.cell(row=r, column=2, value=row.get('country', ''))
        ws.cell(row=r, column=3, value=row.get('kind', ''))
        for ci, tg in enumerate(['Total Sales', 'Personnel', 'Bruttovinst', 'Consultants']):
            base = 4 + ci * 3
            cell = row.get(tg) or {}
            ws.cell(row=r, column=base, value=round((cell.get('facit') or 0) / 1e6, 3))
            ws.cell(row=r, column=base + 1, value=round((cell.get('wh') or 0) / 1e6, 3))
            cl = get_column_letter(base)
            ws.cell(row=r, column=base + 2,
                    value=f'=IF({cl}{r}=0,"",({cl}{r}-{get_column_letter(base+1)}{r})/{cl}{r})')
            ws.cell(row=r, column=base + 2).number_format = '0.0%'
            ws.cell(row=r, column=base).number_format = NUMFMT
            ws.cell(row=r, column=base + 1).number_format = NUMFMT
        r += 1
    for ci, w in enumerate([28, 11, 13] + [11, 11, 9] * 4, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'A4'


def _sheet_validering_2025(wb, dash, val):
    ws = wb.create_sheet('Validering 2025')
    fyo = set(val.get('full_year_only_cids', []))
    ws['A1'] = 'Validering YTD apr 2025 (facit från Mercur "Utfall fg. år")'
    ws['A1'].font = TITLE
    ws['A2'] = f'{len(fyo)} bolag (🔸) har endast helårs-SAFT — jämförs helår 2025 vs Mercur, ej YTD apr.'
    ws['A2'].font = MUTED
    hdrs = ['Status', 'Bolag', 'Land', 'Typ', 'Facit (MSEK)', 'Warehouse (MSEK)', 'Diff (MSEK)', 'Diff %', 'Källa']
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=4, column=ci, value=h)
        c.font = HDR; c.fill = HDR_FILL; c.border = BORDER
    by_cid = {c['company_id']: c for c in dash['companies']}
    r = 5
    for row in val['rows']:
        comp = by_cid.get(row['reporting_cid'])
        is_proxy = _is_proxy(comp) if comp else (row['reporting_cid'] in fyo)
        if is_proxy and comp:
            # BUGGFIX: helår ur periods['202512'] (Cowork läste c['full_year_2025'] = saknades).
            wh_2025 = (comp.get('periods', {}).get('202512', {}) or {}).get('Total Sales', 0) / 1e6
            source, status_emoji = 'SAFT 202512 helår', '🔸'
        else:
            f = (comp.get('facit') if comp else {}) or {}
            wh_2025 = (f.get('ts_2025_wh') or 0) / 1e6
            source, status_emoji = 'YTD apr 2025', emoji_status(f.get('ts_2025_diff_pct'))
        fac = (row.get('Total Sales') or {}).get('facit', 0) / 1e6
        ws.cell(row=r, column=1, value=status_emoji)
        ws.cell(row=r, column=2, value=row['mercur_name'])
        ws.cell(row=r, column=3, value=row.get('country', ''))
        ws.cell(row=r, column=4, value=row.get('kind', ''))
        ws.cell(row=r, column=5, value=round(fac, 3))
        ws.cell(row=r, column=6, value=round(wh_2025, 3))
        ws.cell(row=r, column=7, value=f'=E{r}-F{r}')
        ws.cell(row=r, column=8, value=f'=IF(E{r}=0,"",G{r}/E{r})')
        ws.cell(row=r, column=8).number_format = '0.0%'
        ws.cell(row=r, column=9, value=source)
        for col in (5, 6, 7):
            ws.cell(row=r, column=col).number_format = NUMFMT
        if is_proxy:
            for col in range(1, 10):
                ws.cell(row=r, column=col).fill = PROXY_FILL
        r += 1
    for ci, w in enumerate([8, 28, 11, 13, 13, 14, 12, 10, 20], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'A5'


def _sheet_metod(wb, dash, val):
    ws = wb.create_sheet('Metod')
    fyo = set((val or {}).get('full_year_only_cids', []))
    lines = [
        ('Metod — YTD-nyckeltalsdashboard (v14)', TITLE),
        ('', None),
        ('DATA-KÄLLOR', H2),
        ('  Warehouse: finance-warehouse.fact_balances (Postgres, read-only via db.py)', NORMAL),
        ('  Best source per land: SE: SIE_PSALDO→SIE_VER→SIE→IMP. NO: SAFT→SAFT_VER→IMP.', NORMAL),
        ('  FI/DK/DE/CENTR: IMP. Period-semantik: SIE_PSALDO/IMP monthly, SIE/SIE_VER/SAFT/SAFT_VER ytd.', NORMAL),
        ('  Tecken-normalisering: abs() per (cid, period, top_group). dim_company.currency läses rakt', NORMAL),
        ('  (CENTR-valutorna rättade i prod — ingen override i koden längre).', NORMAL),
        ('  Facit: Mercur Resultaträkning (20).xlsx — 2026 = "Utfall", 2025 = "Utfall fg. år".', NORMAL),
        ('', None),
        ('SAFT_VER — syntetiserad YTD 2025 för NO', H2),
        (f'  NO-bolag som bara levererar helårs-SAF-T (202512) får YTD jan..nov syntetiserad ur', NORMAL),
        ('  SAF-T-journalen (source_kind SAFT_VER). Det ger riktig YoY mot 2026 i stället för proxy.', NORMAL),
        (f'  Kvar som helår-proxy (🔸): {len(fyo)} bolag vars journal inte når tillbaka (t.ex. cid 233', NORMAL),
        ('  Stavanger — journal bara december; cid 157 Hemer — journal från september).', NORMAL),
        ('', None),
        ('PERIODER', H2),
        ('  202604 = YTD apr 2026 · 202504 = YTD apr 2025 · 202512 = helår 2025 (proxy)', NORMAL),
        ('', None),
        ('STATUS-PRICKAR', H2),
        ('  🟢 ≤1% · 🟡 1-5% · 🔴 >5% · ⚪ ej i Mercur · 🔸 helår-proxy (ej jämförbar YTD apr)', NORMAL),
        ('', None),
        ('FX-KURSER', H2),
    ]
    r = 1
    for txt, fnt in lines:
        c = ws.cell(row=r, column=1, value=txt)
        if fnt is not None:
            c.font = fnt
        r += 1
    for per, fx in dash['meta']['fx_assumptions'].items():
        ws.cell(row=r, column=1, value=f"  {per}: " + ', '.join(f'{k}={v}' for k, v in fx.items()))
        r += 1
    ws.column_dimensions['A'].width = 130


def render(dash, validation, fyo, out_path):
    """Bygg arbetsboken och spara till out_path. validation=None → reducerad bok."""
    wb = Workbook()
    wb.remove(wb.active)
    if validation:
        _sheet_sammanfattning(wb, dash, validation)
    _sheet_per_bolag(wb, dash)
    if validation:
        _sheet_validering_2026(wb, validation)
        _sheet_validering_2025(wb, dash, validation)
    _sheet_metod(wb, dash, validation)
    wb.save(out_path)
    return wb
