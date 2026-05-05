"""Ladda valutakurser (NOK, DKK, EUR → SEK) till dim_exchange_rate.

Källa: _params/Valutakurser.xlsx
  Sheet "Genomsnittskurs"  → rate_type='avg'
  Sheet "Constant currency" → rate_type='constant'

Layout per sheet:
  Rad 1: "Variant: ..." (rate_type-etikett, läses ej)
  Rad 2: None, "Dec 2019", "Jan 2020", ...   (månadsrubriker, kolumn B+)
  Rad 3+: "  NOK", val1, val2, ...
           "  DKK", val1, val2, ...
           "  EUR", val1, val2, ...

Körning:
  py load_exchange_rates.py          # laddar alla perioder
  py load_exchange_rates.py --dry-run
"""
from __future__ import annotations

import argparse
from datetime import datetime
from pathlib import Path

import openpyxl

import db
from shared import begin_run, load_config, log

EXCHANGE_FILE = Path(__file__).resolve().parent / "_params" / "Valutakurser.xlsx"
SOURCE_FILE = "_params/Valutakurser.xlsx"

MONTH_SV = {
    "Jan": "01", "Feb": "02", "Mar": "03", "Apr": "04",
    "Maj": "05", "Jun": "06", "Jul": "07", "Aug": "08",
    "Sep": "09", "Okt": "10", "Nov": "11", "Dec": "12",
}

SHEET_RATE_TYPE = {
    "Genomsnittskurs": "avg",
    "Constant currency": "constant",
}


def parse_period(header: str) -> str | None:
    """'Jan 2020' → '202001'. Returnerar None om format okänt."""
    if not header or not isinstance(header, str):
        return None
    parts = header.strip().split()
    if len(parts) != 2:
        return None
    month_sv, year_str = parts
    month = MONTH_SV.get(month_sv)
    if not month or not year_str.isdigit() or len(year_str) != 4:
        return None
    return f"{year_str}{month}"


def parse_sheet(ws) -> list[tuple[str, str, float]]:
    """Returnerar lista av (period, currency, rate) från ett sheet.

    Förväntar rad 2 = månadsrubriker (B2+), rad 3+ = valutor.
    """
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    # Rad 2 (index 1): period-rubriker i kolumn B+
    header_row = rows[1]
    periods: list[str | None] = []
    for cell in header_row[1:]:  # skip col A
        periods.append(parse_period(cell) if cell else None)

    results: list[tuple[str, str, float]] = []
    for row in rows[2:]:
        if not row or row[0] is None:
            continue
        currency = str(row[0]).strip().upper()
        if currency not in ("NOK", "DKK", "EUR"):
            continue
        for i, val in enumerate(row[1:]):
            if i >= len(periods):
                break
            period = periods[i]
            if period is None or val is None:
                continue
            try:
                rate = float(val)
            except (TypeError, ValueError):
                continue
            if rate > 0:
                results.append((period, currency, rate))

    return results


def load_exchange_rates(con, *, dry_run: bool = False) -> dict[str, int]:
    wb = openpyxl.load_workbook(str(EXCHANGE_FILE), data_only=True, read_only=True)
    now = datetime.now()
    counts: dict[str, int] = {}

    for sheet_name, rate_type in SHEET_RATE_TYPE.items():
        if sheet_name not in wb.sheetnames:
            log("WARN", "valuta", f"Sheet '{sheet_name}' saknas i {EXCHANGE_FILE.name}")
            continue

        ws = wb[sheet_name]
        rows = parse_sheet(ws)
        if not rows:
            log("WARN", sheet_name, "Inga kurser hittades")
            continue

        periods = sorted({r[0] for r in rows})
        log("INFO", sheet_name,
            f"rate_type={rate_type}  {len(rows)} kurser  {len(periods)} perioder "
            f"({periods[0]}–{periods[-1]})")

        if dry_run:
            counts[rate_type] = len(rows)
            continue

        db.sync_dim_period(con, periods)

        con.execute("BEGIN")
        try:
            con.execute(
                "DELETE FROM dim_exchange_rate WHERE rate_type = ?", [rate_type]
            )
            con.executemany(
                """INSERT INTO dim_exchange_rate (period, currency, rate_type, rate, loaded_at)
                   VALUES (?, ?, ?, ?, ?)""",
                [(p, c, rate_type, r, now) for p, c, r in rows],
            )
            con.execute("COMMIT")
        except Exception as e:
            con.execute("ROLLBACK")
            log("ERROR", sheet_name, f"DB-fel: {e}")
            counts[rate_type] = 0
            continue

        counts[rate_type] = len(rows)
        log("OK", sheet_name, f"{len(rows)} kurser laddade  rate_type={rate_type}")

    wb.close()
    return counts


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Ladda valutakurser från Valutakurser.xlsx till dim_exchange_rate."
    )
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    begin_run("load_exchange_rates.py", "ALL")
    log("START", "load_exchange_rates.py",
        f"fil={EXCHANGE_FILE.name}  dry_run={args.dry_run}")

    if not EXCHANGE_FILE.exists():
        log("ERROR", "load_exchange_rates.py",
            f"Filen saknas: {EXCHANGE_FILE}")
        return

    con = db.connect()
    try:
        db.init_schema(con)
        counts = load_exchange_rates(con, dry_run=args.dry_run)
        total = sum(counts.values())
        log("DONE", "load_exchange_rates.py",
            f"Totalt {total} kurser laddade  {counts}")
    finally:
        con.close()


if __name__ == "__main__":
    main()
