#!/usr/bin/env python3
"""
process_germany.py  –  German SuSa / Monthly-Value XLSX → INL.xlsx

Kör från C:\\Users\\DidWac\\dev\\finance-automation\\ :
    py process_germany.py               # kör alla bolag
    py process_germany.py 231 245       # kör specifika bolag
    py process_germany.py --dry-run     # visa utan att skriva

Vad scriptet gör:
  1. Läser SuSa / Monthly-Value XLSX per bolag
  2. Extraherar månadsrörelser i kredit-debet-konvention (intäkter +, kostnader -)
  3. Skriver {kod}_{Namn}_{YYYYMM}_INL.xlsx till Germany/output/
  4. Flyttar källfiler till Germany/Referens/

Bolagsspecifika konfigurationer:
  187 (Prosero GmbH):  Monthly-Value XLSX, konton 0–89999, negera alla belopp
  188 (Bofferding):    Monthly-Value XLSX, konton 0–89999, negera alla belopp
  220 (Weckbacher):    SuSa CSV (cp1252, semikolon), amount=-(Soll+Haben), konton 0–8999
  231 (Mittermeier):   SuSa pro Monat, konton 0–8999, Haben−Soll per Mrz-kolumn
  245 (GOLDfunk):      SuSa pro Monat, konton 0–8999, Haben−Soll per Mrz-kolumn
  246 (HW Mechatronic): SuSa Jahresübersicht, Mrz-kolumn, konton 0–8999

Teckensättning SuSa (231/245/246):
  amount = Haben − Soll  →  intäkter +, kostnader −, tillgångsökning −, skuld­ökning +
  (summan av alla konton 0–8999 = 0 per dubbel bokföring)

Teckensättning Monthly-Value (187/188):
  filen lagrar debet-positivt (intäkter negativa, kostnader positiva)
  → negera alla belopp för att nå kredit-debet-konvention

IS-konton:
  187/188: 20 000–89 999  (engelskspråkigt DATEV-system)
  231/245/246: 4 000–8 999  (DATEV SKR03/04)
BS-konton:
  187/188: 0–19 999
  231/245/246: 0–3 999
"""

import argparse
import re
import sys

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
from datetime import date
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("Saknar openpyxl — kör:  py -m pip install openpyxl")

from shared import load_dotterbolag, move_to_referens_safe, save_inl_xlsx, load_config, log, glob_one, begin_run

# ── Paths ──────────────────────────────────────────────────────────────────────
_BASE = Path(__file__).resolve().parent

GET_TESTFILES = Path(load_config()["base_path"])
GERMANY_DIR  = GET_TESTFILES / "extracted" / "Germany"
OUTPUT_DIR   = GERMANY_DIR / "output"
REFERENS_DIR = GERMANY_DIR / "Referens"
DOTTERBOLAG  = _BASE / "_params" / "Dotterbolagslista.xlsx"

# ── Company definitions ────────────────────────────────────────────────────────
# reader    : "monthly_value" | "susa_pro_monat" | "susa_jahresuebersicht" | "susa_csv" | "skip"
# file_glob : glob pattern to discover the main file; supports {period} placeholder
COMPANY_DEFS: dict[str, dict] = {
    "187": dict(reader="monthly_value",        file_glob="187_*monthly*value*.xlsx"),
    "188": dict(reader="monthly_value",        file_glob="188_*monthly*value*.xlsx"),
    "220": dict(reader="susa_csv",             file_glob="220_*Susa*.csv"),
    # 231 Mittermeier: byter mellan "Susa MM.YYYY" och "Auswertung M-YYYY" (jan).
    "231": dict(reader="susa_pro_monat",       file_glob="231_*Susa*.xlsx",
                fallback_glob="231_Auswertung*.xlsx"),
    "245": dict(reader="susa_pro_monat",       file_glob="245_*SuSa*.xlsx"),
    # 246 HW Mechatronic: SUSA-format för 202603, sedan Fennoa-fi-csv från 202604
    # (de bytte bokföringssystem). file_glob försöker SUSA först; om saknad,
    # process_company faller tillbaka till fennoa-csv via reader-namnet.
    "246": dict(reader="fennoa_fi_or_susa",
                file_glob="246_SUSA_{period}*.xlsx",
                fallback_glob="246_tase-kk_*.csv"),
}

# ── Period helpers ─────────────────────────────────────────────────────────────
_MONTH_MAP = {
    "jan": 1,  "feb": 2,  "mar": 3,  "mrz": 3,  "apr": 4,
    "mai": 5,  "may": 5,  "jun": 6,  "jul": 7,  "aug": 8,
    "sep": 9,  "okt": 10, "oct": 10, "nov": 11, "dez": 12, "dec": 12,
}


def prev_month_period() -> str:
    today = date.today()
    if today.month == 1:
        return f"{today.year - 1}12"
    return f"{today.year}{today.month - 1:02d}"


def _match_period(cell_val, period: str) -> bool:
    """True if a header cell value matches YYYYMM period."""
    if cell_val is None:
        return False
    year, month = period[:4], int(period[4:])
    val = str(cell_val).strip().lower()
    m = re.match(r"^([a-zäöü]+)[/\s\-](\d{4})$", val)
    if m:
        abbr = m.group(1)[:3]
        yr = m.group(2)
        return yr == year and _MONTH_MAP.get(abbr) == month
    return False


# ── Amount parsing ─────────────────────────────────────────────────────────────
def parse_de(val) -> float:
    """Parse German number format: '1.234,56' → 1234.56."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return 0.0 if val != val else float(val)
    s = str(val).strip().replace("\xa0", "").replace(" ", "")
    if not s or s in ("-", "+"):
        return 0.0
    s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def parse_amount(val) -> float:
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return 0.0 if val != val else float(val)
    s = str(val).strip().replace("\xa0", "").replace(" ", "")
    if not s or s in ("-", "+"):
        return 0.0
    s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def parse_account(val) -> int | None:
    """Parse account number; strips spaces (e.g. '84 000' → 84000)."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        try:
            return int(float(val))
        except (ValueError, OverflowError):
            return None
    s = str(val).strip().replace(" ", "")
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return None




# ── BS/IS classification ───────────────────────────────────────────────────────
def classify_188(acc: int) -> str | None:
    """'IS' | 'BS' | None (=exclude) for Bofferding English-DATEV numbering."""
    if acc >= 90000:
        return None
    if 20000 <= acc <= 89999:
        return "IS"
    return "BS"


def classify_susa(acc: int) -> str | None:
    """'IS' | 'BS' | None for standard DATEV SKR03/04 SuSa accounts."""
    if acc >= 9000:
        return None          # sub-ledger, contra, or statistical accounts
    if 4000 <= acc <= 8999:
        return "IS"
    return "BS"


# ── Reader: Monthly-Value XLSX (187, 188) ─────────────────────────────────────
def read_monthly_value(
    filepath: Path, period: str
) -> tuple[list, list]:
    """
    Reads 187/188-style 'monthly value' file.

    Finds the column header matching *period* (e.g. 'Mar/2026' for 202603).
    Negates amounts to convert from debit-positive to kredit-debet convention.
    Returns (is_rows, bs_rows).
    """
    wb = openpyxl.load_workbook(str(filepath), data_only=True)
    ws = wb.active

    amount_col: int | None = None
    header_row: int | None = None

    for row_idx, row in enumerate(ws.iter_rows(max_row=5, values_only=True)):
        for col_idx, cell in enumerate(row):
            if _match_period(cell, period):
                amount_col = col_idx
                header_row = row_idx
                break
        if amount_col is not None:
            break

    if amount_col is None:
        wb.close()
        raise ValueError(
            f"Kunde inte hitta månadskolumn för period {period} i {filepath.name}"
        )

    # Account in col 0, name in col 1
    is_rows: list = []
    bs_rows: list = []

    for row in ws.iter_rows(min_row=header_row + 2, values_only=True):
        acc = parse_account(row[0] if row else None)
        if acc is None:
            continue
        kind = classify_188(acc)
        if kind is None:
            continue
        name = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        raw = row[amount_col] if len(row) > amount_col else None
        amt = round(-parse_amount(raw), 2)   # negate: debit-positive → kredit-debet
        if amt == 0.0:
            continue
        if kind == "IS":
            is_rows.append((acc, name, amt))
        else:
            bs_rows.append((acc, name, amt))

    wb.close()
    return is_rows, bs_rows


# ── Reader: SuSa pro Monat (231, 245) ─────────────────────────────────────────
def read_susa_pro_monat(filepath: Path) -> tuple[list, list]:
    """
    Reads 'Summen- und Saldenliste pro Monat' format.

    Expected header (row 2):
      Konto | Beschriftung | EB-Wert | S | H | Saldo | S | H |
      Mrz 2026 Soll | Haben | kum. Werte Soll | Haben

    Amount = col9 (Haben) − col8 (Soll) = kredit-debet sign.
    Excludes accounts >= 9000 (sub-ledger, contra).
    """
    wb = openpyxl.load_workbook(str(filepath), data_only=True)
    ws = wb.active

    is_rows: list = []
    bs_rows: list = []

    for row in ws.iter_rows(min_row=3, values_only=True):
        acc = parse_account(row[0] if row else None)
        if acc is None:
            continue
        kind = classify_susa(acc)
        if kind is None:
            continue
        name = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        soll  = parse_amount(row[8] if len(row) > 8 else None)
        haben = parse_amount(row[9] if len(row) > 9 else None)
        amt = round(haben - soll, 2)
        if amt == 0.0:
            continue
        if kind == "IS":
            is_rows.append((acc, name, amt))
        else:
            bs_rows.append((acc, name, amt))

    wb.close()
    return is_rows, bs_rows


# ── Reader: SuSa Jahresübersicht (246) ────────────────────────────────────────
def read_susa_jahresuebersicht(
    filepath: Path, period: str
) -> tuple[list, list]:
    """
    Reads 'SUSA Jahresübersicht' with one column group per month.

    Header layout per month group: amount | S | H
    Finds the column group matching *period* via header text (e.g. 'Mrz/2026').
    Amount sign: +amount if H-indicator, −amount if S-indicator.
    Excludes accounts >= 9000.
    """
    wb = openpyxl.load_workbook(str(filepath), data_only=True)
    ws = wb.active

    amount_col: int | None = None
    header_row: int | None = None

    for row_idx, row in enumerate(ws.iter_rows(max_row=5, values_only=True)):
        for col_idx, cell in enumerate(row):
            if _match_period(cell, period):
                amount_col = col_idx
                header_row = row_idx
                break
        if amount_col is not None:
            break

    if amount_col is None:
        wb.close()
        raise ValueError(
            f"Kunde inte hitta månadskolumn för period {period} i {filepath.name}"
        )

    s_col = amount_col + 1   # S (Soll/debit) indicator column
    h_col = amount_col + 2   # H (Haben/credit) indicator column

    is_rows: list = []
    bs_rows: list = []

    for row in ws.iter_rows(min_row=header_row + 2, values_only=True):
        acc = parse_account(row[0] if row else None)
        if acc is None:
            continue
        kind = classify_susa(acc)
        if kind is None:
            continue
        if len(row) <= amount_col or row[amount_col] is None:
            continue
        name = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        raw = parse_amount(row[amount_col])
        s_ind = row[s_col] if len(row) > s_col else None
        h_ind = row[h_col] if len(row) > h_col else None
        if h_ind == "H":
            amt = round(raw, 2)
        elif s_ind == "S":
            amt = round(-raw, 2)
        else:
            continue
        if amt == 0.0:
            continue
        if kind == "IS":
            is_rows.append((acc, name, amt))
        else:
            bs_rows.append((acc, name, amt))

    wb.close()
    return is_rows, bs_rows


# ── Reader: Fennoa period-CSV (246 fr.o.m. 2026)  ─────────────────────────────
# 246 HW Mechatronic migrerade till finsk bokföring under 2026. Filerna är samma
# format som 238 ANV i process_finland: ";202601;202602;..." header, finska
# kontonamn, cp1252-encoding, semikolon-separator, kommadecimal med space-tusental.
#
# BS-fil (tase-kk): periodbalanser → månadsändring = col[period] - col[prev_period].
# IS-fil (tuloslaskelma-kk): månadsbelopp direkt i col[period]. Yhteensä-kolumn
# ignoreras.
# Använder finska "tilikartta"-tecken: revenue negativ, expense positiv (SIE-konv.).

import re as _re

_ACC_RE_FI = _re.compile(r"^(\d+)[\s,]+(.+)$")


def _parse_fi_account(text: str):
    text = text.strip()
    m = _ACC_RE_FI.match(text)
    if m:
        return int(m.group(1)), m.group(2).strip()
    return None


def _classify_fi(acc: int) -> str | None:
    """Finska konton: 1–2999 = BS, 3–9 = IS (med samma 9000-cutoff för
    sub-ledger som SuSa)."""
    if acc >= 9000:
        return None
    if 1000 <= acc <= 2999:
        return "BS"
    return "IS"


def _read_fi_csv_rows(filepath: Path) -> list[list[str]]:
    import csv
    with open(filepath, encoding="cp1252", newline="") as fh:
        return [row for row in csv.reader(fh, delimiter=";")]


def _find_period_col_fi(rows: list[list[str]], period: str) -> int:
    """Header row 0 is ';202601;202602;...'. Returnerar kolumn-index för
    target period."""
    if not rows:
        raise ValueError("Tom fil")
    header = rows[0]
    for i, cell in enumerate(header):
        if cell.strip() == period:
            return i
    raise ValueError(f"Period {period} hittades ej i headern {header}")


def read_fennoa_fi_csv(filepath: Path, period: str, *, kind: str) -> list:
    """kind = 'BS' (period_csv_diff) eller 'IS' (period_csv_col)."""
    rows = _read_fi_csv_rows(filepath)
    period_col = _find_period_col_fi(rows, period)
    prev_col = period_col - 1
    out: list = []
    for row in rows[1:]:
        if not row:
            continue
        parsed = _parse_fi_account(row[0])
        if parsed is None:
            continue
        acc, name = parsed
        cls = _classify_fi(acc)
        # För BS-fil läses BS-konton, för IS-fil läses IS-konton — om filen
        # blandar (vissa Fennoa-filer gör det) hoppa över felfacet.
        if cls != kind:
            continue
        if kind == "BS":
            # Månadsändring = col[period] - col[period-1]. För första kolumnen
            # (period 1) är "previous" inte i filen — anta opening = 0.
            curr = parse_amount(row[period_col] if len(row) > period_col else None)
            prev = (parse_amount(row[prev_col]) if prev_col >= 1 and len(row) > prev_col
                    else 0.0)
            amt = round(curr - prev, 2)
            # Spegla SIE-tecken: tillgång ökning = negativ förändring i kredit
            # men positiv i debet → SIE har asset-konton positiva (men ökning
            # ger positivt belopp på debit-sidan). Här gör Finland-238 inget
            # specifikt — vi matchar deras logik.
            if 1 <= acc <= 1999:
                amt = -amt  # asset-flip (matchar process_finland.should_flip)
        else:
            amt = parse_amount(row[period_col] if len(row) > period_col else None)
        if amt == 0.0:
            continue
        out.append((acc, name, amt))
    return out


def read_fennoa_fi_pair(d: Path, period: str) -> tuple[list, list]:
    """Hittar bolagets tase- och tuloslaskelma-csv i d och returnerar (IS, BS)."""
    bs_files = sorted(d.glob("246_tase-kk_*.csv"))
    is_files = sorted(d.glob("246_tuloslaskelma-kk_*.csv"))
    if not bs_files or not is_files:
        raise FileNotFoundError(f"saknar tase-kk/tuloslaskelma-kk csv i {d}")
    bs = read_fennoa_fi_csv(bs_files[0], period, kind="BS")
    is_ = read_fennoa_fi_csv(is_files[0], period, kind="IS")
    return is_, bs


# ── Reader: SuSa CSV (220) ────────────────────────────────────────────────────
def read_susa_csv(filepath: Path) -> tuple[list, list]:
    """
    Reads DATEV SuSa CSV (cp1252, semicolon-delimited).

    Column layout:
      0 Kontonummer | 1 Kontobezeichnung | 2 Anfangsbestand |
      3 Monatsumsatz Soll | 4 Monatsumsatz Haben | ...

    Haben is stored as a negative value.
    amount = -(Soll + Haben)  →  kredit-debet convention (revenues +, costs −).
    Skips rows with no account number (summary/header rows).
    Excludes accounts >= 9000.
    """
    import csv

    is_rows: list = []
    bs_rows: list = []

    with open(filepath, encoding="cp1252", newline="") as fh:
        reader = csv.reader(fh, delimiter=";")
        for row in reader:
            if not row:
                continue
            acc = parse_account(row[0])
            if acc is None:
                continue
            kind = classify_susa(acc)
            if kind is None:
                continue
            name = row[1].strip() if len(row) > 1 else ""
            soll  = parse_de(row[3] if len(row) > 3 else None)
            haben = parse_de(row[4] if len(row) > 4 else None)
            amt = round(-(soll + haben), 2)
            if amt == 0.0:
                continue
            if kind == "IS":
                is_rows.append((acc, name, amt))
            else:
                bs_rows.append((acc, name, amt))

    return is_rows, bs_rows


# ── Referens move ──────────────────────────────────────────────────────────────
def move_to_referens(filename: str, dry_run: bool) -> None:
    src = GERMANY_DIR / filename
    if not src.exists():
        return
    move_to_referens_safe(src, REFERENS_DIR, dry_run)


# ── Process one company ────────────────────────────────────────────────────────
def process_company(
    code: str,
    friendly: str,
    period: str,
    filepath: Path,
    cfg: dict,
    dry_run: bool,
) -> str:
    log("INFO", code, f"{friendly}")

    if cfg["reader"] == "skip":
        log("SKIP", code, "Krypterade filer (.p7m) kan inte bearbetas automatiskt")
        return "skip"

    log("INFO", code, f"Fil: {filepath.name}")

    if not filepath.exists():
        log("SKIP", code, "Filen saknas (redan i Referens?)")
        return "skip"

    try:
        reader = cfg["reader"]
        if reader == "monthly_value":
            is_rows, bs_rows = read_monthly_value(filepath, period)
        elif reader == "susa_pro_monat":
            is_rows, bs_rows = read_susa_pro_monat(filepath)
        elif reader == "susa_jahresuebersicht":
            is_rows, bs_rows = read_susa_jahresuebersicht(filepath, period)
        elif reader == "susa_csv":
            is_rows, bs_rows = read_susa_csv(filepath)
        elif reader == "fennoa_fi_or_susa":
            # Dispatch på filtyp: tase-kk-csv → fennoa-fi; xlsx → SUSA i endera
            # "Jahresübersicht" eller "pro Monat"-format (HW Mechatronic levererar
            # båda i olika månader). Jahres försöks först; om månadskolumnen inte
            # hittas faller vi tillbaka till pro_monat-läsaren.
            if filepath.suffix.lower() == ".csv":
                is_rows, bs_rows = read_fennoa_fi_pair(filepath.parent, period)
            else:
                try:
                    is_rows, bs_rows = read_susa_jahresuebersicht(filepath, period)
                except ValueError:
                    is_rows, bs_rows = read_susa_pro_monat(filepath)
        else:
            log("ERROR", code, f"Okänd reader: {reader}")
            return "error"
    except Exception as e:
        log("ERROR", code, f"Läsfel: {e}")
        return "error"

    total = sum(r[2] for r in is_rows + bs_rows)
    is_warn = abs(total) >= 1.0
    check = "OK" if not is_warn else f"KONTROLLERA ({total:.2f})"
    log("INFO", code, f"IS={len(is_rows)}, BS={len(bs_rows)}  Summa={total:.4f}  {check}")

    out_name = f"{code}_{friendly}_{period}_INL.xlsx"
    out_path = OUTPUT_DIR / out_name

    if not dry_run:
        save_inl_xlsx(is_rows, bs_rows, out_path)
    dry_prefix = "[DRY] " if dry_run else ""
    log("WARN" if is_warn else "OK", code, f"{dry_prefix}Sparad: {out_name}")

    return "warn" if is_warn else "ok"


# ── Main ───────────────────────────────────────────────────────────────────────
def main() -> None:
    global GERMANY_DIR, OUTPUT_DIR, REFERENS_DIR

    parser = argparse.ArgumentParser(
        description="Bearbeta tyska SuSa/Monthly-Value-filer → INL.xlsx"
    )
    parser.add_argument(
        "codes", nargs="*",
        help="Bolagskoder att köra (standard: alla). Ex: py process_germany.py 231 245",
    )
    parser.add_argument(
        "--dry-run", "-n", action="store_true",
        help="Visa vad som skulle hända utan att skriva några filer",
    )
    parser.add_argument(
        "--period", "-p", metavar="YYYYMM", default=None,
        help="Period att köra (t.ex. 202604). Standard: föregående månad.",
    )
    args = parser.parse_args()

    if args.period:
        GERMANY_DIR  = GET_TESTFILES / "extracted" / args.period / "Germany"
        OUTPUT_DIR   = GERMANY_DIR / "output"
        REFERENS_DIR = GERMANY_DIR / "Referens"

    period = args.period or prev_month_period()
    begin_run("process_germany", period)
    dry_label = "  [DRY RUN]" if args.dry_run else ""
    log("START", "process_germany.py", f"period {period}{dry_label}")

    if not GERMANY_DIR.exists():
        sys.exit(f"[ERROR]  Germany-mappen saknas: {GERMANY_DIR}")
    if not DOTTERBOLAG.exists():
        sys.exit(f"[ERROR]  Dotterbolagslistan saknas: {DOTTERBOLAG}")

    friendlies = load_dotterbolag(DOTTERBOLAG)

    if not args.dry_run:
        OUTPUT_DIR.mkdir(exist_ok=True)
        REFERENS_DIR.mkdir(exist_ok=True)

    all_codes  = sorted(COMPANY_DEFS.keys())
    codes_to_run = args.codes if args.codes else all_codes
    stats: dict[str, int] = {"ok": 0, "warn": 0, "skip": 0, "error": 0}

    for code in codes_to_run:
        if code not in COMPANY_DEFS:
            log("ERROR", code, "Okänd bolagskod")
            stats["error"] += 1
            continue

        cfg      = COMPANY_DEFS[code]
        friendly = friendlies.get(int(code), f"Bolag{code}")

        if cfg["reader"] == "skip":
            log("SKIP", code, "Krypterade filer (.p7m) kan inte bearbetas automatiskt")
            stats["skip"] += 1
            continue

        glob_pattern = cfg["file_glob"].format(period=period)
        try:
            filepath = glob_one(GERMANY_DIR, glob_pattern)
        except FileNotFoundError:
            # Stöd för fallback-glob (t.ex. 246 som bytt filformat mitt under året)
            fallback = cfg.get("fallback_glob")
            if fallback:
                try:
                    filepath = glob_one(GERMANY_DIR, fallback.format(period=period))
                except FileNotFoundError:
                    log("SKIP", code, "Källfil saknas (redan i Referens?)")
                    stats["skip"] += 1
                    continue
            else:
                log("SKIP", code, "Källfil saknas (redan i Referens?)")
                stats["skip"] += 1
                continue

        status = process_company(code, friendly, period, filepath, cfg, args.dry_run)
        stats[status] = stats.get(status, 0) + 1

        # Move all remaining {code}_* source files to Referens
        if not args.dry_run:
            REFERENS_DIR.mkdir(exist_ok=True)
        for f in sorted(GERMANY_DIR.glob(f"{code}_*")):
            if f.is_file():
                move_to_referens(f.name, args.dry_run)

    log("DONE", "process_germany.py",
        f"{stats['ok']} OK  {stats['warn']} WARN  {stats['skip']} SKIP  {stats['error']} ERROR")


if __name__ == "__main__":
    main()
