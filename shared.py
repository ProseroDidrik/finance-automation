"""Shared utilities for finance-automation country processing scripts."""
import json
import re
from datetime import date, datetime
from pathlib import Path
import shutil
import sys

DUPE_RE = re.compile(r"\s\(([2-9]|\d{2,})\)\.\w+$")

# Filformats-baserade landsbegränsningar för mail-matchning.
# SIE-filer är ett svenskt redovisningsformat → kan endast gälla svenska bolag.
# SAF-T används i denna pipeline för norska och danska bolag.
_SIE_RE  = re.compile(r"\.sie\b|\bsie[- ]?fil|\bsiefil", re.IGNORECASE)
_SAFT_RE = re.compile(r"\bsaf[- ]?t\b", re.IGNORECASE)


def country_constraint_from_haystacks(haystacks: dict) -> tuple[str, ...] | None:
    """Returnera tillåtna länder baserat på filformatssignal i mailet.

    Inspekterar filename + subject + att_name (inte body/sender — för noisy).
    SIE-signal → ('Sweden',). SAF-T-signal → ('Norway', 'Denmark'). Annars None.
    """
    text = " ".join((
        haystacks.get("filename", ""),
        haystacks.get("subject", ""),
        haystacks.get("att_name", ""),
    ))
    if _SIE_RE.search(text):
        return ("Sweden",)
    if _SAFT_RE.search(text):
        return ("Norway", "Denmark")
    return None

_REPO_ROOT = Path(__file__).resolve().parent
_run_ctx: dict = {"script": None, "period": None, "log_path": None}


def prev_month_period() -> str:
    """Return YYYYMM for the previous calendar month (handles January wrap)."""
    today = date.today()
    if today.month == 1:
        return f"{today.year - 1}12"
    return f"{today.year}{today.month - 1:02d}"


def is_override_for(override: list[int] | None, company_id: int) -> bool:
    """True om --override gäller för company_id.

    argparse-konvention: nargs='*', default=None.
    - None      → flaggan inte angiven → False
    - []        → --override utan args → True (global override)
    - [a, b, …] → --override a b      → True endast för listade IDs
    """
    if override is None:
        return False
    return not override or company_id in override


def begin_run(script_name: str, period: str) -> None:
    """Activate JSONL persistence for subsequent log() calls in this process.

    Writes to _logs/{period}/{script_name}.jsonl. Idempotent within a process.
    Terminal output via log() is unaffected; only adds a side-effect write.
    """
    log_dir = _REPO_ROOT / "_logs" / period
    log_dir.mkdir(parents=True, exist_ok=True)
    _run_ctx["script"] = script_name
    _run_ctx["period"] = period
    _run_ctx["log_path"] = log_dir / f"{script_name}.jsonl"
    _append_event("START", script_name, f"period {period}")


def _append_event(status: str, label, msg: str) -> None:
    if not _run_ctx["log_path"]:
        return
    rec = {
        "ts": datetime.now().isoformat(timespec="seconds"),
        "script": _run_ctx["script"],
        "period": _run_ctx["period"],
        "status": status,
        "label": str(label),
        "msg": msg,
    }
    try:
        with open(_run_ctx["log_path"], "a", encoding="utf-8") as f:
            f.write(json.dumps(rec, ensure_ascii=False) + "\n")
    except OSError:
        pass

try:
    import openpyxl
except ImportError:
    sys.exit("Saknar openpyxl — kör:  py -m pip install openpyxl")

try:
    import pandas as pd
except ImportError:
    sys.exit("Saknar pandas — kör:  py -m pip install pandas openpyxl")


def load_config() -> dict:
    """Load base_path (and other settings) from config.json in the repo root."""
    config_path = Path(__file__).parent / "config.json"
    if not config_path.exists():
        raise FileNotFoundError(
            "config.json saknas. Skapa den i repo-roten med innehållet:\n"
            '  {"base_path": "C:\\\\...\\\\Get testfiles"}'
        )
    with open(config_path, encoding="utf-8") as f:
        return json.load(f)


def log(status: str, label, msg: str = "") -> None:
    """Print a structured log line: [STATUS]  label  msg

    If begin_run() has been called in this process, also append a JSONL event
    to _logs/{period}/{script}.jsonl so a GUI can read run history later.
    """
    tag = f"[{status}]"
    line = f"{tag:<8} {label}"
    if msg:
        line += f"  {msg}"
    print(line)
    _append_event(status, label, msg)


def log_event(status: str, label, msg: str = "") -> None:
    """Append a JSONL event without printing. Useful for events that should be
    visible to the GUI (via _logs/*.jsonl) but not clutter the script's stdout."""
    _append_event(status, label, msg)


def load_dotterbolag(path: Path) -> dict[int, str]:
    """bolagsid → friendly name from Dotterbolagslistan, skips 'consolidated' rows."""
    path = _resolve_master_path(path)
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb["Data For Company Find"]
    result: dict[int, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 5:
            continue
        bolag_id = row[1]
        friendly = row[4]
        kind = row[7] if len(row) > 7 else None
        if str(kind).strip().lower() == "consolidated":
            continue
        if bolag_id and friendly:
            result[int(bolag_id)] = str(friendly).strip()
    wb.close()
    return result


def load_dotterbolag_full(path: Path) -> dict[int, dict]:
    """bolagsid → {name, country, orgnr, domain, kind, ...} from Dotterbolagslistan.

    Includes 'consolidated' rows (caller filters if needed). Used by the GUI
    where Country (col C) is needed in addition to friendly name.

    Acquisition-fields (cols K–P) are returned as None when missing/0:
      closing_date (date), investment_currency (str),
      ev_sek_m, ev_ebitda_ltm, ebitda_ltm, sales_ltm (float).
    """
    from datetime import date, datetime as _dt

    def _to_float(v):
        if v is None or v == "":
            return None
        try:
            f = float(v)
        except (TypeError, ValueError):
            return None
        return f if f != 0 else None

    def _to_date(v):
        if v is None or v == "":
            return None
        if isinstance(v, _dt):
            return v.date()
        if isinstance(v, date):
            return v
        # Fallback: ISO-sträng
        try:
            return _dt.fromisoformat(str(v).strip()[:10]).date()
        except (TypeError, ValueError):
            return None

    path = _resolve_master_path(path)
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb["Data For Company Find"]
    result: dict[int, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 2 or row[1] is None:
            continue
        try:
            bolag_id = int(row[1])
        except (TypeError, ValueError):
            continue
        # Förvärvsår (col D, index 3) — int eller None
        try:
            acq_year = int(row[3]) if len(row) > 3 and row[3] is not None else None
        except (TypeError, ValueError):
            acq_year = None
        # Parent (col G, index 6) — bolagsId till konsoliderat moderbolag
        try:
            parent_id = int(row[6]) if len(row) > 6 and row[6] is not None else None
        except (TypeError, ValueError):
            parent_id = None
        inv_cur_raw = row[11] if len(row) > 11 else None
        inv_cur = str(inv_cur_raw).strip().upper() if inv_cur_raw else None
        result[bolag_id] = {
            "name": str(row[4]).strip() if len(row) > 4 and row[4] else "",
            "country": str(row[2]).strip() if len(row) > 2 and row[2] else "",
            "orgnr": str(row[5]).strip() if len(row) > 5 and row[5] else "",
            "kind": str(row[7]).strip() if len(row) > 7 and row[7] else "",
            "domain": str(row[9]).strip() if len(row) > 9 and row[9] else "",
            "acquisition_year": acq_year,
            "parent_id": parent_id,
            # Förvärvsfält (cols K–P)
            "closing_date": _to_date(row[10]) if len(row) > 10 else None,
            "investment_currency": inv_cur or None,
            "ev_sek_m": _to_float(row[12]) if len(row) > 12 else None,
            "ev_ebitda_ltm": _to_float(row[13]) if len(row) > 13 else None,
            "ebitda_ltm": _to_float(row[14]) if len(row) > 14 else None,
            "sales_ltm": _to_float(row[15]) if len(row) > 15 else None,
        }
    wb.close()
    return result


def load_overrides() -> dict:
    """Load _params/overrides.json (subject/attachment/sender/country/alias overrides for extract).

    Returns a dict with keys: subject_overrides (dict[str, int]),
    attachment_overrides (list of {msg_stem, attachment_substr, bolag_id}),
    sender_overrides (dict[str, int] — sender domain or email substring → bolag_id;
    triggar när delsträngen finns i mailets sender-fält, FÖRE scoring. Användbart
    för bolag med korta/generiska subject som annars förlorar mot starkare
    attachment-träffar),
    country_overrides (dict[str, str]),
    aliases (dict[str, list[str]] — bolag_id → phrases that score full weight when
    found as substring in any haystack source),
    excluded (list[int] — bolag_id som GUI:t gömmer i tabellen; påverkar inte
    matchning eller pipeline). Empty defaults if file missing.
    """
    p = _REPO_ROOT / "_params" / "overrides.json"
    if not p.exists():
        return {
            "subject_overrides": {}, "attachment_overrides": [],
            "sender_overrides": {}, "country_overrides": {}, "aliases": {},
            "excluded": [],
        }
    with open(p, encoding="utf-8") as f:
        data = json.load(f)
    data.setdefault("subject_overrides", {})
    data.setdefault("attachment_overrides", [])
    data.setdefault("sender_overrides", {})
    data.setdefault("country_overrides", {})
    data.setdefault("aliases", {})
    data.setdefault("excluded", [])
    return data


def save_overrides(data: dict) -> None:
    """Atomic-write _params/overrides.json. Used by GUI when toggling exclusions etc."""
    import os
    import tempfile
    p = _REPO_ROOT / "_params" / "overrides.json"
    p.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp = tempfile.mkstemp(prefix=".overrides_", suffix=".tmp", dir=str(p.parent))
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
            f.write("\n")
        os.replace(tmp, p)
    except Exception:
        try:
            os.unlink(tmp)
        except OSError:
            pass
        raise


def safe_dest(dest: Path) -> Path:
    """Return a unique path: append _2, _3, ... if dest already exists."""
    if not dest.exists():
        return dest
    stem, suffix = dest.stem, dest.suffix
    i = 2
    while True:
        candidate = dest.parent / f"{stem}_{i}{suffix}"
        if not candidate.exists():
            return candidate
        i += 1


def move_to_referens_safe(src: Path, referens_dir: Path, dry_run: bool) -> Path:
    """Move src into referens_dir, avoiding filename collisions. Returns actual destination."""
    dst = safe_dest(referens_dir / src.name)
    if dry_run:
        print(f"    [dry] → Referens/{dst.name}")
    else:
        referens_dir.mkdir(parents=True, exist_ok=True)
        shutil.move(str(src), str(dst))
        print(f"    → Referens/{dst.name}")
    return dst


def glob_one(directory: Path, pattern: str) -> Path:
    """Return the best-matching file for pattern in directory.
    Prefers non-duplicate files (unique_path-created copies have ` (2)+` suffix);
    falls back to first match if all matches look like dupes."""
    matches = sorted(directory.glob(pattern))
    if not matches:
        raise FileNotFoundError(directory / pattern)
    non_dupes = [f for f in matches if not DUPE_RE.search(f.name)]
    return non_dupes[0] if non_dupes else matches[0]


def save_inl_xlsx(is_rows: list, bs_rows: list, output_path: Path) -> None:
    """Write IS+BS rows to INL.xlsx.

    Layout: empty row 1, then data rows with cols A=account, B=name, C=amount,
    D='IS' or 'BS'. Column D lets the DB loader assign statement_type without
    per-country heuristics; manual readers ignore it.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    records = [{"A": None, "B": None, "C": None, "D": None}]
    for acc, name, amt in is_rows:
        records.append({"A": acc, "B": name, "C": amt, "D": "IS"})
    for acc, name, amt in bs_rows:
        records.append({"A": acc, "B": name, "C": amt, "D": "BS"})
    df = pd.DataFrame(records)
    with pd.ExcelWriter(str(output_path), engine="openpyxl") as writer:
        df.to_excel(writer, index=False, header=False, sheet_name="Sheet1")


# ----- Azure Blob fallback för Dotterbolagslistan -----------------------------
#
# I molnet (App Service / Container Job) finns inte _params/ på filsystemet —
# Dropbox-mappen ligger på laptopen. Master-filen pushas separat till Blob via
# scripts/push_master.py. När shared.load_dotterbolag*-funktionerna körs i
# molnet hämtas filen från Blob:en istället. Lokalt (där _params/-filen finns)
# används den direkt och Blob rörs aldrig.
#
# Konfiguration:
#   MASTER_BLOB_URL  — full URL till blobben, t.ex.
#                      https://acct.blob.core.windows.net/master/Dotterbolagslista.xlsx
#                      Auth: DefaultAzureCredential (Managed Identity i App
#                      Service, az-cli-token lokalt).

import os as _os
import tempfile as _tempfile

_MASTER_CACHE: Path | None = None  # download-cache för processens livstid


def _resolve_master_path(local_path: Path) -> Path:
    """Returnera en användbar fil-path till Dotterbolagslistan.

    Prio:
    1. Om local_path finns (laptop / volume-mountad container) → använd den.
    2. Annars, om MASTER_BLOB_URL är satt → ladda ner till temp och cache:a
       per process. Spara i en modul-global så upprepade anrop inom samma
       process inte triggar nya downloads.
    3. Annars → höj samma FileNotFoundError som tidigare.
    """
    local_path = Path(local_path)
    if local_path.exists():
        return local_path

    global _MASTER_CACHE
    if _MASTER_CACHE is not None and _MASTER_CACHE.exists():
        return _MASTER_CACHE

    blob_url = _os.environ.get("MASTER_BLOB_URL")
    if not blob_url:
        raise FileNotFoundError(
            f"Master-filen saknas lokalt ({local_path}) och MASTER_BLOB_URL "
            "är inte satt — kan inte falla tillbaka till Azure Blob."
        )

    try:
        from azure.identity import DefaultAzureCredential
        from azure.storage.blob import BlobClient
    except ImportError as e:
        raise RuntimeError(
            "azure-identity / azure-storage-blob saknas — pip install dem. "
            f"Underliggande fel: {e}"
        ) from e

    cred = DefaultAzureCredential()
    client = BlobClient.from_blob_url(blob_url, credential=cred)
    tmp_dir = Path(_tempfile.gettempdir()) / "_master_cache"
    tmp_dir.mkdir(parents=True, exist_ok=True)
    cache_path = tmp_dir / local_path.name
    with open(cache_path, "wb") as f:
        downloader = client.download_blob()
        downloader.readinto(f)
    _MASTER_CACHE = cache_path
    return cache_path
