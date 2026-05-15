"""
process_finland.py  –  Finnish Balansräkning + Resultaträkning → INL.xlsx

Output:  {code}_{FriendlyName}_{YYYYMM}_INL.xlsx
Layout:  empty row 1, then IS rows, then BS rows
         A = account number,  B = account name,  C = amount
Signs:   BS accounts 1-1999 (or 100-199999 for 6-digit) are multiplied by -1
Skip:    237X accounts (årets resultat), zero amounts, summary rows, bold rows (134)
Verify:  column C must sum to 0
"""

import argparse
import csv
import os
import re
import sys
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import pandas as pd
import xlrd

from shared import move_to_referens_safe, save_inl_xlsx, load_config, log, glob_one, DUPE_RE, begin_run


# ---------------------------------------------------------------------------
# Paths  (reassigned in main() when --period is given)
# ---------------------------------------------------------------------------

FINLAND_DIR  = Path(load_config()["base_path"]) / "extracted" / "Finland"
OUTPUT_DIR   = FINLAND_DIR / "output"
REFERENS_DIR = FINLAND_DIR / "Referens"

_DRY_RUN = False
PERIOD   = ""   # set in main()


# ---------------------------------------------------------------------------
# Sheet-name discovery (used by 177 Lukkoluket)
# ---------------------------------------------------------------------------

def _find_sheet(filepath: str | Path, pattern: str) -> str:
    from openpyxl import load_workbook as _load
    wb = _load(str(filepath), read_only=True)
    for name in wb.sheetnames:
        if re.search(pattern, name, re.IGNORECASE):
            wb.close()
            return name
    wb.close()
    raise ValueError(f"No sheet matching '{pattern}' in {Path(filepath).name}")


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

def parse_amount(val) -> float:
    if val is None or (isinstance(val, float) and val != val):
        return 0.0
    s = str(val).strip().replace("\xa0", "").replace(" ", "")
    s = s.replace(" ", "").replace(",", ".")          # Finnish thousands/decimal
    if s in ("", "nan", "0.00", "0"):
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


_ACC_RE = re.compile(r"^(\d+)[\s,]+(.+)$")

def split_account(text: str):
    """Return (int_acc, name) or None for non-account rows."""
    text = text.strip()
    m = _ACC_RE.match(text)
    if m:
        return int(m.group(1)), m.group(2).strip()
    return None


def _normalize4(account: int) -> int:
    """Return the 4-digit prefix of any account number (works for 4, 5, 6 digits)."""
    s = str(account)
    return int(s[:4]) if len(s) > 4 else account


def is_237x(account: int) -> bool:
    """True for årets-resultat accounts (237X family, any digit count)."""
    return 2370 <= _normalize4(account) <= 2379


def should_flip(account: int) -> bool:
    """BS asset accounts (1–1999 in 4-digit space) need sign flip."""
    return 1 <= _normalize4(account) <= 1999


def apply_flip(amount: float, account: int) -> float:
    return -amount if should_flip(account) else amount


def extract_company_name(raw: str) -> str:
    name = str(raw).strip()
    for sfx in [" Oy", " Ab", " AS", " A/S", " Ltd", " OY", " AB", " ry"]:
        if name.endswith(sfx):
            name = name[: -len(sfx)].strip()
            break
    return name


_MONTH_NAMES = {
    "january": 1,  "february": 2,  "march": 3,    "april": 4,
    "may": 5,      "june": 6,      "july": 7,     "august": 8,
    "september": 9, "october": 10, "november": 11, "december": 12,
    "tammikuu": 1, "helmikuu": 2,  "maaliskuu": 3, "huhtikuu": 4,
    "toukokuu": 5, "kesäkuu": 6,   "heinäkuu": 7,  "elokuu": 8,
    "syyskuu": 9,  "lokakuu": 10,  "marraskuu": 11, "joulukuu": 12,
}


def extract_period(text: str) -> str:
    """Extract YYYYMM from any period string (dates, period codes, month names)."""
    m = re.search(r"(\d{1,2})/(\d{4})", text)
    if m:
        return f"{m.group(2)}{int(m.group(1)):02d}"
    m = re.search(r"(\d{1,2})\.(\d{1,2})\.(\d{4})", text)
    if m:
        return f"{m.group(3)}{int(m.group(2)):02d}"
    m = re.search(r"(20\d{4})", text)
    if m:
        return m.group(1)
    # "March 2026" / "Maaliskuu 2026"
    m = re.search(r"(\w+)\s+(20\d{2})", text)
    if m:
        month_num = _MONTH_NAMES.get(m.group(1).lower())
        if month_num:
            return f"{m.group(2)}{month_num:02d}"
    return ""


def _find_period_col(rows: list, target_period: str) -> int | None:
    """Scan first rows for the data-column-header row and return the column whose
    period matches target_period.

    Prioriterar rader med >=2 period-celler — det är den riktiga kolumn-header-
    raden (e.g. 'Account|01.04.2026-30.04.2026|01.04.2025-30.04.2025|...').
    Skippar 'Date range'-rader som har period bara i en kolumn med fel offset.
    Fallback: greedy första match (för enkla layouter med 1 period-kolumn).
    """
    candidate_rows: list[tuple[int, list[tuple[int, str]]]] = []
    for r_idx, row in enumerate(rows[:10]):
        periods_in_row: list[tuple[int, str]] = []
        for i, cell in enumerate(row):
            p = extract_period(str(cell).strip() if cell is not None else "")
            if p:
                periods_in_row.append((i, p))
        if len(periods_in_row) >= 2:
            candidate_rows.append((r_idx, periods_in_row))
    if candidate_rows:
        for i, p in candidate_rows[0][1]:
            if p == target_period:
                return i
    # Fallback: greedy
    for row in rows[:10]:
        for i, cell in enumerate(row):
            if extract_period(str(cell).strip() if cell is not None else "") == target_period:
                return i
    return None


def _fennoa_month_range(period: str) -> str:
    """period=202604 → '(01.04.2026-30.04.2026)'. För att glob:a månads-Fennoa-export."""
    from calendar import monthrange
    yyyy, mm = period[:4], period[4:6]
    last_day = monthrange(int(yyyy), int(mm))[1]
    return f"(01.{mm}.{yyyy}-{last_day:02d}.{mm}.{yyyy})"


def verify_sum(is_rows, bs_rows) -> float:
    return sum(r[2] for r in is_rows + bs_rows)


def move_to_referens(filename: str):
    src = FINLAND_DIR / filename
    if src.exists():
        move_to_referens_safe(src, REFERENS_DIR, dry_run=_DRY_RUN)


def move_pdfs_to_referens():
    for f in FINLAND_DIR.iterdir():
        if f.is_file() and f.suffix.lower() == ".pdf":
            move_to_referens(f.name)


def _detect_csv_enc(filepath: str) -> str:
    """UTF-16-LE if null bytes in first 4 bytes (no-BOM variant), else latin-1."""
    with open(filepath, "rb") as fh:
        header = fh.read(4)
    if header[:2] in (b"\xff\xfe", b"\xfe\xff"):
        return "utf-16"
    if b"\x00" in header:
        return "utf-16-le"
    return "latin-1"


def _try_glob(d: Path, *patterns: str) -> Path:
    """Try patterns in order; return best non-dupe match. Raises FileNotFoundError if none match."""
    for pat in patterns:
        matches = sorted(d.glob(pat))
        if matches:
            non_dupes = [f for f in matches if not DUPE_RE.search(f.name)]
            return non_dupes[0] if non_dupes else matches[0]
    raise FileNotFoundError(d / patterns[0])


# ---------------------------------------------------------------------------
# Format A  –  "Fennoa CSV" (146, 161-IS, 170, 181-IS, 182-IS)
#   col 0 = "XXXX Name",  col 1 = latest month amount
#   encoding auto-detected (latin-1 or utf-16-le), separator ;
# ---------------------------------------------------------------------------

def read_fennoa_csv(filepath: str, target_period: str | None = None) -> list:
    """Fennoa-CSV: kolumn 0 = 'NNNN Namn', kolumn 1 = belopp default.

    Vissa Fennoa-exporter har flera datum-kolumner (FY föregående år | månad |
    YTD aktuell period). Om target_period är satt scannas header-raderna för
    att hitta den kolumn vars period matchar (via extract_period()). Fallback
    till kolumn 1 om ingen match.
    """
    enc = _detect_csv_enc(filepath)
    if enc in ("utf-16-le", "utf-16"):
        return read_csv_col_enc(filepath, col=1, encoding=enc, target_period=target_period)
    df = pd.read_csv(filepath, header=None, encoding=enc, sep=";", dtype=str)
    col = 1
    if target_period:
        # Skapa list-of-rows och delegera till _find_period_col (samma logik som
        # för utf-16-filerna — prioriterar multi-period-header-rader framför
        # 'Date range'-etiketter med fel kolumn-offset).
        rows_for_scan = [
            [df.iloc[i, j] if pd.notna(df.iloc[i, j]) else "" for j in range(len(df.columns))]
            for i in range(min(10, len(df)))
        ]
        found = _find_period_col(rows_for_scan, target_period)
        if found is not None:
            col = found
    rows = []
    for _, row in df.iterrows():
        cell = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ""
        parsed = split_account(cell)
        if parsed is None:
            continue
        acc, name = parsed
        amt = parse_amount(row.iloc[col] if col < len(row) else None)
        if amt == 0.0:
            continue
        rows.append((acc, name, amt))
    return rows


# ---------------------------------------------------------------------------
# Format B  –  "Muutos CSV/XLSX"  (161-BS, 173-BS, 181-BS, 182-BS, 185-BS, 145-BS)
#   col 0 = "XXXX Name",  col 2 = Muutos (change)
#   encoding latin-1, separator ; (CSV) or default (XLSX)
# ---------------------------------------------------------------------------

def _collect_muutos(df: pd.DataFrame) -> list:
    rows = []
    for _, row in df.iterrows():
        cell = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ""
        parsed = split_account(cell)
        if parsed is None:
            continue
        acc, name = parsed
        amt = parse_amount(row.iloc[2])
        if amt == 0.0:
            continue
        rows.append((acc, name, amt))
    return rows


def read_muutos_csv(filepath: str) -> list:
    enc = _detect_csv_enc(filepath)
    if enc in ("utf-16-le", "utf-16"):
        return read_muutos_csv_enc(filepath, enc)
    df = pd.read_csv(filepath, header=None, encoding=enc, sep=";", dtype=str)
    return _collect_muutos(df)


def read_muutos_xlsx(filepath: str, sheet_name=0) -> list:
    df = pd.read_excel(filepath, sheet_name=sheet_name, header=None, dtype=str)
    return _collect_muutos(df)


# ---------------------------------------------------------------------------
# Format C  –  "Income-only XLSX"  (173-IS, 145-IS)
#   col 0 = "XXXX Name",  col 1 = amount
# ---------------------------------------------------------------------------

def read_income_only_xlsx(filepath: str, sheet_name=0) -> list:
    df = pd.read_excel(filepath, sheet_name=sheet_name, header=None, dtype=str)
    rows = []
    for _, row in df.iterrows():
        cell = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ""
        parsed = split_account(cell)
        if parsed is None:
            continue
        acc, name = parsed
        amt = parse_amount(row.iloc[1])
        if amt == 0.0:
            continue
        rows.append((acc, name, amt))
    return rows


# ---------------------------------------------------------------------------
# Format D  –  "Period XLS"  (166 IS)
#   row 0 = period labels (202601, 202602, 202603)
#   col 0 = "XXXX  Name",  find col with target_period → monthly IS amount
#
# Format D2 – "Period XLS diff" (166 BS)
#   Same layout but col values are CUMULATIVE balances → change = col_target - col_prev
# ---------------------------------------------------------------------------

def _find_period_col_df(df: pd.DataFrame, target_period: str, filepath: str) -> int:
    for j, val in enumerate(df.iloc[0]):
        if pd.notna(val) and extract_period(str(val)) == target_period:
            return j
    raise ValueError(f"Period column {target_period} not found in {filepath}")


def read_period_xls(filepath: str, target_period: str) -> list:
    df = pd.read_excel(filepath, header=None, dtype=str)
    period_col = _find_period_col_df(df, target_period, filepath)
    rows = []
    for _, row in df.iterrows():
        cell = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ""
        parsed = split_account(cell)
        if parsed is None:
            continue
        acc, name = parsed
        amt = parse_amount(row.iloc[period_col])
        if amt == 0.0:
            continue
        rows.append((acc, name, amt))
    return rows


def read_period_xls_diff(filepath: str, target_period: str) -> list:
    """166-style BS: values are period-end balances → compute col_target - col_prev."""
    df = pd.read_excel(filepath, header=None, dtype=str)
    period_col = _find_period_col_df(df, target_period, filepath)
    prev_col = period_col - 1   # previous month column
    rows = []
    for _, row in df.iterrows():
        cell = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ""
        parsed = split_account(cell)
        if parsed is None:
            continue
        acc, name = parsed
        curr = parse_amount(row.iloc[period_col])
        prev = parse_amount(row.iloc[prev_col]) if prev_col >= 0 else 0.0
        amt = curr - prev
        if amt == 0.0:
            continue
        rows.append((acc, name, amt))
    return rows


# ---------------------------------------------------------------------------
# Format E  –  "Turvatalo XLSX"  (153)
#   col 0 = "XXXX, Name",  col 1 = monthly amount (single data column)
# ---------------------------------------------------------------------------

def read_turvatalo_xlsx(filepath: str) -> list:
    df = pd.read_excel(filepath, header=None, dtype=str)
    rows = []
    for _, row in df.iterrows():
        cell = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ""
        parsed = split_account(cell)
        if parsed is None:
            continue
        acc, name = parsed
        amt = parse_amount(row.iloc[1])
        if amt == 0.0:
            continue
        rows.append((acc, name, amt))
    return rows


# ---------------------------------------------------------------------------
# Format F  –  "Ajan Lukko XLSX"  (179)
#   Wide format, account in col 3 ("XXXX   Name"),  period col found via "N (N)" header
# ---------------------------------------------------------------------------

def _find_col_by_pattern(df: pd.DataFrame, header_row: int, pattern: str) -> int:
    for j, val in enumerate(df.iloc[header_row]):
        if pd.notna(val) and re.search(pattern, str(val), re.IGNORECASE):
            return j
    raise ValueError(f"Column matching '{pattern}' not found in row {header_row}")


def read_ajan_lukko_xlsx(filepath: str, period: str) -> list:
    month = int(period[4:])
    pattern = rf"{month}\s*\({month}\)"
    df = pd.read_excel(filepath, header=None, dtype=str)
    period_col = _find_col_by_pattern(df, header_row=5, pattern=pattern)
    rows = []
    for _, row in df.iterrows():
        cell = str(row.iloc[3]) if pd.notna(row.iloc[3]) else ""
        parsed = split_account(cell)
        if parsed is None:
            continue
        acc, name = parsed
        amt = parse_amount(row.iloc[period_col])
        if amt == 0.0:
            continue
        rows.append((acc, name, amt))
    return rows


# ---------------------------------------------------------------------------
# Format G  –  "Kausi XLS/XLSM"  (134, 196)
#   col 0 = account number (numeric),  col 1 = name,  col 2 = Kausi (period)
#   For 134: skip bold rows (xlrd)
# ---------------------------------------------------------------------------

def _is_bold_xlrd(wb, ws, row_idx: int) -> bool:
    row = ws.row(row_idx)
    if len(row) < 2:
        return False
    xf = wb.xf_list[row[1].xf_index]
    return bool(wb.font_list[xf.font_index].bold)


def read_kausi_xls(filepath: str, skip_bold: bool = False) -> list:
    ext = os.path.splitext(filepath)[1].lower()
    if ext == ".xls" and skip_bold:
        return _read_kausi_xls_xlrd(filepath)
    if ext in (".xlsx", ".xlsm") and skip_bold:
        return _read_kausi_xlsx_openpyxl(filepath)
    df = pd.read_excel(filepath, header=None, dtype=str)
    return _collect_kausi_df(df)


def _read_kausi_xls_xlrd(filepath: str) -> list:
    wb = xlrd.open_workbook(filepath, formatting_info=True)
    ws = wb.sheets()[0]
    rows = []
    for i in range(ws.nrows):
        if _is_bold_xlrd(wb, ws, i):
            continue
        row = ws.row(i)
        val0 = row[0].value if row else None
        if not val0:
            continue
        try:
            acc = int(float(val0))
        except (ValueError, TypeError):
            continue
        name = str(row[1].value).strip() if len(row) > 1 else ""
        amt_raw = row[2].value if len(row) > 2 else None
        amt = parse_amount(amt_raw)
        if amt == 0.0:
            continue
        rows.append((acc, name, amt))
    return rows


def _read_kausi_xlsx_openpyxl(filepath: str) -> list:
    from openpyxl import load_workbook
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows():
        if len(row) < 3:
            continue
        cell1 = row[1]
        if cell1.font and cell1.font.bold:
            continue
        cell0 = row[0]
        if cell0.value is None:
            continue
        try:
            acc = int(float(str(cell0.value)))
        except (ValueError, TypeError):
            continue
        name = str(cell1.value).strip() if cell1.value else ""
        amt = parse_amount(row[2].value)
        if amt == 0.0:
            continue
        rows.append((acc, name, amt))
    return rows


def _collect_kausi_df(df: pd.DataFrame) -> list:
    rows = []
    for _, row in df.iterrows():
        val0 = row.iloc[0]
        if pd.isna(val0) or str(val0).strip() == "":
            continue
        try:
            acc = int(float(str(val0).strip()))
        except (ValueError, TypeError):
            continue
        name = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ""
        amt = parse_amount(row.iloc[2])
        if amt == 0.0:
            continue
        rows.append((acc, name, amt))
    return rows


# ---------------------------------------------------------------------------
# Format N  –  SpreadsheetML XML 2003 (134 Arvolukko jan-2026 övergångsformat)
#   .xls-fil men faktiskt XML. Kolumner: 1=konto, 2=namn, 3=Kausi (period),
#   4=Tilikausi (fiscal year balance), 5=Edellinen tilikausi (prev year), 6=%.
#   Använder col 3 (Kausi) — månadens bevegelse.
# ---------------------------------------------------------------------------

def read_spreadsheetml_xls(filepath: str, amount_col: int = 3) -> list:
    import xml.etree.ElementTree as ET
    ns = {"ss": "urn:schemas-microsoft-com:office:spreadsheet"}
    INDEX = "{urn:schemas-microsoft-com:office:spreadsheet}Index"
    STYLE = "{urn:schemas-microsoft-com:office:spreadsheet}StyleID"
    tree = ET.parse(filepath)
    root = tree.getroot()

    # Bygg en bold-karta: style_id → True om Font[ss:Bold='1'] (eller om
    # föräldra-stilen är bold, t.ex. B3 utan egen Bold men ärver från ss:Parent).
    bold_styles: set[str] = set()
    style_parents: dict[str, str] = {}
    for s in root.findall("ss:Styles/ss:Style", ns):
        sid = s.get("{urn:schemas-microsoft-com:office:spreadsheet}ID")
        if not sid:
            continue
        parent = s.get("{urn:schemas-microsoft-com:office:spreadsheet}Parent")
        if parent:
            style_parents[sid] = parent
        font = s.find("ss:Font", ns)
        if font is not None and font.get(
            "{urn:schemas-microsoft-com:office:spreadsheet}Bold") == "1":
            bold_styles.add(sid)

    def _is_bold(style_id: str | None) -> bool:
        cur = style_id
        seen: set[str] = set()
        while cur and cur not in seen:
            if cur in bold_styles:
                return True
            seen.add(cur)
            cur = style_parents.get(cur)
        return False

    rows_out: list = []
    for r in root.findall(".//ss:Row", ns):
        col = 0
        by_col: dict[int, str] = {}
        first_style: str | None = None
        for c in r.findall("ss:Cell", ns):
            idx = c.get(INDEX)
            col = int(idx) if idx else col + 1
            if first_style is None:
                first_style = c.get(STYLE)
            data = c.find("ss:Data", ns)
            if data is not None and data.text:
                by_col[col] = data.text
        # Hoppa över bold-rader (summa-rader som "LIIKEVAIHTO YHTEENSÄ" är
        # styled B3 → ärver bold från sin parent).
        if _is_bold(first_style):
            continue
        acc_raw = (by_col.get(1) or "").strip()
        if not acc_raw or not acc_raw.isdigit():
            continue
        try:
            acc = int(acc_raw)
        except ValueError:
            continue
        name = (by_col.get(2) or "").strip()
        amt = parse_amount(by_col.get(amount_col))
        if amt == 0.0:
            continue
        rows_out.append((acc, name, amt))
    return rows_out


# ---------------------------------------------------------------------------
# Format M  –  "Sparse-row XLSX" (182 jan-2026 övergångsformat)
#   col 0 = 'NNNN[N]  Name', de flesta cellerna tomma + en eller flera beloppsceller.
#   pick='first' tar första numeriska värdet (rätt för IS-filer som bara har en
#   periodkolumn). pick='last' tar sista numeriska (rätt för BS-filer där sista
#   kolumnen är muutos = månadsändring; tidigare kolumner är period-balanser).
# ---------------------------------------------------------------------------

def read_sparse_row_xlsx(filepath: str, *, pick: str = "first") -> list:
    df = pd.read_excel(filepath, header=None, dtype=object)
    rows = []
    for _, row in df.iterrows():
        cell = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ""
        parsed = split_account(cell)
        if parsed is None:
            continue
        acc, name = parsed
        cells = row.iloc[1:]
        iter_cells = reversed(list(cells)) if pick == "last" else list(cells)
        amt = 0.0
        for v in iter_cells:
            if isinstance(v, (int, float)) and not pd.isna(v):
                amt = float(v)
                break
            parsed_amt = parse_amount(v)
            if parsed_amt != 0.0:
                amt = parsed_amt
                break
        if amt == 0.0:
            continue
        rows.append((acc, name, amt))
    return rows


# ---------------------------------------------------------------------------
# Format H  –  "Avaava CSV UTF-8"  (199)
#   Semicolon separated, UTF-8 BOM, amounts may have spaces (thousands sep)
#   BS: col 2 = Jakson muutos,  IS: col 1 = period amount
# ---------------------------------------------------------------------------

def read_avaava_csv(filepath: str, amount_col: int) -> list:
    rows = []
    with open(filepath, encoding="utf-8-sig", newline="") as fh:
        reader = csv.reader(fh, delimiter=";")
        for line in reader:
            if not line:
                continue
            cell = line[0].strip()
            parsed = split_account(cell)
            if parsed is None:
                continue
            acc, name = parsed
            if len(line) <= amount_col:
                continue
            amt = parse_amount(line[amount_col])
            if amt == 0.0:
                continue
            rows.append((acc, name, amt))
    return rows


# ---------------------------------------------------------------------------
# Format I  –  Generic CSV with configurable column and encoding
#   Used for: UTF-16-LE muutos (193/195 BS), UTF-16-LE fennoa (195 IS),
#             UTF-16-LE monthly col 3 (193 IS)
# ---------------------------------------------------------------------------

def _read_csv_rows(filepath: str, encoding: str) -> list[list[str]]:
    import io as _io, csv as _csv
    with open(filepath, encoding=encoding) as fh:
        content = fh.read()
    return list(_csv.reader(_io.StringIO(content), delimiter=";", quotechar='"'))


def read_csv_col_enc(filepath: str, col: int, encoding: str = "latin-1",
                     target_period: str = None) -> list:
    """col 0 = 'NNNN Name', col <col> = amount. Handles any encoding.
    If target_period given, scans header rows to find the right column dynamically."""
    all_rows = _read_csv_rows(filepath, encoding)
    if target_period:
        found = _find_period_col(all_rows, target_period)
        if found is not None:
            col = found
    rows = []
    for row in all_rows:
        cell = row[0].strip() if row else ""
        parsed = split_account(cell)
        if parsed is None:
            continue
        acc, name = parsed
        amt = parse_amount(row[col] if len(row) > col else None)
        if amt == 0.0:
            continue
        rows.append((acc, name, amt))
    return rows


def read_muutos_csv_enc(filepath: str, encoding: str = "latin-1") -> list:
    """Muutos-style (col 2 = change) with configurable encoding."""
    rows = []
    for row in _read_csv_rows(filepath, encoding):
        cell = row[0].strip() if row else ""
        parsed = split_account(cell)
        if parsed is None:
            continue
        acc, name = parsed
        amt = parse_amount(row[2] if len(row) > 2 else None)
        if amt == 0.0:
            continue
        rows.append((acc, name, amt))
    return rows


# ---------------------------------------------------------------------------
# Format J  –  Emsec XLSX  (215)
#   BS: col 1 = 'NNNN Name', col 3 = change (numeric or '\xa0'-formatted string)
#   PL: col 1 = 'NNNN Name', col 2 = monthly amount
# ---------------------------------------------------------------------------

def read_emsec_bs_xlsx(filepath: str) -> list:
    df = pd.read_excel(filepath, header=None, dtype=object)
    rows = []
    for _, row in df.iterrows():
        cell = str(row.iloc[1]) if len(row) > 1 and row.iloc[1] is not None else ""
        parsed = split_account(cell)
        if parsed is None:
            continue
        acc, name = parsed
        amt = parse_amount(row.iloc[3] if len(row) > 3 else None)
        if amt == 0.0:
            continue
        rows.append((acc, name, amt))
    return rows


def read_emsec_pl_xlsx(filepath: str) -> list:
    df = pd.read_excel(filepath, header=None, dtype=object)
    rows = []
    for _, row in df.iterrows():
        cell = str(row.iloc[1]) if len(row) > 1 and row.iloc[1] is not None else ""
        parsed = split_account(cell)
        if parsed is None:
            continue
        acc, name = parsed
        amt = parse_amount(row.iloc[2] if len(row) > 2 else None)
        if amt == 0.0:
            continue
        rows.append((acc, name, amt))
    return rows


# ---------------------------------------------------------------------------
# Format K  –  JM Lukko combined BS+IS XLSX  (221)
#   Single sheet; col 0 = 'NNNN Name'.
#   Layouts varierar per månad:
#     - Jan (period 1): bara YTD-kolumn (col 1 = monthly = YTD).
#     - Feb+ : col 1 = YTD, col 2 = månadsändring → använd col 2.
#   Filter by 4-digit account prefix: 1000-2999 → BS, 3000-9999 → IS.
# ---------------------------------------------------------------------------

def read_combined_xlsx_col2(filepath: str, acc_min_4d: int = 0, acc_max_4d: int = 9999) -> list:
    df = pd.read_excel(filepath, header=None, dtype=str)
    # Auto-detect column: föredra col 2 (månadsändring) om den finns,
    # fallback till col 1 (jan-fil med bara YTD). _find_period_col() funkar inte
    # här eftersom JM:s header har datumintervall ("1.1.2026 - 31.1.2026"), inte
    # period-format som extract_period() känner igen direkt.
    amt_col = 2 if df.shape[1] > 2 else 1
    rows = []
    for _, row in df.iterrows():
        cell = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ""
        parsed = split_account(cell)
        if parsed is None:
            continue
        acc, name = parsed
        if not (acc_min_4d <= _normalize4(acc) <= acc_max_4d):
            continue
        amt = parse_amount(row.iloc[amt_col] if len(row) > amt_col else None)
        if amt == 0.0:
            continue
        rows.append((acc, name, amt))
    return rows


# ---------------------------------------------------------------------------
# Format L  –  ANV period-column CSV  (238)
#   Header row: ;202601;202602;202603;
#   BS: col3 - col2 (period-end balances → monthly change)
#   IS: col3 directly (monthly amounts)
#   Encoding cp1252; amounts may use space as thousands sep.
# ---------------------------------------------------------------------------

def _parse_period_csv_lines(filepath: str):
    with open(filepath, encoding="cp1252", newline="") as fh:
        for line in fh:
            cols = line.rstrip("\n").split(";")
            cell = cols[0].strip() if cols else ""
            parsed = split_account(cell)
            if parsed is None:
                continue
            yield parsed, cols


def _period_csv_target_col(filepath: str, target_period: str) -> int:
    """Find the column index matching target_period in the header of a period CSV."""
    with open(filepath, encoding="cp1252") as fh:
        header = fh.readline()
    cols = header.rstrip("\n").split(";")
    for i, c in enumerate(cols):
        if extract_period(c.strip()) == target_period:
            return i
    # Fallback: last YYYYMM-looking column
    last = None
    for i, c in enumerate(cols):
        if re.match(r"^\s*20\d{4}\s*$", c):
            last = i
    return last if last is not None else 3


def read_period_csv_diff(filepath: str, target_period: str) -> list:
    """238 BS: period-column CSV. Computes change = col[target] - col[target-1]."""
    target_col = _period_csv_target_col(filepath, target_period)
    rows = []
    for (acc, name), cols in _parse_period_csv_lines(filepath):
        curr = parse_amount(cols[target_col]) if len(cols) > target_col else 0.0
        prev = parse_amount(cols[target_col - 1]) if len(cols) > target_col - 1 else 0.0
        amt = round(curr - prev, 2)
        if amt == 0.0:
            continue
        rows.append((acc, name, amt))
    return rows


def read_period_csv_col(filepath: str, target_period: str) -> list:
    """238 IS: period-column CSV. Uses the column matching target_period."""
    target_col = _period_csv_target_col(filepath, target_period)
    rows = []
    for (acc, name), cols in _parse_period_csv_lines(filepath):
        amt = parse_amount(cols[target_col]) if len(cols) > target_col else 0.0
        if amt == 0.0:
            continue
        rows.append((acc, name, amt))
    return rows


# ---------------------------------------------------------------------------
# Common post-processing (flip signs, exclude 237X)
# ---------------------------------------------------------------------------

def post_process(raw_rows: list, flip: bool = True) -> list:
    result = []
    for acc, name, amt in raw_rows:
        if is_237x(acc):
            continue
        if flip:
            amt = apply_flip(amt, acc)
        result.append((acc, name, round(amt, 2)))
    return result


# ---------------------------------------------------------------------------
# Company processor
# ---------------------------------------------------------------------------

def process_company(
    code: str,
    friendly_name: str,
    period: str,
    bs_rows_raw: list,
    is_rows_raw: list,
) -> str:
    bs_rows = post_process(bs_rows_raw, flip=True)
    is_rows = post_process(is_rows_raw, flip=False)

    total = verify_sum(is_rows, bs_rows)
    is_warn = abs(total) > 0.11
    check = "OK (~0)" if not is_warn else f"WARN: {total:.2f}"
    log("INFO", code, f"IS={len(is_rows)}, BS={len(bs_rows)}  Sum={total:.4f}  {check}")

    filename = f"{code}_{friendly_name}_{period}_INL.xlsx"
    if _DRY_RUN:
        log("OK", code, f"[DRY] Skulle spara: {filename}")
    else:
        save_inl_xlsx(is_rows, bs_rows, OUTPUT_DIR / filename)
        log("WARN" if is_warn else "OK", code, f"Sparad: {filename}")

    return "warn" if is_warn else "ok"


# ---------------------------------------------------------------------------
# Individual company definitions  (each takes d: Path = FINLAND_DIR)
# ---------------------------------------------------------------------------

def run_146(d: Path):
    return process_company(
        code="146", friendly_name="Avain-Asema", period=PERIOD,
        bs_rows_raw=read_fennoa_csv(str(glob_one(d, "146_*Balansr*.csv")), target_period=PERIOD),
        is_rows_raw=read_fennoa_csv(str(glob_one(d, "146_*Resultatr*.csv")), target_period=PERIOD),
    )


def run_134(d: Path):
    # 134 har två filformat under 2026:
    #   - Standard .xls (xlrd): finska namn ("tase", "tuloslaskelma")
    #   - SpreadsheetML XML (.xls med XML-innehåll): engelska namn (jan-2026)
    # Detekteras via filhuvud (b'<?xml ...').
    def _is_xml(p: Path) -> bool:
        with open(p, "rb") as fh:
            return fh.read(5).startswith(b"<?xml")

    bs_f = _try_glob(d, "134_*tase*.xls",
                     "134_*[Bb]alance*[Ss]heet*.xls")
    is_f = _try_glob(d, "134_*tuloslaskelma*.xls",
                     "134_*[Ii]ncome*statement*.xls")
    bs_raw = (read_spreadsheetml_xls(str(bs_f)) if _is_xml(bs_f)
              else read_kausi_xls(str(bs_f), skip_bold=True))
    is_raw = (read_spreadsheetml_xls(str(is_f)) if _is_xml(is_f)
              else read_kausi_xls(str(is_f), skip_bold=True))
    return process_company(
        code="134", friendly_name="Arvolukko", period=PERIOD,
        bs_rows_raw=bs_raw, is_rows_raw=is_raw,
    )


def run_153(d: Path):
    # 153 har växlat mellan "Balans" (svenskt namn, t.o.m. 202603) och "Balance"
    # (engelskt namn, fr.o.m. 202604) i BS-filen.
    return process_company(
        code="153", friendly_name="Turvatalo", period=PERIOD,
        bs_rows_raw=read_turvatalo_xlsx(str(_try_glob(d,
            "153_*[Bb]alans*.xlsx", "153_*[Bb]alance*.xlsx"))),
        is_rows_raw=read_turvatalo_xlsx(str(glob_one(d, "153_*[Ii]ncome*.xlsx"))),
    )


def run_161(d: Path):
    bs_f = _try_glob(d, "161_*[Tt]ase*.xlsx", "161_*[Tt]ase*.csv")
    is_f = _try_glob(d, "161_*[Tt]uloslaskelma*.xlsx", "161_*[Tt]uloslaskelma*.csv")
    bs_raw = (read_muutos_xlsx if bs_f.suffix.lower() == ".xlsx" else read_muutos_csv)(str(bs_f))
    is_raw = (read_income_only_xlsx if is_f.suffix.lower() == ".xlsx" else read_fennoa_csv)(str(is_f))
    return process_company(
        code="161", friendly_name="Lukitustekniikka-STY", period=PERIOD,
        bs_rows_raw=bs_raw, is_rows_raw=is_raw,
    )


def run_166(d: Path):
    return process_company(
        code="166", friendly_name="Lukkoassa", period=PERIOD,
        bs_rows_raw=read_period_xls_diff(str(glob_one(d, "166_*[Tt]ase*kk*.xls")), PERIOD),
        is_rows_raw=read_period_xls(str(glob_one(d, "166_*[Tt]uloslaskelma*kk*.xls")), PERIOD),
    )


def run_173(d: Path):
    return process_company(
        code="173", friendly_name="Avainahjo", period=PERIOD,
        bs_rows_raw=read_muutos_xlsx(str(glob_one(d, "173_*[Bb]alance*.xlsx"))),
        is_rows_raw=read_income_only_xlsx(str(glob_one(d, "173_*[Ii]ncome*.xlsx"))),
    )


def run_179(d: Path):
    return process_company(
        code="179", friendly_name="Ajan-Lukko", period=PERIOD,
        bs_rows_raw=read_ajan_lukko_xlsx(str(glob_one(d, "179_*balance*sheet*.xlsx")), PERIOD),
        is_rows_raw=read_ajan_lukko_xlsx(str(glob_one(d, "179_*income*statement*.xlsx")), PERIOD),
    )


def run_181(d: Path):
    return process_company(
        code="181", friendly_name="Tele-Projekti", period=PERIOD,
        bs_rows_raw=read_muutos_csv(str(glob_one(d, "181_*[Tt]ase*.csv"))),
        is_rows_raw=read_fennoa_csv(str(glob_one(d, "181_*[Tt]uloslaskelma*.csv"))),
    )


def run_196(d: Path):
    # ST Hälytys-filerna har växlat mellan finskt namn (tase) och engelskt
    # (balance sheet). IS-glob:en matchade redan "profit & loss statement"
    # eftersom * mellan profit och loss matchar valfri text inkl. "& ".
    return process_company(
        code="196", friendly_name="ST-Halytys", period=PERIOD,
        bs_rows_raw=read_kausi_xls(str(_try_glob(d,
            "196_*tase*.xlsm", "196_*[Bb]alance*[Ss]heet*.xlsm")), skip_bold=True),
        is_rows_raw=read_kausi_xls(str(glob_one(d, "196_*profit*loss*.xlsm")), skip_bold=True),
    )


def run_199(d: Path):
    return process_company(
        code="199", friendly_name="Etela-Halytintekniikka", period=PERIOD,
        bs_rows_raw=read_avaava_csv(str(glob_one(d, "199_*[Tt]ase*.csv")), amount_col=2),
        is_rows_raw=read_avaava_csv(str(glob_one(d, "199_*[Tt]ulos*.csv")), amount_col=1),
    )


def run_170(d: Path):
    return process_company(
        code="170", friendly_name="PAP", period=PERIOD,
        bs_rows_raw=read_fennoa_csv(str(glob_one(d, "170_*tase*.csv"))),
        is_rows_raw=read_fennoa_csv(str(glob_one(d, "170_*tulos*.csv"))),
    )


def run_177(d: Path):
    f = str(glob_one(d, "177_*.xlsx"))
    bs_sheet = _find_sheet(f, r"LL Balance Sheet")
    is_sheet = _find_sheet(f, r"LL Income Statement")
    return process_company(
        code="177", friendly_name="Lukkoluket", period=PERIOD,
        bs_rows_raw=read_muutos_xlsx(f, sheet_name=bs_sheet),
        is_rows_raw=read_income_only_xlsx(f, sheet_name=is_sheet),
    )


def run_182(d: Path):
    # 202602+ använder Fennoa-CSV; jan-2026 övergångsformat var sparse XLSX
    # (en numerisk cell per rad; BS = muutos sist, IS = period först).
    bs_f = _try_glob(d, "182_*[Tt]ase*.csv", "182_*[Tt]ase*.xlsx")
    is_f = _try_glob(d, "182_*[Tt]uloslaskelma*.csv", "182_*[Tt]uloslaskelma*.xlsx")
    bs_raw = (read_muutos_csv(str(bs_f)) if bs_f.suffix.lower() == ".csv"
              else read_sparse_row_xlsx(str(bs_f), pick="last"))
    is_raw = (read_fennoa_csv(str(is_f)) if is_f.suffix.lower() == ".csv"
              else read_sparse_row_xlsx(str(is_f), pick="first"))
    return process_company(
        code="182", friendly_name="THV", period=PERIOD,
        bs_rows_raw=bs_raw, is_rows_raw=is_raw,
    )


def run_185(d: Path):
    # Tar månads-filen (01.MM.YYYY-DD.MM.YYYY) i stället för YTD (01.01-DD.MM).
    month = _fennoa_month_range(PERIOD)
    return process_company(
        code="185", friendly_name="Suomen-Turvalukko", period=PERIOD,
        bs_rows_raw=read_muutos_csv(str(glob_one(d, f"185_*[Bb]alance*sheet*{month}*.csv"))),
        is_rows_raw=read_fennoa_csv(str(glob_one(d, f"185_*[Ii]ncome*statement*{month}*.csv")), target_period=PERIOD),
    )


def run_193(d: Path):
    return process_company(
        code="193", friendly_name="Suomen-Turvakonsultit", period=PERIOD,
        bs_rows_raw=read_muutos_csv_enc(
            str(glob_one(d, "193_*[Bb]alance*sheet*.csv")), encoding="utf-16-le",
        ),
        is_rows_raw=read_csv_col_enc(
            str(glob_one(d, "193_*[Pp]rofit*loss*.csv")),
            col=3, encoding="utf-16-le", target_period=PERIOD,
        ),
    )


def run_195(d: Path):
    # Tar månads-filen (01.MM.YYYY-DD.MM.YYYY) i stället för YTD (01.01-DD.MM)
    # för både Tase och Tuloslaskelma. Muutos i månadsfilen = månadsändring.
    month = _fennoa_month_range(PERIOD)
    return process_company(
        code="195", friendly_name="Meri-Lapin", period=PERIOD,
        bs_rows_raw=read_muutos_csv_enc(
            str(glob_one(d, f"195_*[Tt]ase*{month}*.csv")), encoding="utf-16-le",
        ),
        is_rows_raw=read_csv_col_enc(
            str(glob_one(d, f"195_Tuloslaskelma_*{month}*.csv")),
            col=1, encoding="utf-16-le", target_period=PERIOD,
        ),
    )


def run_215(d: Path):
    return process_company(
        code="215", friendly_name="Emsec", period=PERIOD,
        bs_rows_raw=read_emsec_bs_xlsx(str(glob_one(d, "215_*BS*.xlsx"))),
        is_rows_raw=read_emsec_pl_xlsx(str(glob_one(d, "215_*PL*.xlsx"))),
    )


def run_221(d: Path):
    f = str(glob_one(d, "221_*.xlsx"))
    return process_company(
        code="221", friendly_name="JM-Lukko", period=PERIOD,
        bs_rows_raw=read_combined_xlsx_col2(f, acc_min_4d=1000, acc_max_4d=2999),
        is_rows_raw=read_combined_xlsx_col2(f, acc_min_4d=3000, acc_max_4d=9999),
    )


def run_238(d: Path):
    return process_company(
        code="238", friendly_name="ANV", period=PERIOD,
        bs_rows_raw=read_period_csv_diff(str(glob_one(d, "238_tase-kk_*.csv")), PERIOD),
        is_rows_raw=read_period_csv_col(str(glob_one(d, "238_tuloslaskelma-kk_*.csv")), PERIOD),
    )


def run_145(d: Path):
    # 145 (Prosero Security Oy) växlar filformat per månad:
    #   - 202601: utf-16-le CSV (Tase / Tuloslaskelma)
    #   - 202602/03: binär XLS, finskt namn
    #   - 202604: XLSX med engelska namn (Balance_sheet / Income_statement)
    # Mailen bifogar dessutom helår-2025-referensfiler — filtrera per
    # period-månadsintervall så vi inte plockar fel.
    month = _fennoa_month_range(PERIOD)
    bs_f = _try_glob(
        d,
        f"145_*[Tt]ase*{month}*.xls*",
        f"145_*[Tt]ase*{month}*.csv",
        f"145_*[Bb]alance*[Ss]heet*.xls*",
        f"145_*[Bb]alance*[Ss]heet*.csv",
    )
    is_f = _try_glob(
        d,
        f"145_*[Tt]uloslaskelma*{month}*.xls*",
        f"145_*[Tt]uloslaskelma*{month}*.csv",
        f"145_*[Ii]ncome*[Ss]tatement*{month}*.xls*",
        f"145_*[Ii]ncome*[Ss]tatement*.xls*",
    )
    bs_raw = (read_muutos_csv_enc(str(bs_f), encoding="utf-16-le")
              if bs_f.suffix.lower() == ".csv" else read_muutos_xlsx(str(bs_f)))
    # 145:s IS-csv har bara ETT period-värde i header (col 2 har "Päivämääräväli"
    # och visar månadsintervallet — det är bara en metadata-cell, inte en
    # kolumnetikett). _find_period_col-skannern skulle annars välja fel kolumn.
    is_raw = (read_csv_col_enc(str(is_f), col=1, encoding="utf-16-le")
              if is_f.suffix.lower() == ".csv"
              else read_income_only_xlsx(str(is_f)))
    return process_company(
        code="145", friendly_name="Prosero Security Oy", period=PERIOD,
        bs_rows_raw=bs_raw, is_rows_raw=is_raw,
    )


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

RUNNERS = {
    "134": run_134,
    "145": run_145,
    "146": run_146,
    "153": run_153,
    "161": run_161,
    "166": run_166,
    "170": run_170,
    "173": run_173,
    "177": run_177,
    "179": run_179,
    "181": run_181,
    "182": run_182,
    "185": run_185,
    "193": run_193,
    "195": run_195,
    "196": run_196,
    "199": run_199,
    "215": run_215,
    "221": run_221,
    "238": run_238,
}

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Processera finska bolag → INL.xlsx")
    parser.add_argument("--dry-run", "-n", action="store_true", help="Visa utan att skriva")
    parser.add_argument(
        "--period", "-p", metavar="YYYYMM", default=None,
        help="Period att köra (t.ex. 202604). Standard: föregående månad.",
    )
    parser.add_argument("codes", nargs="*", help="Bolagskoder (lämna tomt för alla)")
    args = parser.parse_args()

    _DRY_RUN = args.dry_run

    if args.period:
        PERIOD = args.period
        FINLAND_DIR  = Path(load_config()["base_path"]) / "extracted" / args.period / "Finland"
        OUTPUT_DIR   = FINLAND_DIR / "output"
        REFERENS_DIR = FINLAND_DIR / "Referens"
    else:
        from datetime import date as _date
        today = _date.today()
        PERIOD = (
            f"{today.year - 1}12" if today.month == 1
            else f"{today.year}{today.month - 1:02d}"
        )

    begin_run("process_finland", PERIOD)
    dry_label = "  [DRY RUN]" if _DRY_RUN else ""
    log("START", "process_finland.py", f"period {PERIOD}{dry_label}")

    if not FINLAND_DIR.exists():
        import sys as _sys
        _sys.exit(f"[ERROR]  Finland-mappen saknas: {FINLAND_DIR}")

    if not _DRY_RUN:
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        REFERENS_DIR.mkdir(parents=True, exist_ok=True)

    move_pdfs_to_referens()

    codes = args.codes if args.codes else sorted(RUNNERS)
    stats: dict[str, int] = {"ok": 0, "warn": 0, "skip": 0, "error": 0}

    for code in codes:
        if code not in RUNNERS:
            log("ERROR", code, "Okänd bolagskod")
            stats["error"] += 1
            continue
        try:
            status = RUNNERS[code](FINLAND_DIR)
            stats[status] = stats.get(status, 0) + 1
        except FileNotFoundError as e:
            fname = Path(str(e)).name if e.args else "okänd fil"
            log("SKIP", code, f"Källfil saknas: {fname}")
            stats["skip"] += 1
        except Exception as e:
            log("ERROR", code, f"Fel: {e}")
            stats["error"] += 1
        finally:
            # Move all {code}_* source files to Referens regardless of outcome
            for f in sorted(FINLAND_DIR.glob(f"{code}_*")):
                if f.is_file():
                    move_to_referens(f.name)

    log("DONE", "process_finland.py",
        f"{stats['ok']} OK  {stats['warn']} WARN  {stats['skip']} SKIP  {stats['error']} ERROR")
